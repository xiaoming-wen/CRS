import logging
import mimetypes
import zipfile
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Request, Query
from fastapi.responses import FileResponse, StreamingResponse, Response
from openpyxl import Workbook, load_workbook
from starlette.datastructures import UploadFile as StarletteUploadFile
from sqlalchemy import func, and_, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional, Tuple, Union
from app.datetime_utils import utc_now, ensure_utc
import os
import uuid
import re

from app.eight_digit_id import allocate_eight_digit_id, validate_eight_digit_id
from app.database import get_db
from app.alt_auth.context import get_current_alt_identity, get_optional_alt_identity
from app.alt_auth.database import get_alt_auth_db
from app.alt_auth.models import AltAuthUserRecord
from app.permissions import Permission, require_permission
from app.models.user import File as FileModel
from app.models.competition import (
    Competition,
    CompetitionEnrollment,
    CompetitionEnrollmentScope,
    CompetitionEnrollmentStatus,
    CompetitionExpertAssignment,
    CompetitionExpertTeamAssignment,
    CompetitionPromotion,
    CompetitionQuestionAnswer,
    CompetitionQuestionAnswerStatus,
    CompetitionStage,
    COMPETITION_QUESTION_COUNT,
    Team,
    TeamMember,
    TeamJoinRequest,
    TeamJoinRequestStatus,
    TeamInvite,
    TeamInviteStatus,
    TeamStatus,
    Submission,
    SubmissionStatus,
    Review,
    CompetitionTeamQuestionGrade,
)
from app.schemas import (
    CompetitionCreate,
    CompetitionUpdate,
    CompetitionResponse,
    CompetitionExamPapers,
    CompetitionExamPaperSlot,
    CompetitionSubmissionQuestionConfig,
    CompetitionSubmissionQuestionConfigByTrack,
    CompetitionEnrollmentCreate,
    CompetitionEnrollmentResponse,
    CompetitionPromotionCreate,
    CompetitionPromotionResponse,
    CompetitionPromotionImportResult,
    CompetitionPromotionImportItemResult,
    CompetitionPromotionCandidateTeam,
    CompetitionPromotionCandidatesResponse,
    MyEnrollmentResponse,
    MyRejectedTeamItem,
    TeamCreate,
    TeamResponse,
    TeamPatch,
    TeamInviteMember,
    TeamDetailResponse,
    TeamMemberResponse,
    TeamJoinRequestResponse,
    TeamJoinRequestReview,
    TeamInviteResponse,
    TeamInviteAction,
    IndividualParticipantItem,
    TeamParticipantDetailResponse,
    TeamMemberWithUserResponse,
    TeamTransferCaptain,
    AltUserAdminPatch,
    AltUserAdminUpdateResult,
    SchoolAdminTeamReviewListResponse,
    SchoolAdminTeamReviewItem,
    SchoolAdminTeamMemberItem,
    TeamSchoolReviewRequest,
    TeamSchoolReviewResult,
    SchoolAdminSetTeamAdvisorRequest,
    SchoolAdminSetTeamAdvisorResult,
    TeamSetSecondAdvisorRequest,
    TeamSetSecondAdvisorResult,
    SchoolAdminSetTeamDivisionTrackRequest,
    SchoolAdminSetTeamDivisionTrackResult,
    SchoolAdminProxyTeamCreate,
    SchoolAdminProxyEnrollRequest,
    SchoolAdminProxyEnrollResult,
    SchoolAdminApplicationStatus,
    SchoolAdminApplicationMeResponse,
    SchoolAdminApplicationListResponse,
    SchoolAdminApplicationListItem,
    SchoolAdminApplicationReviewRequest,
    SchoolAdminApplicationReviewResult,
    CompetitionExpertsListResponse,
    CompetitionExpertListItem,
    CompetitionExpertAssignedTeam,
    CompetitionExpertAssignRequest,
    SubmissionCreate,
    SubmissionCreateWrapped,
    SubmissionResponse,
    CompetitionQuestionAnswerResponse,
    CompetitionQuestionAnswerSlot,
    CompetitionQuestionAnswersBoard,
    CompetitionQuestionAnswersTeamOverview,
    CompetitionQuestionAnswersOverviewResponse,
    CompetitionQuestionAnswersSubmitResult,
    ReviewGrade,
    ReviewResponse,
    TeamQuestionGradeRequest,
    TeamQuestionGradeResponse,
    CompetitionScoreSummaryResponse,
    CompetitionScoreTeamItem,
    CompetitionScoreRankingItem,
    CompetitionScoreRankingResponse,
    CompetitionDivision,
    CompetitionWorkTrack,
    MyCompetitionScoresResponse,
    SubmissionForStudentScoreResponse,
)

router = APIRouter(prefix="/competitions", tags=["Competition Management"])
logger = logging.getLogger(__name__)

SUBMISSION_UPLOAD_DIR = "competition_submissions"
os.makedirs(SUBMISSION_UPLOAD_DIR, exist_ok=True)

QUESTION_ANSWER_UPLOAD_DIR = "competition_question_answers"
os.makedirs(QUESTION_ANSWER_UPLOAD_DIR, exist_ok=True)
MAX_QUESTION_ANSWER_BYTES = 100 * 1024 * 1024

COMPETITION_QR_DIR = "competition_qr_codes"
os.makedirs(COMPETITION_QR_DIR, exist_ok=True)
MAX_QR_IMAGE_BYTES = 5 * 1024 * 1024
_ALLOWED_QR_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
_ALLOWED_QR_MIME = {"image/png", "image/jpeg", "image/gif", "image/webp"}

COMPETITION_LOGO_DIR = "competition_logos"
os.makedirs(COMPETITION_LOGO_DIR, exist_ok=True)
MAX_LOGO_IMAGE_BYTES = 5 * 1024 * 1024
_ALLOWED_LOGO_EXT = _ALLOWED_QR_EXT
_ALLOWED_LOGO_MIME = _ALLOWED_QR_MIME

COMPETITION_EXAM_PAPER_DIR = "competition_exam_papers"
os.makedirs(COMPETITION_EXAM_PAPER_DIR, exist_ok=True)
MAX_EXAM_PAPER_BYTES = 100 * 1024 * 1024
_ALLOWED_EXAM_PAPER_EXT = {".pdf", ".doc", ".docx", ".zip"}

SCHOOL_ADMIN_PHOTO_DIR = "school_admin_applications"
os.makedirs(SCHOOL_ADMIN_PHOTO_DIR, exist_ok=True)
MAX_SCHOOL_ADMIN_PHOTO_BYTES = 5 * 1024 * 1024


def _alt_users_by_id(adb: Session, ids: set[int]) -> dict[int, AltAuthUserRecord]:
    if not ids:
        return {}
    rows = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id.in_(list(ids))).all()
    return {r.id: r for r in rows}


def _effective_alt_role(role: Optional[str]) -> str:
    return (role or "student").strip()


def _require_super_admin_identity(identity: AltAuthUserRecord) -> None:
    if _effective_alt_role(identity.role) != "super_admin":
        raise HTTPException(status_code=403, detail="Only super_admin can perform this operation")


def _expert_gate_ok(identity: AltAuthUserRecord) -> bool:
    """专家且已由管理员核验。"""
    return _effective_alt_role(identity.role) == "expert" and bool(getattr(identity, "expert_verified", False))


def _is_assigned_competition_expert(db: Session, competition_id: int, expert_user_id: int) -> bool:
    return (
        db.query(CompetitionExpertAssignment)
        .filter(
            CompetitionExpertAssignment.competition_id == competition_id,
            CompetitionExpertAssignment.expert_id == expert_user_id,
        )
        .first()
        is not None
    )


def _assigned_team_ids_for_expert(
    db: Session, competition_id: int, expert_user_id: int
) -> set[int]:
    rows = (
        db.query(CompetitionExpertTeamAssignment.team_id)
        .filter(
            CompetitionExpertTeamAssignment.competition_id == competition_id,
            CompetitionExpertTeamAssignment.expert_id == expert_user_id,
        )
        .all()
    )
    return {int(r[0]) for r in rows}


def _is_assigned_expert_for_team(
    db: Session, competition_id: int, team_id: int, expert_user_id: int
) -> bool:
    if not _is_assigned_competition_expert(db, competition_id, expert_user_id):
        return False
    return (
        db.query(CompetitionExpertTeamAssignment)
        .filter(
            CompetitionExpertTeamAssignment.competition_id == competition_id,
            CompetitionExpertTeamAssignment.expert_id == expert_user_id,
            CompetitionExpertTeamAssignment.team_id == team_id,
        )
        .first()
        is not None
    )


def _can_view_all_competition_submissions(db: Session, competition_id: int, identity: AltAuthUserRecord) -> bool:
    """超级管理员可看竞赛全部作品/队伍。专家须按队伍指派，不走此「全部」门。"""
    if _effective_alt_role(identity.role) == "super_admin":
        return True
    return False


def _is_competition_assigned_expert(db: Session, competition_id: int, identity: AltAuthUserRecord) -> bool:
    return _expert_gate_ok(identity) and _is_assigned_competition_expert(
        db, competition_id, identity.id
    )


def _can_view_competition_score_insights(db: Session, competition_id: int, identity: AltAuthUserRecord) -> bool:
    """管理员或已指派专家对竞赛查看评分汇总 / 排行榜。"""
    if _can_view_all_competition_submissions(db, competition_id, identity):
        return True
    return _is_competition_assigned_expert(db, competition_id, identity)


def _can_view_full_participant_rosters(db: Session, competition_id: int, identity: AltAuthUserRecord) -> bool:
    """参赛者花名册：超管看全部；专家仅看已指派队伍（由调用方再过滤）。"""
    if _can_view_all_competition_submissions(db, competition_id, identity):
        return True
    return _is_competition_assigned_expert(db, competition_id, identity)


def _ensure_alt_principal_is_student(adb: Session, user_id: int) -> AltAuthUserRecord:
    validate_eight_digit_id(user_id, label="用户ID")
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == user_id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="User not found")
    if _effective_alt_role(row.role) != "student":
        raise HTTPException(status_code=400, detail="Target must be a student account")
    return row


def _ensure_alt_principal_is_advisor(adb: Session, user_id: int) -> AltAuthUserRecord:
    validate_eight_digit_id(user_id, label="用户ID")
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == user_id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="指导老师用户不存在")
    if _effective_alt_role(row.role) not in {"advisor", "teacher"}:
        raise HTTPException(status_code=400, detail="指定用户须为指导老师账号")
    return row


def _looks_like_eight_digit_id(raw: Optional[str]) -> Optional[int]:
    """纯 8 位数字则解析为用户/竞赛 ID，否则返回 None。"""
    s = (raw or "").strip()
    if not s.isdigit() or len(s) != 8:
        return None
    n = int(s)
    try:
        validate_eight_digit_id(n, label="用户ID")
    except HTTPException:
        return None
    return n


def _match_advisors_by_name(adb: Session, name: str) -> list[AltAuthUserRecord]:
    key = (name or "").strip()
    if not key:
        return []
    key_lower = key.lower()
    matches: list[AltAuthUserRecord] = []
    for row in adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.is_active.is_(True)).all():
        if _effective_alt_role(row.role) not in {"advisor", "teacher"}:
            continue
        full_name = (row.full_name or "").strip()
        username = (row.username or "").strip()
        if full_name.lower() == key_lower or username.lower() == key_lower:
            matches.append(row)
    return matches


def _match_students_by_name(adb: Session, name: str) -> list[AltAuthUserRecord]:
    key = (name or "").strip()
    if not key:
        return []
    key_lower = key.lower()
    matches: list[AltAuthUserRecord] = []
    for row in adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.is_active.is_(True)).all():
        if _effective_alt_role(row.role) != "student":
            continue
        full_name = (row.full_name or "").strip()
        username = (row.username or "").strip()
        if full_name.lower() == key_lower or username.lower() == key_lower:
            matches.append(row)
    return matches


def _resolve_advisor_by_name(adb: Session, name: str) -> AltAuthUserRecord:
    key = (name or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="指导老师姓名不能为空")
    matches = _match_advisors_by_name(adb, key)
    if not matches:
        raise HTTPException(status_code=404, detail="未找到该指导老师，请确认姓名是否正确")
    if len(matches) > 1:
        raise HTTPException(status_code=400, detail="存在多位同名指导老师，请填写更准确的姓名")
    return matches[0]


def _try_resolve_advisor_by_name(adb: Session, name: str) -> Optional[AltAuthUserRecord]:
    key = (name or "").strip()
    if not key:
        return None
    matches = _match_advisors_by_name(adb, key)
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise HTTPException(status_code=400, detail="存在多位同名指导老师，请填写更准确的姓名")
    return None


def _resolve_advisor_ref(adb: Session, ref: str, *, label: str = "指导老师") -> AltAuthUserRecord:
    """按 8 位用户 ID、姓名或用户名解析指导老师。"""
    key = (ref or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail=f"{label}不能为空")
    nid = _looks_like_eight_digit_id(key)
    if nid is not None:
        return _ensure_alt_principal_is_advisor(adb, nid)
    matches = _match_advisors_by_name(adb, key)
    if not matches:
        raise HTTPException(status_code=404, detail=f"未找到该{label}，请确认姓名、用户名或用户 ID")
    if len(matches) > 1:
        raise HTTPException(status_code=400, detail=f"存在多位同名{label}，请改用 8 位用户 ID")
    return matches[0]


def _resolve_student_ref(adb: Session, ref: str, *, label: str = "学生") -> AltAuthUserRecord:
    """按 8 位用户 ID、姓名或用户名解析学生。"""
    key = (ref or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail=f"{label}不能为空")
    nid = _looks_like_eight_digit_id(key)
    if nid is not None:
        return _ensure_alt_principal_is_student(adb, nid)
    matches = _match_students_by_name(adb, key)
    if not matches:
        raise HTTPException(status_code=404, detail=f"未找到该{label}，请确认姓名、用户名或用户 ID")
    if len(matches) > 1:
        raise HTTPException(status_code=400, detail=f"存在多位同名{label}，请改用 8 位用户 ID")
    return matches[0]


def _team_advisor_display_name(
    team: Team,
    users_by_id: Optional[dict[int, AltAuthUserRecord]] = None,
) -> Optional[str]:
    stored = getattr(team, "advisor_name", None)
    if stored is not None and str(stored).strip():
        return str(stored).strip()
    advisor_id = team.created_by_advisor_id
    if advisor_id and users_by_id:
        adv = users_by_id.get(advisor_id)
        if adv:
            return _display_user_name(adv, advisor_id)
    return None


def _team_second_advisor_display_name(
    team: Team,
    users_by_id: Optional[dict[int, AltAuthUserRecord]] = None,
) -> Optional[str]:
    stored = getattr(team, "second_advisor_name", None)
    if stored is not None and str(stored).strip():
        return str(stored).strip()
    advisor_id = getattr(team, "second_advisor_id", None)
    if advisor_id and users_by_id:
        adv = users_by_id.get(int(advisor_id))
        if adv:
            return _display_user_name(adv, int(advisor_id))
    return None


ADVISOR_MAX_TEAMS_PER_DIVISION_TRACK = 4
ADVISOR_MAX_FIRST_TEAMS_PER_DIVISION_TRACK = 2


def _advisor_quota_team_statuses() -> Tuple[str, ...]:
    return (TeamStatus.PENDING_SCHOOL_REVIEW, TeamStatus.ACTIVE)


def _division_track_label_cn(division: Optional[str], work_track: Optional[str]) -> str:
    div = _division_label_cn(division)
    track_map = {"works": "作品", "software": "软件", "hardware": "硬件"}
    track = track_map.get(str(work_track or "").strip().lower(), str(work_track or "").strip() or "未设赛道")
    return f"{div}·{track}"


def _count_advisor_teams_in_scope(
    db: Session,
    *,
    competition_id: int,
    advisor_id: int,
    division: str,
    work_track: str,
    exclude_team_id: Optional[int] = None,
) -> Tuple[int, int]:
    """
    返回 (作为第一指导老师的队伍数, 作为第一或第二指导老师的队伍总数)。
    仅统计待校审/已通过队伍，按竞赛+组别+赛道。
    """
    div = str(division or CompetitionDivision.DEFAULT.value).strip().lower()
    track = _normalize_optional_work_track(work_track) or str(work_track or "").strip().lower()
    q = db.query(Team).filter(
        Team.competition_id == competition_id,
        Team.status.in_(_advisor_quota_team_statuses()),
        Team.division == div,
        Team.work_track == track,
        or_(
            Team.created_by_advisor_id == advisor_id,
            Team.second_advisor_id == advisor_id,
        ),
    )
    if exclude_team_id is not None:
        q = q.filter(Team.id != int(exclude_team_id))
    rows = q.all()
    as_first = sum(
        1
        for t in rows
        if t.created_by_advisor_id is not None and int(t.created_by_advisor_id) == int(advisor_id)
    )
    return as_first, len(rows)


def _assert_advisor_quota(
    db: Session,
    *,
    competition_id: int,
    advisor_id: int,
    division: str,
    work_track: str,
    as_first: bool,
    exclude_team_id: Optional[int] = None,
    advisor_label: str = "该指导老师",
) -> None:
    as_first_count, total_count = _count_advisor_teams_in_scope(
        db,
        competition_id=competition_id,
        advisor_id=advisor_id,
        division=division,
        work_track=work_track,
        exclude_team_id=exclude_team_id,
    )
    scope_label = _division_track_label_cn(division, work_track)
    if as_first and as_first_count >= ADVISOR_MAX_FIRST_TEAMS_PER_DIVISION_TRACK:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{advisor_label}在「{scope_label}」已作为第一指导老师指导 "
                f"{as_first_count} 支队伍，不得超过 "
                f"{ADVISOR_MAX_FIRST_TEAMS_PER_DIVISION_TRACK} 支"
            ),
        )
    if total_count >= ADVISOR_MAX_TEAMS_PER_DIVISION_TRACK:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{advisor_label}在「{scope_label}」指导队伍总数已达 "
                f"{total_count} 支，不得超过 {ADVISOR_MAX_TEAMS_PER_DIVISION_TRACK} 支"
                f"（含第一/第二指导老师）"
            ),
        )


def _resolve_advisor_pair_for_assignment(
    *,
    first_id: Optional[int],
    first_name: Optional[str],
    second_id: Optional[int],
    second_name: Optional[str],
    first_advisor_slot: Optional[str] = None,
) -> Tuple[Optional[int], Optional[str], Optional[int], Optional[str]]:
    """按 first_advisor_slot 交换第一/第二指导老师栏位。"""
    slot = str(first_advisor_slot or "primary").strip().lower()
    if slot == "secondary":
        return second_id, second_name, first_id, first_name
    return first_id, first_name, second_id, second_name


def _assert_team_advisors_quota(
    db: Session,
    *,
    competition_id: int,
    division: str,
    work_track: str,
    first_advisor_id: Optional[int],
    second_advisor_id: Optional[int],
    exclude_team_id: Optional[int] = None,
) -> None:
    if first_advisor_id is not None and second_advisor_id is not None:
        if int(first_advisor_id) == int(second_advisor_id):
            raise HTTPException(status_code=400, detail="第一与第二指导老师不能为同一人")
    if first_advisor_id is not None:
        _assert_advisor_quota(
            db,
            competition_id=competition_id,
            advisor_id=int(first_advisor_id),
            division=division,
            work_track=work_track,
            as_first=True,
            exclude_team_id=exclude_team_id,
            advisor_label="第一指导老师",
        )
    if second_advisor_id is not None:
        _assert_advisor_quota(
            db,
            competition_id=competition_id,
            advisor_id=int(second_advisor_id),
            division=division,
            work_track=work_track,
            as_first=False,
            exclude_team_id=exclude_team_id,
            advisor_label="第二指导老师",
        )


def _team_advisor_managed(team: Team, identity_id: int) -> bool:
    uid = int(identity_id)
    if team.created_by_advisor_id is not None and int(team.created_by_advisor_id) == uid:
        return True
    second_id = getattr(team, "second_advisor_id", None)
    return second_id is not None and int(second_id) == uid


def _can_manage_team_composition(team: Team, identity: AltAuthUserRecord) -> bool:
    """队长，或第一/第二指导老师可调整队员。"""
    if team.captain_id == identity.id:
        return True
    return _effective_alt_role(identity.role) in {"advisor", "teacher"} and _team_advisor_managed(
        team, identity.id
    )


def _normalize_school_name(raw: Optional[str]) -> str:
    return (raw or "").strip()


def _school_admin_application_status(user: AltAuthUserRecord) -> SchoolAdminApplicationStatus:
    raw = (getattr(user, "school_admin_application_status", None) or "").strip().lower()
    if raw == SchoolAdminApplicationStatus.PENDING.value:
        return SchoolAdminApplicationStatus.PENDING
    if raw == SchoolAdminApplicationStatus.APPROVED.value:
        return SchoolAdminApplicationStatus.APPROVED
    if raw == SchoolAdminApplicationStatus.REJECTED.value:
        return SchoolAdminApplicationStatus.REJECTED
    return SchoolAdminApplicationStatus.NOT_SUBMITTED


def _school_admin_can_review_teams(user: AltAuthUserRecord) -> bool:
    return (
        _effective_alt_role(user.role) == "school_admin"
        and bool(getattr(user, "school_admin_verified", False))
    )


def _require_school_admin_role(identity: AltAuthUserRecord) -> None:
    if _effective_alt_role(identity.role) != "school_admin":
        raise HTTPException(status_code=403, detail="Only school_admin can perform this operation")
    if not _normalize_school_name(getattr(identity, "school", None)):
        raise HTTPException(status_code=400, detail="School admin account must have school configured")


def _require_school_admin_identity(identity: AltAuthUserRecord) -> None:
    _require_school_admin_role(identity)
    require_permission(identity.role, Permission.REVIEW_TEAMS)
    if not _school_admin_can_review_teams(identity):
        raise HTTPException(
            status_code=403,
            detail="School admin account pending verification; submit application with photo and wait for super_admin approval",
        )


def _resolve_school_admin_photo_fs_path(stored: str) -> str:
    if not stored or not stored.strip():
        raise HTTPException(status_code=404, detail="Photo not found")
    base_dir = os.path.abspath(SCHOOL_ADMIN_PHOTO_DIR)
    if os.path.isabs(stored):
        full = os.path.abspath(stored)
    else:
        full = os.path.abspath(os.path.normpath(os.path.join(os.getcwd(), stored)))
    if not full.startswith(base_dir + os.sep) and full != base_dir:
        raise HTTPException(status_code=404, detail="Photo not found")
    if not os.path.isfile(full):
        raise HTTPException(status_code=404, detail="Photo missing on server")
    return full


async def _save_school_admin_photo(upload: StarletteUploadFile, user_id: int) -> str:
    if not upload.filename or not str(upload.filename).strip():
        raise HTTPException(status_code=400, detail="photo requires a filename")
    ext = os.path.splitext(upload.filename)[1].lower()
    if ext not in _ALLOWED_QR_EXT:
        raise HTTPException(status_code=400, detail="Unsupported photo format")
    mime = (upload.content_type or "").split(";")[0].strip().lower()
    if mime and mime not in _ALLOWED_QR_MIME:
        raise HTTPException(status_code=400, detail="Unsupported photo mime type")
    content = await upload.read()
    if len(content) > MAX_SCHOOL_ADMIN_PHOTO_BYTES:
        raise HTTPException(status_code=400, detail="Photo too large (max 5MB)")
    if not content:
        raise HTTPException(status_code=400, detail="Empty photo file")
    filename = f"user_{user_id}_{uuid.uuid4().hex}{ext}"
    rel_path = f"{SCHOOL_ADMIN_PHOTO_DIR}/{filename}"
    fs_path = os.path.abspath(os.path.join(os.getcwd(), rel_path.replace("/", os.sep)))
    base_dir = os.path.abspath(SCHOOL_ADMIN_PHOTO_DIR)
    if not fs_path.startswith(base_dir + os.sep):
        raise HTTPException(status_code=400, detail="Invalid photo path")
    with open(fs_path, "wb") as f:
        f.write(content)
    return rel_path.replace("\\", "/")


def _delete_school_admin_photo(stored: Optional[str]) -> None:
    if not stored:
        return
    try:
        fs = _resolve_school_admin_photo_fs_path(stored)
        if os.path.isfile(fs):
            os.remove(fs)
    except HTTPException:
        pass
    except Exception as e:
        logger.warning("Failed to remove school admin photo %s: %s", stored, e)


def _team_composition_open_statuses() -> Tuple[str, ...]:
    """允许调整队员/队名的队伍状态（含待校审）。"""
    return (TeamStatus.PENDING_SCHOOL_REVIEW, TeamStatus.ACTIVE)


def _add_student_to_team(
    db: Session,
    competition: Competition,
    team: Team,
    student_id: int,
    *,
    is_captain: bool = False,
) -> TeamMember:
    member = TeamMember(team_id=team.id, user_id=student_id, is_captain=is_captain)
    db.add(member)
    team_division = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    team_work_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    if not team_work_track:
        raise HTTPException(status_code=400, detail="队伍未设置赛道 work_track")
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        student_id,
        team_work_track,
        team_division,
        team_id=team.id,
        allow_same_team=True,
    )
    row_any = _get_enrollment_by_work_track(
        db, competition.id, student_id, team_work_track
    )
    if row_any and row_any.status == CompetitionEnrollmentStatus.WITHDRAWN:
        row_any.team_id = team.id
        row_any.enrollment_scope = CompetitionEnrollmentScope.TEAM
        row_any.is_captain = is_captain
        row_any.status = CompetitionEnrollmentStatus.ENROLLED
        row_any.division = team_division
        row_any.work_track = team_work_track
    elif row_any and row_any.status == CompetitionEnrollmentStatus.ENROLLED:
        row_any.team_id = team.id
        row_any.is_captain = is_captain
        row_any.division = team_division
    else:
        db.add(
            CompetitionEnrollment(
                competition_id=competition.id,
                student_id=student_id,
                team_id=team.id,
                enrollment_scope=CompetitionEnrollmentScope.TEAM,
                division=team_division,
                work_track=team_work_track,
                is_captain=is_captain,
                status=CompetitionEnrollmentStatus.ENROLLED,
            )
        )
    return member


def _build_team_join_request_response(
    req: TeamJoinRequest,
    users_by_id: dict[int, AltAuthUserRecord],
) -> TeamJoinRequestResponse:
    u = users_by_id.get(req.user_id)
    return TeamJoinRequestResponse(
        id=req.id,
        team_id=req.team_id,
        user_id=req.user_id,
        username=(u.username or "") if u else "",
        full_name=u.full_name if u else None,
        status=req.status,
        created_at=req.created_at,
        reviewed_at=req.reviewed_at,
        reviewed_by_id=req.reviewed_by_id,
    )


def _resolve_team_school(adb: Session, captain_id: int) -> str:
    captain = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == captain_id).first()
    if captain is None:
        raise HTTPException(status_code=400, detail="Captain account not found")
    school = _normalize_school_name(getattr(captain, "school", None))
    if not school:
        raise HTTPException(status_code=400, detail="Captain must have school configured")
    return school


def _team_matches_school(team: Team, school: str) -> bool:
    team_school = _normalize_school_name(getattr(team, "school", None))
    return team_school.casefold() == _normalize_school_name(school).casefold()


def _team_ids_with_submitted_work(db: Session, teams: List[Team]) -> set[int]:
    """批量判断哪些队伍已提交作品（分题正式提交或压缩包）。"""
    if not teams:
        return set()
    by_comp: dict[int, List[int]] = {}
    for t in teams:
        by_comp.setdefault(int(t.competition_id), []).append(int(t.id))
    submitted: set[int] = set()
    for cid, tids in by_comp.items():
        qa_ids = {
            int(r[0])
            for r in (
                db.query(CompetitionQuestionAnswer.team_id)
                .filter(
                    CompetitionQuestionAnswer.competition_id == cid,
                    CompetitionQuestionAnswer.team_id.in_(tids),
                    CompetitionQuestionAnswer.status == CompetitionQuestionAnswerStatus.SUBMITTED,
                )
                .distinct()
                .all()
            )
            if r and r[0] is not None
        }
        zip_ids = {
            int(r[0])
            for r in (
                db.query(Submission.team_id)
                .filter(
                    Submission.competition_id == cid,
                    Submission.team_id.in_(tids),
                )
                .distinct()
                .all()
            )
            if r and r[0] is not None
        }
        submitted |= qa_ids | zip_ids
    return submitted


def _resolved_team_captain_id(team: Team) -> Optional[int]:
    """仅以已正式入队成员为准解析队长；邀请未接受时不把 designated captain_id 当队长。"""
    members = list(team.members or [])
    for m in members:
        if m.is_captain and m.user_id is not None:
            return int(m.user_id)
    if team.captain_id is None:
        return None
    designated = int(team.captain_id)
    for m in members:
        if m.user_id is not None and int(m.user_id) == designated:
            return designated
    return None


def _assert_team_ready_for_school_approve(team: Team) -> None:
    members = list(team.members or [])
    if not members:
        raise HTTPException(
            status_code=400,
            detail="队伍尚无正式队员（邀请未被接受），无法通过审核",
        )
    if _resolved_team_captain_id(team) is None:
        raise HTTPException(
            status_code=400,
            detail="队伍尚无正式队长，无法通过审核",
        )


def _build_school_admin_team_item(
    team: Team,
    competition: Competition,
    users_by_id: dict[int, AltAuthUserRecord],
    *,
    work_submitted: bool = False,
) -> SchoolAdminTeamReviewItem:
    advisor_id = team.created_by_advisor_id
    advisor_name = _team_advisor_display_name(team, users_by_id)
    second_advisor_id = getattr(team, "second_advisor_id", None)
    second_advisor_name = _team_second_advisor_display_name(team, users_by_id)

    captain_id = _resolved_team_captain_id(team)
    captain = users_by_id.get(captain_id) if captain_id is not None else None
    captain_name = (captain.full_name or captain.username) if captain else None

    members_out: List[SchoolAdminTeamMemberItem] = []
    for m in sorted(team.members, key=lambda x: (not x.is_captain, x.joined_at or utc_now(), x.id)):
        u = users_by_id.get(m.user_id)
        members_out.append(
            SchoolAdminTeamMemberItem(
                user_id=m.user_id,
                username=u.username if u else "",
                full_name=u.full_name if u else None,
                is_captain=m.is_captain,
            )
        )

    return SchoolAdminTeamReviewItem(
        team_id=team.id,
        competition_id=competition.id,
        competition_name=competition.name,
        competition_start_at=competition.start_at,
        competition_end_at=competition.end_at,
        school=getattr(team, "school", None),
        advisor_name=advisor_name,
        advisor_id=advisor_id,
        second_advisor_name=second_advisor_name,
        second_advisor_id=int(second_advisor_id) if second_advisor_id is not None else None,
        team_name=team.name,
        captain_name=captain_name,
        captain_id=captain_id,
        members=members_out,
        division=getattr(team, "division", None),
        work_track=getattr(team, "work_track", None),
        status=TeamStatus(team.status),
        review_feedback=getattr(team, "review_feedback", None),
        reviewed_at=getattr(team, "reviewed_at", None),
        created_at=team.created_at,
        work_submitted=bool(work_submitted),
    )


def _strip_team_name(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    return s or None


def _assert_team_name_unique_in_competition(
    db: Session,
    competition_id: int,
    name: Optional[str],
    *,
    exclude_team_id: Optional[int] = None,
) -> None:
    """同一竞赛下待校审/已通过队伍的队名须唯一（忽略大小写）。"""
    tn = _strip_team_name(name)
    if not tn:
        return
    target_key = tn.casefold()
    q = (
        db.query(Team)
        .filter(
            Team.competition_id == competition_id,
            Team.status.in_(_team_composition_open_statuses()),
            Team.name.isnot(None),
        )
    )
    if exclude_team_id is not None:
        q = q.filter(Team.id != int(exclude_team_id))
    for existing in q.all():
        other = _strip_team_name(existing.name)
        if other and other.casefold() == target_key:
            raise HTTPException(status_code=400, detail="队名已存在，请更换其他队名")


def _form_optional_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _form_bool(val, default: bool) -> bool:
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s == "":
        return default
    return s in ("1", "true", "yes", "on")


def _form_optional_choice(val, allowed: Tuple[str, ...]) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"invalid value {val!r}; expected one of {list(allowed)}",
        )
    return s


def _pick_qr_upload(form, key: str):
    q = form.get(key)
    if isinstance(q, StarletteUploadFile) and q.filename and str(q.filename).strip():
        return q
    return None


def _competition_create_from_form(form) -> CompetitionCreate:
    name_raw = form.get("name")
    if name_raw is None or not str(name_raw).strip():
        raise HTTPException(status_code=400, detail="name is required")
    payload = {
        "name": str(name_raw).strip(),
        "description": _form_optional_str(form.get("description")),
        "rules_text": _form_optional_str(form.get("rules_text")),
        "target_audience": _form_optional_str(form.get("target_audience")),
        "contact_name": _form_optional_str(form.get("contact_name")),
        "contact_phone": _form_optional_str(form.get("contact_phone")),
        "location": _form_optional_str(form.get("location")),
        "environment": _form_optional_str(form.get("environment")),
        "start_at": _form_optional_str(form.get("start_at")),
        "end_at": _form_optional_str(form.get("end_at")),
        "allow_individual": _form_bool(form.get("allow_individual"), True),
        "allow_team": _form_bool(form.get("allow_team"), True),
    }
    division_mode = _form_optional_choice(form.get("division_mode"), ("single", "dual"))
    if division_mode is not None:
        payload["division_mode"] = division_mode
    qr_layout = _form_optional_choice(form.get("qr_layout"), ("shared", "separate"))
    if qr_layout is not None:
        payload["qr_layout"] = qr_layout
    stage_mode = _form_optional_choice(form.get("stage_mode"), ("single", "prelim_final"))
    if stage_mode is not None:
        payload["stage_mode"] = stage_mode
    final_start = _form_optional_str(form.get("final_start_at"))
    if final_start is not None:
        payload["final_start_at"] = final_start
    final_end = _form_optional_str(form.get("final_end_at"))
    if final_end is not None:
        payload["final_end_at"] = final_end
    try:
        return CompetitionCreate.model_validate(payload)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid form payload: {e}") from e


def _competition_update_from_form(form) -> CompetitionUpdate:
    """multipart 修改竞赛：仅解析表单中出现的字段（与 JSON exclude_unset 语义一致）。"""
    payload = {}
    if "name" in form:
        name_raw = form.get("name")
        if name_raw is None or not str(name_raw).strip():
            raise HTTPException(status_code=400, detail="name cannot be empty")
        payload["name"] = str(name_raw).strip()
    for key in (
        "description",
        "rules_text",
        "target_audience",
        "contact_name",
        "contact_phone",
        "location",
        "environment",
        "start_at",
        "end_at",
        "final_start_at",
        "final_end_at",
    ):
        if key in form:
            payload[key] = _form_optional_str(form.get(key))
    if "allow_individual" in form:
        payload["allow_individual"] = _form_bool(form.get("allow_individual"), True)
    if "allow_team" in form:
        payload["allow_team"] = _form_bool(form.get("allow_team"), True)
    if "division_mode" in form:
        division_mode = _form_optional_choice(form.get("division_mode"), ("single", "dual"))
        if division_mode is not None:
            payload["division_mode"] = division_mode
    if "qr_layout" in form:
        qr_layout = _form_optional_choice(form.get("qr_layout"), ("shared", "separate"))
        if qr_layout is not None:
            payload["qr_layout"] = qr_layout
    if "stage_mode" in form:
        stage_mode = _form_optional_choice(form.get("stage_mode"), ("single", "prelim_final"))
        if stage_mode is not None:
            payload["stage_mode"] = stage_mode
    try:
        return CompetitionUpdate.model_validate(payload)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid form payload: {e}") from e


def _delete_stored_qr_file(stored: Optional[str]) -> None:
    if not stored:
        return
    try:
        fs = _resolve_qr_fs_path(stored)
        if os.path.isfile(fs):
            os.remove(fs)
    except Exception as e:
        logger.warning("Failed to remove competition QR file %s: %s", stored, e)


def _normalize_stored_qr_path(rel_path: str) -> str:
    return rel_path.replace("\\", "/")


def _resolve_qr_fs_path(stored: str) -> str:
    """将库中记录的相对路径解析为绝对路径，并限制在 competition_qr_codes 目录内。"""
    if not stored or not stored.strip():
        raise HTTPException(status_code=404, detail="QR code not found")
    base_dir = os.path.abspath(COMPETITION_QR_DIR)
    if os.path.isabs(stored):
        full = os.path.abspath(stored)
    else:
        full = os.path.abspath(os.path.normpath(os.path.join(os.getcwd(), stored)))
    if not full.startswith(base_dir + os.sep) and full != base_dir:
        raise HTTPException(status_code=404, detail="QR code not found")
    return full


async def _save_qr_code_upload(upload: StarletteUploadFile, competition_id: int) -> str:
    if not upload.filename or not str(upload.filename).strip():
        raise HTTPException(status_code=400, detail="qr_code_image requires a filename")
    ext = os.path.splitext(upload.filename)[1].lower()
    if ext not in _ALLOWED_QR_EXT:
        raise HTTPException(
            status_code=400,
            detail=f"qr_code_image: only image extensions {sorted(_ALLOWED_QR_EXT)} allowed",
        )
    ct = (upload.content_type or "").split(";")[0].strip().lower()
    if ct and ct not in _ALLOWED_QR_MIME:
        raise HTTPException(status_code=400, detail="qr_code_image: invalid image content type")

    body = await upload.read()
    if len(body) > MAX_QR_IMAGE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"qr_code_image too large (max {MAX_QR_IMAGE_BYTES // (1024 * 1024)} MiB)",
        )

    fname = f"comp_{competition_id}_{uuid.uuid4().hex}{ext}"
    rel = _normalize_stored_qr_path(os.path.join(COMPETITION_QR_DIR, fname))
    abs_path = _resolve_qr_fs_path(rel)
    with open(abs_path, "wb") as f:
        f.write(body)
    return rel


def _delete_stored_logo_file(stored: Optional[str]) -> None:
    if not stored:
        return
    try:
        fs = _resolve_logo_fs_path(stored)
        if os.path.isfile(fs):
            os.remove(fs)
    except Exception as e:
        logger.warning("Failed to remove competition logo file %s: %s", stored, e)


def _resolve_logo_fs_path(stored: str) -> str:
    """将库中记录的相对路径解析为绝对路径，并限制在 competition_logos 目录内。"""
    if not stored or not stored.strip():
        raise HTTPException(status_code=404, detail="Logo not found")
    base_dir = os.path.abspath(COMPETITION_LOGO_DIR)
    if os.path.isabs(stored):
        full = os.path.abspath(stored)
    else:
        full = os.path.abspath(os.path.normpath(os.path.join(os.getcwd(), stored)))
    if not full.startswith(base_dir + os.sep) and full != base_dir:
        raise HTTPException(status_code=404, detail="Logo not found")
    return full


async def _save_logo_upload(upload: StarletteUploadFile, competition_id: int) -> str:
    if not upload.filename or not str(upload.filename).strip():
        raise HTTPException(status_code=400, detail="logo_image requires a filename")
    ext = os.path.splitext(upload.filename)[1].lower()
    if ext not in _ALLOWED_LOGO_EXT:
        raise HTTPException(
            status_code=400,
            detail=f"logo_image: only image extensions {sorted(_ALLOWED_LOGO_EXT)} allowed",
        )
    ct = (upload.content_type or "").split(";")[0].strip().lower()
    if ct and ct not in _ALLOWED_LOGO_MIME:
        raise HTTPException(status_code=400, detail="logo_image: invalid image content type")

    body = await upload.read()
    if len(body) > MAX_LOGO_IMAGE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"logo_image too large (max {MAX_LOGO_IMAGE_BYTES // (1024 * 1024)} MiB)",
        )

    fname = f"comp_{competition_id}_{uuid.uuid4().hex}{ext}"
    rel = _normalize_stored_qr_path(os.path.join(COMPETITION_LOGO_DIR, fname))
    abs_path = _resolve_logo_fs_path(rel)
    with open(abs_path, "wb") as f:
        f.write(body)
    return rel


def _get_competition(db: Session, competition_id: int) -> Competition:
    validate_eight_digit_id(competition_id, label="竞赛ID")
    competition = db.query(Competition).filter(Competition.id == competition_id).first()
    if not competition:
        raise HTTPException(status_code=404, detail="Competition not found")
    return competition


def _competition_stage(competition: Competition) -> str:
    return str(getattr(competition, "stage", None) or CompetitionStage.SINGLE).lower()


def _is_final_stage(competition: Competition) -> bool:
    return _competition_stage(competition) == CompetitionStage.FINAL


def _is_preliminary_stage(competition: Competition) -> bool:
    return _competition_stage(competition) == CompetitionStage.PRELIMINARY


def _user_has_final_promotion(
    db: Session,
    final_competition: Competition,
    user_id: int,
    team_id: Optional[int] = None,
) -> bool:
    """决赛准入：用户属于已晋级决赛队伍，或个人晋级名单。"""
    if team_id is not None:
        promo = (
            db.query(CompetitionPromotion)
            .filter(
                CompetitionPromotion.to_competition_id == final_competition.id,
                CompetitionPromotion.final_team_id == team_id,
            )
            .first()
        )
        if promo:
            return True
    # 作为决赛队伍成员
    member_team_ids = [
        int(r[0])
        for r in db.query(TeamMember.team_id)
        .join(Team, Team.id == TeamMember.team_id)
        .filter(
            TeamMember.user_id == user_id,
            Team.competition_id == final_competition.id,
        )
        .all()
    ]
    if member_team_ids:
        promo = (
            db.query(CompetitionPromotion)
            .filter(
                CompetitionPromotion.to_competition_id == final_competition.id,
                CompetitionPromotion.final_team_id.in_(member_team_ids),
            )
            .first()
        )
        if promo:
            return True
    promo_student = (
        db.query(CompetitionPromotion)
        .filter(
            CompetitionPromotion.to_competition_id == final_competition.id,
            CompetitionPromotion.source_student_id == user_id,
        )
        .first()
    )
    return promo_student is not None


def _assert_final_stage_participant(
    db: Session,
    competition: Competition,
    user_id: int,
    team_id: Optional[int] = None,
) -> None:
    if not _is_final_stage(competition):
        return
    if _user_has_final_promotion(db, competition, user_id, team_id=team_id):
        return
    raise HTTPException(
        status_code=403,
        detail="决赛仅晋级队伍可参加，无需重新报名；未晋级账号无法参赛",
    )


def _assert_final_stage_open_create_blocked(competition: Competition) -> None:
    """决赛不允许公开建队，须由管理员晋级时自动建队。"""
    if _is_final_stage(competition):
        raise HTTPException(
            status_code=403,
            detail="决赛沿用初赛晋级队伍，不可自行创建队伍",
        )


def _assert_final_stage_roster_frozen(competition: Competition) -> None:
    """决赛名单冻结：不可邀请/申请加入，仅晋级时复制的初赛队伍。"""
    if _is_final_stage(competition):
        raise HTTPException(
            status_code=403,
            detail="决赛沿用初赛晋级队伍名单，不可新增或申请加入队员",
        )


def _assert_competition_uses_question_answers(competition: Competition) -> None:
    """竞赛存在即可使用分题答案；具体赛道由 _assert_team_uses_question_answers 约束。"""
    if competition is None:
        raise HTTPException(status_code=404, detail="Competition not found")


def _normalize_optional_work_track(raw) -> Optional[str]:
    if raw is None:
        return None
    v = getattr(raw, "value", raw)
    s = str(v).strip().lower() if v is not None else ""
    if s in ("works", "software", "hardware"):
        return s
    return None


def _peek_team_work_track(db: Session, competition_id: int, team: Optional[Team]) -> Optional[str]:
    """队伍赛道：优先 team.work_track，否则取该队任一有效报名上的 work_track。未设置时返回 None。"""
    if team is None:
        return None
    track = _normalize_optional_work_track(getattr(team, "work_track", None))
    if track:
        return track
    row = (
        db.query(CompetitionEnrollment)
        .filter(
            CompetitionEnrollment.competition_id == competition_id,
            CompetitionEnrollment.team_id == team.id,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
        )
        .order_by(CompetitionEnrollment.id.asc())
        .first()
    )
    return _normalize_optional_work_track(getattr(row, "work_track", None) if row else None)


def _resolve_team_work_track(db: Session, competition_id: int, team: Team) -> str:
    """队伍赛道：优先 team.work_track，否则取该队任一有效报名上的 work_track。"""
    track = _peek_team_work_track(db, competition_id, team)
    if track:
        return track
    raise HTTPException(
        status_code=400,
        detail="队伍未设置赛道，请先选择作品 / 软件 / 硬件赛道后再提交",
    )


def _resolve_individual_work_track(
    db: Session, competition_id: int, student_id: int
) -> str:
    row = _get_enrollment_by_scope(
        db, competition_id, student_id, CompetitionEnrollmentScope.INDIVIDUAL
    )
    if not row or row.status != CompetitionEnrollmentStatus.ENROLLED:
        raise HTTPException(status_code=400, detail="未找到有效的个人报名")
    track = _normalize_optional_work_track(getattr(row, "work_track", None))
    if track:
        return track
    raise HTTPException(
        status_code=400,
        detail="个人报名未设置赛道，请先选择作品 / 软件 / 硬件赛道后再提交",
    )


def _assert_work_track_allows_zip(work_track: str) -> None:
    if work_track != "works":
        raise HTTPException(
            status_code=400,
            detail="仅作品赛道可上传压缩包；软件 / 硬件赛道请使用分题答案上传",
        )


def _assert_work_track_allows_question_answers(work_track: str) -> None:
    if work_track == "works":
        raise HTTPException(
            status_code=400,
            detail="作品赛道请上传压缩包作品；软件 / 硬件赛道才可使用分题答案上传",
        )
    if work_track not in ("software", "hardware"):
        raise HTTPException(
            status_code=400,
            detail="当前赛道不支持分题答案上传",
        )


def _assert_competition_uses_zip_submission(competition: Competition) -> None:
    """保留旧调用点：竞赛级不再一律禁止，改由赛道校验。"""
    if competition is None:
        raise HTTPException(status_code=404, detail="Competition not found")


def _paired_final_competition(db: Session, prelim: Competition) -> Competition:
    if not _is_preliminary_stage(prelim):
        raise HTTPException(status_code=400, detail="仅初赛竞赛支持晋级操作")
    paired_id = getattr(prelim, "paired_competition_id", None)
    if paired_id is None:
        raise HTTPException(status_code=400, detail="该初赛未关联决赛")
    final = db.query(Competition).filter(Competition.id == int(paired_id)).first()
    if not final or not _is_final_stage(final):
        raise HTTPException(status_code=400, detail="关联决赛不存在或阶段无效")
    return final


def _enroll_student_on_final_team(
    db: Session,
    final: Competition,
    student_id: int,
    team: Team,
    is_captain: bool,
    division: str,
    work_track: Optional[str],
) -> None:
    row = _get_enrollment_by_scope(
        db, final.id, student_id, CompetitionEnrollmentScope.TEAM
    )
    if row and row.status == CompetitionEnrollmentStatus.ENROLLED:
        row.team_id = team.id
        row.is_captain = is_captain
        row.division = division
        if work_track:
            row.work_track = work_track
        return
    if row and row.status == CompetitionEnrollmentStatus.WITHDRAWN:
        row.team_id = team.id
        row.enrollment_scope = CompetitionEnrollmentScope.TEAM
        row.is_captain = is_captain
        row.status = CompetitionEnrollmentStatus.ENROLLED
        row.division = division
        row.work_track = work_track
        return
    db.add(
        CompetitionEnrollment(
            competition_id=final.id,
            student_id=student_id,
            team_id=team.id,
            enrollment_scope=CompetitionEnrollmentScope.TEAM,
            division=division,
            work_track=work_track,
            is_captain=is_captain,
            status=CompetitionEnrollmentStatus.ENROLLED,
        )
    )


def _promote_prelim_team_to_final(
    db: Session,
    prelim: Competition,
    final: Competition,
    source_team: Team,
    promoted_by: int,
) -> CompetitionPromotion:
    existing = (
        db.query(CompetitionPromotion)
        .filter(
            CompetitionPromotion.to_competition_id == final.id,
            CompetitionPromotion.source_team_id == source_team.id,
        )
        .first()
    )
    if existing:
        return existing

    if source_team.status != TeamStatus.ACTIVE:
        raise HTTPException(
            status_code=400,
            detail=f"队伍 {source_team.id} 须校审通过后方可晋级",
        )
    if int(source_team.competition_id) != int(prelim.id):
        raise HTTPException(status_code=400, detail=f"队伍 {source_team.id} 不属于该初赛")

    members = (
        db.query(TeamMember).filter(TeamMember.team_id == source_team.id).all()
    )
    if not members:
        raise HTTPException(status_code=400, detail=f"队伍 {source_team.id} 无成员")

    division = str(getattr(source_team, "division", None) or "default")
    work_track = _peek_team_work_track(db, prelim.id, source_team) or getattr(
        source_team, "work_track", None
    )
    final_team = Team(
        id=allocate_eight_digit_id(db, Team),
        competition_id=final.id,
        name=source_team.name,
        captain_id=source_team.captain_id,
        created_by_advisor_id=source_team.created_by_advisor_id,
        advisor_name=source_team.advisor_name,
        second_advisor_id=getattr(source_team, "second_advisor_id", None),
        second_advisor_name=getattr(source_team, "second_advisor_name", None),
        school=source_team.school,
        division=division,
        work_track=work_track,
        status=TeamStatus.ACTIVE,
        reviewed_by_id=source_team.reviewed_by_id,
        reviewed_at=source_team.reviewed_at or utc_now(),
        review_feedback=source_team.review_feedback,
    )
    db.add(final_team)
    db.flush()

    for m in members:
        db.add(
            TeamMember(
                team_id=final_team.id,
                user_id=m.user_id,
                is_captain=bool(m.is_captain),
            )
        )
        _enroll_student_on_final_team(
            db,
            final,
            m.user_id,
            final_team,
            bool(m.is_captain),
            division,
            work_track,
        )

    promo = CompetitionPromotion(
        id=allocate_eight_digit_id(db, CompetitionPromotion),
        from_competition_id=prelim.id,
        to_competition_id=final.id,
        source_team_id=source_team.id,
        source_student_id=None,
        final_team_id=final_team.id,
        promoted_by=promoted_by,
    )
    db.add(promo)
    db.flush()
    return promo


# 分享链接匿名可读：仅已发布或已锁定（closed）的竞赛
_SHAREABLE_COMPETITION_STATUSES = ("published", "closed")


def _ensure_competition_shareable(competition: Competition) -> None:
    """未登录访客仅可查看已发布/已结束竞赛，草稿等返回 404 避免泄露。"""
    if competition.status not in _SHAREABLE_COMPETITION_STATUSES:
        raise HTTPException(status_code=404, detail="Competition not found")


def _resolve_competition_qr_storage_path(
    competition: Competition, division: Optional[str] = None
) -> Optional[str]:
    """按 division_mode / qr_layout / division 解析二维码存储路径。"""
    mode = str(getattr(competition, "division_mode", None) or "single").lower()
    layout = str(getattr(competition, "qr_layout", None) or "shared").lower()
    if mode == "dual" and layout == "separate":
        if division == "undergraduate":
            return competition.qr_code_path_undergraduate or competition.qr_code_path
        if division == "vocational":
            return competition.qr_code_path_vocational or competition.qr_code_path
        return None
    return competition.qr_code_path


def _ensure_competition_published_for_papers(competition: Competition) -> None:
    """发布试卷 / 分享 URL：须竞赛已 published 或 closed。"""
    if competition.status not in _SHAREABLE_COMPETITION_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="Competition must be published before publishing exam papers or sharing URL",
        )


def _normalize_exam_paper_division(competition: Competition, division: Optional[str]) -> str:
    """试卷组别：一律按本科 / 高职分槽（不再使用 default 不分学历槽位）。"""
    raw = (division or "").strip().lower() if division is not None else ""
    if raw in ("undergraduate", "vocational"):
        return raw
    if raw in ("", "default"):
        raise HTTPException(
            status_code=400,
            detail="须指定组别：undergraduate（本科）或 vocational（高职）",
        )
    raise HTTPException(status_code=400, detail="Invalid division")


def _resolve_enrollment_division(
    competition: Competition, requested: Optional[str]
) -> str:
    """报名/建队写入的组别：须 undergraduate（本科）或 vocational（高职）。"""
    raw = None
    if requested is not None:
        raw = getattr(requested, "value", requested)
        raw = str(raw).strip().lower() if raw is not None else ""
    if raw in ("undergraduate", "vocational"):
        return raw
    raise HTTPException(
        status_code=400,
        detail="组别必选：本科（undergraduate）或高职（vocational）",
    )


def _resolve_work_track(requested) -> str:
    """报名/建队赛道：works / software / hardware。"""
    raw = None
    if requested is not None:
        raw = getattr(requested, "value", requested)
        raw = str(raw).strip().lower() if raw is not None else ""
    if raw in ("works", "software", "hardware"):
        return raw
    raise HTTPException(
        status_code=400,
        detail="赛道必选：作品（works）/ 软件（software）/ 硬件（hardware）",
    )


def _exam_paper_path_and_filename(
    competition: Competition, division: str
) -> Tuple[Optional[str], Optional[str]]:
    if division == "undergraduate":
        return (
            getattr(competition, "exam_paper_path_undergraduate", None),
            getattr(competition, "exam_paper_filename_undergraduate", None),
        )
    if division == "vocational":
        return (
            getattr(competition, "exam_paper_path_vocational", None),
            getattr(competition, "exam_paper_filename_vocational", None),
        )
    return (
        getattr(competition, "exam_paper_path", None),
        getattr(competition, "exam_paper_filename", None),
    )


def _set_exam_paper_path_and_filename(
    competition: Competition, division: str, path: Optional[str], filename: Optional[str]
) -> None:
    if division == "undergraduate":
        competition.exam_paper_path_undergraduate = path
        competition.exam_paper_filename_undergraduate = filename
    elif division == "vocational":
        competition.exam_paper_path_vocational = path
        competition.exam_paper_filename_vocational = filename
    else:
        competition.exam_paper_path = path
        competition.exam_paper_filename = filename


def _delete_stored_exam_paper_file(stored: Optional[str]) -> None:
    if not stored:
        return
    try:
        fs = _resolve_exam_paper_fs_path(stored)
        if os.path.isfile(fs):
            os.remove(fs)
    except Exception as e:
        logger.warning("Failed to remove exam paper file %s: %s", stored, e)


def _resolve_exam_paper_fs_path(stored: str) -> str:
    if not stored or not stored.strip():
        raise HTTPException(status_code=404, detail="Exam paper not found")
    base_dir = os.path.abspath(COMPETITION_EXAM_PAPER_DIR)
    if os.path.isabs(stored):
        full = os.path.abspath(stored)
    else:
        full = os.path.abspath(os.path.normpath(os.path.join(os.getcwd(), stored)))
    if not full.startswith(base_dir + os.sep) and full != base_dir:
        raise HTTPException(status_code=404, detail="Exam paper not found")
    return full


def _exam_paper_slot(
    competition_id: int,
    path: Optional[str],
    filename: Optional[str],
    division: str,
    work_track: Optional[str] = None,
) -> CompetitionExamPaperSlot:
    published = bool(path and str(path).strip())
    download_url = None
    if published:
        q = f"division={division}"
        if work_track:
            q += f"&work_track={work_track}"
        download_url = f"/api/v1/competitions/{competition_id}/exam-papers/download?{q}"
    return CompetitionExamPaperSlot(
        published=published,
        filename=filename if published else None,
        download_url=download_url,
        work_track=work_track,
        division=division,
    )


def _build_exam_papers_meta(competition: Competition) -> CompetitionExamPapers:
    from app.competition_exam_config import WORK_TRACKS, get_exam_papers_by_track_map

    cid = competition.id
    # 试卷一律按本科 / 高职分槽；仍读取旧 default 槽以便兼容展示/回退
    divisions = ["undergraduate", "vocational", "default"]
    by_track = {}
    track_map = get_exam_papers_by_track_map(competition)

    for div in divisions:
        by_track[div] = {}
        for track in WORK_TRACKS:
            path = None
            name = None
            # 严格按「组别 + 赛道」取值，禁止把其它赛道/旧组别槽位串到本赛道
            if div in track_map and track in track_map[div]:
                path = track_map[div][track].get("path")
                name = track_map[div][track].get("filename")
            by_track[div][track] = _exam_paper_slot(cid, path, name, div, track)

    def _legacy_slot(div: str) -> CompetitionExamPaperSlot:
        # 顶层 undergraduate/vocational 仅表示旧「按组别、不分赛道」槽位，不再取某赛道文件冒充
        path, name = _exam_paper_path_and_filename(competition, div)
        return _exam_paper_slot(cid, path, name, div)

    # 对外主字段仅本科/高职；by_track 可含 default 供兼容
    return CompetitionExamPapers(
        undergraduate=_legacy_slot("undergraduate"),
        vocational=_legacy_slot("vocational"),
        default=_legacy_slot("default"),
        by_track=by_track,
    )


async def _save_exam_paper_upload(
    upload: StarletteUploadFile,
    competition_id: int,
    division: str,
    work_track: Optional[str] = None,
) -> Tuple[str, str]:
    if not upload.filename or not str(upload.filename).strip():
        raise HTTPException(status_code=400, detail="exam paper requires a filename")
    original = os.path.basename(str(upload.filename).strip())
    ext = os.path.splitext(original)[1].lower()
    if ext not in _ALLOWED_EXAM_PAPER_EXT:
        raise HTTPException(
            status_code=400,
            detail=f"exam paper: only extensions {sorted(_ALLOWED_EXAM_PAPER_EXT)} allowed",
        )
    body = await upload.read()
    if len(body) > MAX_EXAM_PAPER_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"exam paper too large (max {MAX_EXAM_PAPER_BYTES // (1024 * 1024)} MiB)",
        )
    safe_div = division.replace("/", "_")
    track_part = f"_{work_track}" if work_track else ""
    fname = f"comp_{competition_id}_{safe_div}{track_part}_{uuid.uuid4().hex}{ext}"
    rel = _normalize_stored_qr_path(os.path.join(COMPETITION_EXAM_PAPER_DIR, fname))
    abs_path = _resolve_exam_paper_fs_path(rel)
    with open(abs_path, "wb") as f:
        f.write(body)
    return rel, original[:255]


def _exam_paper_requires_division_match(competition: Competition) -> bool:
    """试卷已按本科/高职分槽，下载时须与报名/队伍组别一致。"""
    return True


def _can_download_exam_paper(
    db: Session,
    competition: Competition,
    identity: AltAuthUserRecord,
    division: str,
    work_track: Optional[str] = None,
) -> bool:
    role = _effective_alt_role(identity.role)
    cid = competition.id
    match_division = _exam_paper_requires_division_match(competition)
    track = (work_track or "").strip().lower() or None
    if role == "student":
        q = db.query(CompetitionEnrollment).filter(
            CompetitionEnrollment.competition_id == cid,
            CompetitionEnrollment.student_id == identity.id,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
        )
        if match_division:
            q = q.filter(CompetitionEnrollment.division == division)
        if track:
            q = q.filter(CompetitionEnrollment.work_track == track)
        rows = q.all()
        if not rows:
            return False
        # 组队报名：仅队伍校审通过（active）可下载；个人报名无队伍则可下
        for row in rows:
            if row.team_id is None:
                return True
            team = db.query(Team).filter(Team.id == row.team_id).first()
            if team is not None and team.status == TeamStatus.ACTIVE:
                return True
        return False
    # 指导老师 / 教师不可下载试卷
    return False


def _resolve_identity_work_track_for_paper(
    db: Session,
    competition: Competition,
    identity: AltAuthUserRecord,
    division: str,
    requested: Optional[str],
) -> str:
    raw = (requested or "").strip().lower() if requested is not None else ""
    if raw in ("works", "software", "hardware"):
        return raw
    role = _effective_alt_role(identity.role)
    match_division = _exam_paper_requires_division_match(competition)
    if role == "student":
        q = db.query(CompetitionEnrollment).filter(
            CompetitionEnrollment.competition_id == competition.id,
            CompetitionEnrollment.student_id == identity.id,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
        )
        if match_division:
            q = q.filter(CompetitionEnrollment.division == division)
        row = q.order_by(CompetitionEnrollment.id.desc()).first()
        track = str(getattr(row, "work_track", None) or "").strip().lower() if row else ""
        if track in ("works", "software", "hardware"):
            return track
    if role in {"advisor", "teacher"}:
        q = db.query(Team).filter(
            Team.competition_id == competition.id,
            Team.status == TeamStatus.ACTIVE,
            or_(
                Team.created_by_advisor_id == identity.id,
                Team.second_advisor_id == identity.id,
            ),
        )
        if match_division:
            q = q.filter(Team.division == division)
        team = q.order_by(Team.id.desc()).first()
        track = str(getattr(team, "work_track", None) or "").strip().lower() if team else ""
        if track in ("works", "software", "hardware"):
            return track
    raise HTTPException(
        status_code=400,
        detail="请指定赛道 work_track（works/software/hardware），或先完成对应赛道报名/建队",
    )


def _is_enrollment_closed(competition: Competition) -> bool:
    """
    「停止报名」条件（与业务上“锁定竞赛”含义一致）：
    1) 状态为 closed（管理员手动关闭报名）
    2) 到达/超过 end_at（自动停止报名）
    """
    if competition.status == "closed":
        return True
    if competition.end_at is None:
        return False
    return utc_now() >= ensure_utc(competition.end_at)


def _ensure_enrollment_open(competition: Competition) -> None:
    """禁止在已停止报名后：新报名 / 新建队伍 / 加入队伍。"""
    if _is_enrollment_closed(competition):
        raise HTTPException(
            status_code=400,
            detail="Competition enrollment is closed (status closed or past end date).",
        )


def _ensure_competition_allows_submissions(competition: Competition) -> None:
    """
    作品提交：已发布（published）或已锁定报名（closed）均允许已参赛用户继续提交；
    草稿（draft）不允许。
    """
    if competition.status not in ("published", "closed"):
        raise HTTPException(
            status_code=400,
            detail="Competition is not accepting submissions (must be published or closed)",
        )


def _ensure_competition_ended_for_export(competition: Competition) -> None:
    """赛后导出：报名已关闭或已过结束时间。"""
    if not _is_enrollment_closed(competition):
        raise HTTPException(
            status_code=400,
            detail="竞赛尚未结束，结束后才可导出答案（状态为「已结束/closed」或已过结束时间）",
        )


def _validate_question_no(
    question_no: int,
    competition: Optional[Competition] = None,
    work_track: Optional[str] = None,
) -> int:
    from app.competition_exam_config import get_competition_question_count

    max_n = (
        get_competition_question_count(competition, work_track)
        if competition is not None
        else COMPETITION_QUESTION_COUNT
    )
    max_n = max(1, min(COMPETITION_QUESTION_COUNT, int(max_n)))
    if question_no < 1 or question_no > max_n:
        raise HTTPException(
            status_code=400,
            detail=f"question_no must be between 1 and {max_n}",
        )
    return question_no


def _safe_export_filename(name: str, fallback: str = "answer") -> str:
    base = os.path.basename(name or "").strip() or fallback
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", base).strip(" .")
    return base or fallback


def _team_export_label(team: Team) -> str:
    """导出压缩包内文件夹/文件名：优先队伍名；空名时回退「队伍{id}」。"""
    raw = (getattr(team, "name", None) or "").strip()
    if raw:
        return _safe_export_filename(raw, fallback=f"队伍{team.id}")
    return _safe_export_filename(f"队伍{team.id}", fallback=f"队伍{team.id}")


def _unique_team_export_labels(teams: list) -> dict:
    """
    为导出生成唯一队伍标签：同名队伍在队名后追加 _队伍ID，避免 zip 内重名覆盖。
    返回 {team_id: label}。
    """
    counts: dict[str, int] = {}
    for team in teams:
        label = _team_export_label(team)
        counts[label] = counts.get(label, 0) + 1
    used: dict[str, int] = {}
    out: dict = {}
    for team in teams:
        base = _team_export_label(team)
        if counts.get(base, 0) <= 1:
            out[int(team.id)] = base
            continue
        # 重名：队名_队伍ID
        unique = _safe_export_filename(f"{base}_{team.id}", fallback=f"队伍{team.id}")
        # 极端情况下仍冲突再追加序号
        n = used.get(unique, 0)
        used[unique] = n + 1
        if n > 0:
            unique = _safe_export_filename(f"{unique}_{n + 1}", fallback=f"队伍{team.id}")
        out[int(team.id)] = unique
    return out


def _question_folder_name(
    question_no: int,
    competition: Optional[Competition] = None,
    work_track: Optional[str] = None,
) -> str:
    from app.competition_exam_config import question_name_map

    if competition is not None:
        names = question_name_map(competition, work_track)
        if question_no in names:
            return names[question_no]
    return f"第{question_no}题"


def _question_answer_download_filename(
    *,
    team: Optional[Team],
    competition: Optional[Competition],
    question_no: int,
    work_track: Optional[str],
    original_filename: Optional[str],
) -> str:
    """下载名：队伍名+题目名称，扩展名保持原文件不变。"""
    raw_name = (original_filename or "").strip() or "answer"
    _root, ext = os.path.splitext(os.path.basename(raw_name))
    team_name = ""
    if team is not None:
        team_name = (getattr(team, "name", None) or "").strip()
        if not team_name:
            team_name = f"队伍{team.id}"
    team_part = team_name or "队伍"
    question_part = _question_folder_name(question_no, competition, work_track)
    stem = _safe_export_filename(f"{team_part}{question_part}", fallback="answer")
    return f"{stem}{ext}" if ext else stem


def _question_answer_response(
    db: Session, row: CompetitionQuestionAnswer
) -> CompetitionQuestionAnswerResponse:
    file_record = db.query(FileModel).filter(FileModel.id == row.file_id).first()
    return CompetitionQuestionAnswerResponse(
        id=row.id,
        competition_id=row.competition_id,
        team_id=row.team_id,
        question_no=row.question_no,
        submitter_id=row.submitter_id,
        file_id=row.file_id,
        filename=file_record.filename if file_record else None,
        status=getattr(row, "status", None) or CompetitionQuestionAnswerStatus.DRAFT,
        uploaded_at=row.uploaded_at,
        submitted_at=getattr(row, "submitted_at", None),
    )


def _slot_from_row(
    db: Session, question_no: int, row: Optional[CompetitionQuestionAnswer]
) -> CompetitionQuestionAnswerSlot:
    if not row:
        return CompetitionQuestionAnswerSlot(
            question_no=question_no, uploaded=False, submitted=False, answer=None
        )
    status = getattr(row, "status", None) or CompetitionQuestionAnswerStatus.DRAFT
    return CompetitionQuestionAnswerSlot(
        question_no=question_no,
        uploaded=True,
        submitted=status == CompetitionQuestionAnswerStatus.SUBMITTED,
        answer=_question_answer_response(db, row),
    )


def _require_active_team_member_for_answers(
    db: Session,
    competition: Competition,
    team_id: int,
    identity: AltAuthUserRecord,
) -> Team:
    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.competition_id == competition.id)
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    if team.status != TeamStatus.ACTIVE:
        raise HTTPException(
            status_code=400,
            detail="Team must be approved by school admin before uploading answers",
        )
    member = (
        db.query(TeamMember)
        .filter(TeamMember.team_id == team.id, TeamMember.user_id == identity.id)
        .first()
    )
    if not member:
        raise HTTPException(status_code=403, detail="User is not a team member")
    return team


def _can_view_team_question_answers(
    db: Session, competition_id: int, team_id: int, identity: AltAuthUserRecord
) -> bool:
    if _can_view_all_competition_submissions(db, competition_id, identity):
        return True
    if _expert_gate_ok(identity) and _is_assigned_expert_for_team(
        db, competition_id, team_id, identity.id
    ):
        return True
    if _effective_alt_role(identity.role) != "student":
        return False
    return (
        db.query(TeamMember)
        .join(Team)
        .filter(
            TeamMember.user_id == identity.id,
            TeamMember.team_id == team_id,
            Team.competition_id == competition_id,
        )
        .first()
        is not None
    )


def _team_has_formal_submitted_answers(db: Session, competition_id: int, team_id: int) -> bool:
    row = (
        db.query(CompetitionQuestionAnswer.id)
        .filter(
            CompetitionQuestionAnswer.competition_id == competition_id,
            CompetitionQuestionAnswer.team_id == team_id,
            CompetitionQuestionAnswer.status == CompetitionQuestionAnswerStatus.SUBMITTED,
        )
        .first()
    )
    return row is not None


def _team_has_zip_submission(db: Session, competition_id: int, team_id: int) -> bool:
    row = (
        db.query(Submission.id)
        .filter(
            Submission.competition_id == competition_id,
            Submission.team_id == team_id,
        )
        .first()
    )
    return row is not None


def _team_has_submitted_work(db: Session, competition_id: int, team_id: int) -> bool:
    """队伍是否已正式提交作品（分题答案或压缩包任一即可）。"""
    return _team_has_formal_submitted_answers(db, competition_id, team_id) or _team_has_zip_submission(
        db, competition_id, team_id
    )


def _purge_prior_zip_submissions(
    db: Session,
    competition_id: int,
    *,
    team_id: Optional[int] = None,
    student_id: Optional[int] = None,
) -> None:
    """同一队伍（或同一学生个人赛道）再次上传压缩包时，删除旧提交，仅保留即将写入的最新记录。"""
    q = db.query(Submission).filter(Submission.competition_id == int(competition_id))
    if team_id is not None:
        q = q.filter(Submission.team_id == int(team_id))
    else:
        if student_id is None:
            return
        q = q.filter(Submission.team_id.is_(None), Submission.student_id == int(student_id))
    old_rows = q.all()
    if not old_rows:
        return
    old_ids = [int(s.id) for s in old_rows]
    db.query(Review).filter(Review.submission_id.in_(old_ids)).delete(synchronize_session=False)
    for s in old_rows:
        db.delete(s)
    db.flush()


def _keep_latest_submissions_only(rows: List[Submission]) -> List[Submission]:
    """列表去重：有 team_id 的按队伍保留最新一条；个人提交按 student_id 保留最新一条。"""
    if not rows:
        return []

    def _recency(s: Submission):
        ts = ensure_utc(s.submitted_at) if getattr(s, "submitted_at", None) else None
        epoch = ts.timestamp() if ts is not None else 0.0
        return (epoch, int(getattr(s, "id", 0) or 0))

    ordered = sorted(rows, key=_recency, reverse=True)
    seen: set = set()
    out: List[Submission] = []
    for s in ordered:
        if s.team_id is not None:
            key = f"team:{int(s.team_id)}"
        else:
            key = f"student:{int(s.student_id)}"
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


def _team_names_by_ids(db: Session, team_ids: set) -> dict:
    ids = {int(x) for x in team_ids if x is not None}
    if not ids:
        return {}
    rows = db.query(Team.id, Team.name).filter(Team.id.in_(ids)).all()
    return {int(tid): (name or "").strip() or f"队伍{tid}" for tid, name in rows}


def _submission_to_response(
    submission: Submission,
    team_names: Optional[dict] = None,
) -> SubmissionResponse:
    base = SubmissionResponse.model_validate(submission)
    team_name = None
    if submission.team_id is not None:
        tid = int(submission.team_id)
        if team_names and tid in team_names:
            team_name = team_names[tid]
        else:
            team_name = f"队伍{tid}"
    return base.model_copy(update={"team_name": team_name})


def _submissions_to_responses(db: Session, rows: List[Submission]) -> List[SubmissionResponse]:
    team_ids = {int(s.team_id) for s in rows if s.team_id is not None}
    names = _team_names_by_ids(db, team_ids)
    return [_submission_to_response(s, names) for s in rows]


def _submission_download_filename(db: Session, submission: Submission, file_record: FileModel) -> str:
    """下载文件名：队伍提交优先用队伍名；保留原扩展名。"""
    original = (file_record.filename or "").strip() or os.path.basename(file_record.file_path or "") or "submission.zip"
    _root, ext = os.path.splitext(os.path.basename(original))
    if not ext:
        ext = ".zip"
    if submission.team_id is not None:
        team = db.query(Team).filter(Team.id == int(submission.team_id)).first()
        label = _team_export_label(team) if team is not None else f"队伍{int(submission.team_id)}"
        return f"{label}{ext}"
    safe = _safe_export_filename(original, fallback=f"submission_{submission.id}{ext}")
    if not os.path.splitext(safe)[1]:
        return f"{safe}{ext}"
    return safe


def _assert_team_answers_not_locked(db: Session, competition_id: int, team_id: int) -> None:
    if _team_has_formal_submitted_answers(db, competition_id, team_id):
        raise HTTPException(status_code=400, detail="作品已正式提交，无法再上传或删除题目文件")


def _assert_team_roster_mutable(db: Session, competition_id: int, team_id: int) -> None:
    if _team_has_submitted_work(db, competition_id, team_id):
        raise HTTPException(
            status_code=400,
            detail="作品已提交，无法再变更队伍成员（邀请、移除、转让或退队）",
        )


def _assert_team_school_meta_mutable(db: Session, competition_id: int, team_id: int) -> None:
    """队伍提交作品后，禁止校管修改指导老师 / 组别 / 赛道。"""
    if _team_has_submitted_work(db, competition_id, team_id):
        raise HTTPException(
            status_code=400,
            detail="作品已提交，无法再添加/修改指导老师或修改组别/赛道",
        )


def _assert_student_same_school_as_team(
    adb: Session,
    team: Team,
    student_id: int,
    *,
    label: str = "学生",
) -> None:
    """组队要求：队长与队员须属同一学校。"""
    team_school = _normalize_school_name(getattr(team, "school", None))
    if not team_school:
        team_school = _resolve_team_school(adb, int(team.captain_id))
    stu = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == int(student_id)).first()
    if stu is None:
        raise HTTPException(status_code=404, detail=f"{label}不存在")
    stu_school = _normalize_school_name(getattr(stu, "school", None))
    if not stu_school:
        raise HTTPException(status_code=400, detail=f"{label}未配置学校，无法组队")
    if stu_school.casefold() != team_school.casefold():
        raise HTTPException(
            status_code=400,
            detail="仅允许同一学校的学生组队，该学生与队伍/队长学校不一致",
        )


def _assert_student_ids_same_school(adb: Session, student_ids: List[int]) -> str:
    """校验多名学生同一学校，返回规范化校名。"""
    ids = []
    seen = set()
    for sid in student_ids:
        n = int(sid)
        if n not in seen:
            seen.add(n)
            ids.append(n)
    if not ids:
        raise HTTPException(status_code=400, detail="缺少学生信息")
    schools: List[str] = []
    for sid in ids:
        u = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == sid).first()
        if u is None:
            raise HTTPException(status_code=404, detail=f"学生 {sid} 不存在")
        sch = _normalize_school_name(getattr(u, "school", None))
        if not sch:
            name = (u.full_name or u.username or str(sid)).strip()
            raise HTTPException(status_code=400, detail=f"学生「{name}」未配置学校，无法组队")
        schools.append(sch)
    base = schools[0]
    for sch in schools[1:]:
        if sch.casefold() != base.casefold():
            raise HTTPException(
                status_code=400,
                detail="仅允许同一学校的学生组队，队长与队员必须属于同一学校",
            )
    return base


def _build_team_invite_response(
    inv: TeamInvite,
    *,
    team: Optional[Team] = None,
    competition: Optional[Competition] = None,
    inviter: Optional[AltAuthUserRecord] = None,
) -> TeamInviteResponse:
    inviter_name = None
    if inviter is not None:
        inviter_name = (inviter.full_name or inviter.username or "").strip() or None
    return TeamInviteResponse(
        id=inv.id,
        team_id=inv.team_id,
        competition_id=inv.competition_id,
        competition_name=(competition.name if competition else None),
        team_name=(team.name if team else None),
        invitee_id=inv.invitee_id,
        inviter_id=inv.inviter_id,
        inviter_name=inviter_name,
        as_captain=bool(inv.as_captain),
        status=inv.status,
        created_at=inv.created_at,
        responded_at=inv.responded_at,
    )


def _create_pending_team_invite(
    db: Session,
    *,
    team: Team,
    competition: Competition,
    invitee_id: int,
    inviter_id: int,
    as_captain: bool = False,
) -> TeamInvite:
    existing = (
        db.query(TeamInvite)
        .filter(
            TeamInvite.team_id == team.id,
            TeamInvite.invitee_id == invitee_id,
            TeamInvite.status == TeamInviteStatus.PENDING,
        )
        .first()
    )
    if existing:
        existing.as_captain = bool(as_captain)
        existing.inviter_id = inviter_id
        return existing
    inv = TeamInvite(
        team_id=team.id,
        competition_id=competition.id,
        invitee_id=invitee_id,
        inviter_id=inviter_id,
        as_captain=bool(as_captain),
        status=TeamInviteStatus.PENDING,
    )
    db.add(inv)
    return inv


def _cancel_pending_team_side_effects(db: Session, team: Team, *, now=None) -> None:
    """队伍驳回时关闭待处理入队邀请与入队申请，避免学生端卡住。"""
    ts = now or utc_now()
    pending_invites = (
        db.query(TeamInvite)
        .filter(
            TeamInvite.team_id == team.id,
            TeamInvite.status == TeamInviteStatus.PENDING,
        )
        .all()
    )
    for inv in pending_invites:
        inv.status = TeamInviteStatus.CANCELLED
        inv.responded_at = ts
    pending_joins = (
        db.query(TeamJoinRequest)
        .filter(
            TeamJoinRequest.team_id == team.id,
            TeamJoinRequest.status == TeamJoinRequestStatus.PENDING,
        )
        .all()
    )
    for req in pending_joins:
        req.status = TeamJoinRequestStatus.REJECTED
        req.reviewed_at = ts


def _handle_team_invite_rejected(db: Session, inv: TeamInvite, team: Optional[Team]) -> None:
    """
    指定队长拒绝邀请且队伍尚无正式成员时：自动驳回待审队伍，避免校管看到「空壳队」并误通过。
    若已有其他正式队员，则将队长转给队内成员。
    """
    if team is None:
        return
    if team.status not in _team_composition_open_statuses():
        return

    is_designated_captain = bool(inv.as_captain) or (
        team.captain_id is not None and int(team.captain_id) == int(inv.invitee_id)
    )
    if not is_designated_captain:
        return

    already_member = (
        db.query(TeamMember)
        .filter(TeamMember.team_id == team.id, TeamMember.user_id == inv.invitee_id)
        .first()
    )
    if already_member:
        return

    members = (
        db.query(TeamMember)
        .filter(TeamMember.team_id == team.id)
        .order_by(TeamMember.id.asc())
        .all()
    )
    now = utc_now()
    if not members:
        if team.status == TeamStatus.PENDING_SCHOOL_REVIEW:
            team.status = TeamStatus.REJECTED
            team.reviewed_at = now
            team.review_feedback = "指定队长拒绝入队邀请，队伍已自动取消待审"
            _cancel_pending_team_side_effects(db, team, now=now)
        return

    # 已有队员：把队长转给队内第一人（优先已有 is_captain）
    new_cap = next((m for m in members if m.is_captain), members[0])
    for m in members:
        m.is_captain = int(m.user_id) == int(new_cap.user_id)
    team.captain_id = int(new_cap.user_id)


def _build_question_answers_board(
    db: Session, competition_id: int, team_id: int
) -> CompetitionQuestionAnswersBoard:
    from app.competition_exam_config import get_competition_question_count

    competition = db.query(Competition).filter(Competition.id == competition_id).first()
    team = db.query(Team).filter(Team.id == team_id, Team.competition_id == competition_id).first()
    track = None
    if team is not None:
        try:
            track = _resolve_team_work_track(db, competition_id, team)
        except HTTPException:
            track = str(getattr(team, "work_track", None) or "").strip().lower() or None
    q_count = (
        get_competition_question_count(competition, track)
        if competition
        else COMPETITION_QUESTION_COUNT
    )
    rows = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.competition_id == competition_id,
            CompetitionQuestionAnswer.team_id == team_id,
        )
        .all()
    )
    by_q = {r.question_no: r for r in rows}
    slots: List[CompetitionQuestionAnswerSlot] = []
    submitted_count = 0
    draft_count = 0
    for q in range(1, q_count + 1):
        row = by_q.get(q)
        slot = _slot_from_row(db, q, row)
        slots.append(slot)
        if slot.submitted:
            submitted_count += 1
        elif slot.uploaded:
            draft_count += 1
    return CompetitionQuestionAnswersBoard(
        competition_id=competition_id,
        team_id=team_id,
        question_count=q_count,
        submitted_count=submitted_count,
        draft_count=draft_count,
        slots=slots,
    )


def _write_answer_file_into_zip(
    zf: zipfile.ZipFile,
    arcname_dir: str,
    file_record: FileModel,
) -> None:
    """将答案文件写入 zip 内指定目录；源文件缺失时写入占位说明。"""
    folder = arcname_dir.rstrip("/") + "/"
    # 保证空目录也有条目
    zf.writestr(folder, "")
    if not file_record or not file_record.file_path or not os.path.exists(file_record.file_path):
        zf.writestr(folder + "MISSING.txt", "答案文件在服务器上缺失\n")
        return
    fname = _safe_export_filename(file_record.filename or os.path.basename(file_record.file_path))
    zf.write(file_record.file_path, folder + fname)


def _filter_teams_by_work_track(
    db: Session,
    competition_id: int,
    teams: List[Team],
    work_track: Optional[str],
) -> List[Team]:
    track = (work_track or "").strip().lower() or None
    if not track:
        return list(teams)
    out: List[Team] = []
    for team in teams:
        try:
            t = _resolve_team_work_track(db, competition_id, team)
        except HTTPException:
            t = str(getattr(team, "work_track", None) or "").strip().lower() or ""
        if t == track:
            out.append(team)
    return out


def _build_answers_export_zip(
    db: Session,
    competition: Competition,
    mode: str,
    work_track: Optional[str] = None,
    allowed_team_ids: Optional[set] = None,
) -> BytesIO:
    from app.competition_exam_config import get_competition_question_count

    teams = (
        db.query(Team)
        .filter(Team.competition_id == competition.id, Team.status == TeamStatus.ACTIVE)
        .order_by(Team.id.asc())
        .all()
    )
    track = (work_track or "").strip().lower() or None
    if track in ("software", "hardware"):
        teams = _filter_teams_by_work_track(db, competition.id, teams, track)
    if allowed_team_ids is not None:
        teams = [t for t in teams if int(t.id) in allowed_team_ids]
    q_count = get_competition_question_count(competition, track)
    q_count = max(1, min(COMPETITION_QUESTION_COUNT, int(q_count)))

    answers = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.competition_id == competition.id,
            CompetitionQuestionAnswer.status == CompetitionQuestionAnswerStatus.SUBMITTED,
        )
        .all()
    )
    team_ids = {t.id for t in teams}
    if track in ("software", "hardware"):
        answers = [a for a in answers if a.team_id in team_ids]
    file_ids = {a.file_id for a in answers}
    files_by_id = {
        f.id: f for f in db.query(FileModel).filter(FileModel.id.in_(file_ids)).all()
    } if file_ids else {}
    answers_map: dict[tuple[int, int], CompetitionQuestionAnswer] = {
        (a.team_id, a.question_no): a for a in answers
    }
    team_labels = _unique_team_export_labels(teams)

    outer = BytesIO()
    with zipfile.ZipFile(outer, "w", compression=zipfile.ZIP_DEFLATED) as outer_zf:
        if mode == "by_team":
            for team in teams:
                inner_buf = BytesIO()
                with zipfile.ZipFile(inner_buf, "w", compression=zipfile.ZIP_DEFLATED) as inner_zf:
                    for q in range(1, q_count + 1):
                        folder = _question_folder_name(q, competition, track)
                        ans = answers_map.get((team.id, q))
                        if ans:
                            _write_answer_file_into_zip(
                                inner_zf, folder, files_by_id.get(ans.file_id)
                            )
                        else:
                            inner_zf.writestr(folder + "/", "")
                inner_buf.seek(0)
                team_label = team_labels.get(int(team.id)) or _team_export_label(team)
                outer_zf.writestr(f"{team_label}.zip", inner_buf.read())
        elif mode == "by_question":
            for q in range(1, q_count + 1):
                inner_buf = BytesIO()
                with zipfile.ZipFile(inner_buf, "w", compression=zipfile.ZIP_DEFLATED) as inner_zf:
                    for team in teams:
                        folder = team_labels.get(int(team.id)) or _team_export_label(team)
                        ans = answers_map.get((team.id, q))
                        if ans:
                            _write_answer_file_into_zip(
                                inner_zf, folder, files_by_id.get(ans.file_id)
                            )
                        else:
                            inner_zf.writestr(folder + "/", "")
                inner_buf.seek(0)
                q_name = _safe_export_filename(
                    _question_folder_name(q, competition, track), fallback=f"第{q}题"
                )
                outer_zf.writestr(f"{q_name}.zip", inner_buf.read())
        else:
            raise HTTPException(status_code=400, detail="mode must be by_team or by_question")

    outer.seek(0)
    return outer


def _build_works_submissions_export_zip(
    db: Session,
    competition: Competition,
    allowed_team_ids: Optional[set] = None,
) -> BytesIO:
    """作品赛道：按队伍导出压缩包作品，外层 zip 内含「队伍名.zip」。"""
    teams = (
        db.query(Team)
        .filter(Team.competition_id == competition.id, Team.status == TeamStatus.ACTIVE)
        .order_by(Team.id.asc())
        .all()
    )
    works_teams = _filter_teams_by_work_track(db, competition.id, teams, "works")
    works_team_ids = {t.id for t in works_teams}
    if allowed_team_ids is not None:
        works_team_ids = {tid for tid in works_team_ids if int(tid) in allowed_team_ids}
        works_teams = [t for t in works_teams if int(t.id) in works_team_ids]
    teams_by_id = {int(t.id): t for t in works_teams}
    team_labels = _unique_team_export_labels(works_teams)

    submissions = (
        db.query(Submission)
        .filter(Submission.competition_id == competition.id)
        .order_by(Submission.submitted_at.desc())
        .all()
    )
    # 每队取最新一条带文件的提交；个人作品赛（无 team_id）按 student_id 归组
    latest_by_key: dict[str, Submission] = {}
    for sub in submissions:
        if sub.team_id is not None:
            if int(sub.team_id) not in works_team_ids:
                continue
            key = f"team:{int(sub.team_id)}"
        else:
            if allowed_team_ids is not None:
                continue
            # 个人提交：须对应作品赛道有效报名
            enroll = (
                db.query(CompetitionEnrollment)
                .filter(
                    CompetitionEnrollment.competition_id == competition.id,
                    CompetitionEnrollment.student_id == sub.student_id,
                    CompetitionEnrollment.enrollment_scope == CompetitionEnrollmentScope.INDIVIDUAL,
                    CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
                )
                .first()
            )
            track = _normalize_optional_work_track(getattr(enroll, "work_track", None) if enroll else None)
            if track != "works":
                continue
            key = f"student:{int(sub.student_id)}"
        if key in latest_by_key:
            continue
        if not sub.file_id:
            continue
        latest_by_key[key] = sub

    file_ids = {s.file_id for s in latest_by_key.values() if s.file_id}
    files_by_id = {
        f.id: f for f in db.query(FileModel).filter(FileModel.id.in_(file_ids)).all()
    } if file_ids else {}

    outer = BytesIO()
    with zipfile.ZipFile(outer, "w", compression=zipfile.ZIP_DEFLATED) as outer_zf:
        for key, sub in sorted(latest_by_key.items(), key=lambda kv: kv[0]):
            inner_buf = BytesIO()
            with zipfile.ZipFile(inner_buf, "w", compression=zipfile.ZIP_DEFLATED) as inner_zf:
                _write_answer_file_into_zip(
                    inner_zf, "作品", files_by_id.get(sub.file_id)
                )
            inner_buf.seek(0)
            if sub.team_id is not None:
                tid = int(sub.team_id)
                label = team_labels.get(tid)
                if not label:
                    team = teams_by_id.get(tid)
                    label = _team_export_label(team) if team else f"队伍{tid}"
                arc = f"{label}.zip"
            else:
                arc = f"student_{int(sub.student_id)}.zip"
            outer_zf.writestr(arc, inner_buf.read())

    outer.seek(0)
    return outer


def _enrollment_scope_for_team(team_id: Optional[int]) -> str:
    return (
        CompetitionEnrollmentScope.TEAM
        if team_id is not None
        else CompetitionEnrollmentScope.INDIVIDUAL
    )


def _list_enrollments_for_student(
    db: Session,
    competition_id: int,
    student_id: int,
    *,
    enrolled_only: bool = False,
) -> List[CompetitionEnrollment]:
    q = db.query(CompetitionEnrollment).filter(
        CompetitionEnrollment.competition_id == competition_id,
        CompetitionEnrollment.student_id == student_id,
    )
    if enrolled_only:
        q = q.filter(CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED)
    return q.order_by(CompetitionEnrollment.id.asc()).all()


def _get_enrollment_by_scope(
    db: Session,
    competition_id: int,
    student_id: int,
    scope: str,
    work_track: Optional[str] = None,
) -> Optional[CompetitionEnrollment]:
    """
    按 scope 取报名。若传入 work_track 则精确匹配该赛道；
    未传时返回该 scope 下任意一条（兼容旧调用，优先 enrolled）。
    """
    track = _normalize_optional_work_track(work_track)
    q = db.query(CompetitionEnrollment).filter(
        CompetitionEnrollment.competition_id == competition_id,
        CompetitionEnrollment.student_id == student_id,
        CompetitionEnrollment.enrollment_scope == scope,
    )
    if track:
        q = q.filter(CompetitionEnrollment.work_track == track)
        return q.first()
    rows = q.order_by(
        CompetitionEnrollment.status.asc(),  # enrolled 通常排在前（字母序）
        CompetitionEnrollment.id.asc(),
    ).all()
    for row in rows:
        if row.status == CompetitionEnrollmentStatus.ENROLLED:
            return row
    return rows[0] if rows else None


def _get_enrollment_by_work_track(
    db: Session,
    competition_id: int,
    student_id: int,
    work_track: str,
) -> Optional[CompetitionEnrollment]:
    track = _normalize_optional_work_track(work_track)
    if not track:
        return None
    return (
        db.query(CompetitionEnrollment)
        .filter(
            CompetitionEnrollment.competition_id == competition_id,
            CompetitionEnrollment.student_id == student_id,
            CompetitionEnrollment.work_track == track,
        )
        .first()
    )


def _has_active_enrollment_in_scope(
    db: Session,
    competition_id: int,
    student_id: int,
    scope: str,
    work_track: Optional[str] = None,
) -> bool:
    if work_track:
        row = _get_enrollment_by_scope(
            db, competition_id, student_id, scope, work_track=work_track
        )
        return row is not None and row.status == CompetitionEnrollmentStatus.ENROLLED
    rows = _list_enrollments_for_student(
        db, competition_id, student_id, enrolled_only=True
    )
    return any(r.enrollment_scope == scope for r in rows)


def _active_work_tracks_for_student(
    db: Session, competition_id: int, student_id: int
) -> List[str]:
    tracks: List[str] = []
    for row in _list_enrollments_for_student(
        db, competition_id, student_id, enrolled_only=True
    ):
        t = _normalize_optional_work_track(getattr(row, "work_track", None))
        if t and t not in tracks:
            tracks.append(t)
    return tracks


def _assert_student_can_enroll_work_track(
    db: Session,
    competition_id: int,
    student_id: int,
    work_track: str,
    division: str,
    *,
    team_id: Optional[int] = None,
    allow_same_team: bool = True,
) -> None:
    """
    同一竞赛：作品/软件/硬件各最多一次；组别须与已有有效报名一致；
    同赛道已报名其他队伍则拒绝。
    """
    track = _normalize_optional_work_track(work_track) or _resolve_work_track(work_track)
    div = str(division or CompetitionDivision.DEFAULT.value).strip().lower()
    existing = _get_enrollment_by_work_track(db, competition_id, student_id, track)
    if existing and existing.status == CompetitionEnrollmentStatus.ENROLLED:
        if (
            allow_same_team
            and team_id is not None
            and existing.team_id is not None
            and int(existing.team_id) == int(team_id)
        ):
            return
        raise HTTPException(
            status_code=400,
            detail=f"已在该竞赛报名赛道 {track}，同一赛道不可重复报名",
        )

    active = _list_enrollments_for_student(
        db, competition_id, student_id, enrolled_only=True
    )
    for row in active:
        row_div = str(getattr(row, "division", None) or "").strip().lower()
        if row_div and div and row_div != "default" and div != "default" and row_div != div:
            raise HTTPException(
                status_code=400,
                detail="同一竞赛只能选择一个学历组别（本科或高职）",
            )

    tracks = _active_work_tracks_for_student(db, competition_id, student_id)
    if track not in tracks and len(tracks) >= 3:
        raise HTTPException(
            status_code=400,
            detail="同一竞赛最多报名作品/软件/硬件三个赛道",
        )


def _ensure_active_individual_enrollment(db: Session, competition_id: int, user_id: int) -> None:
    """个人赛道提交：必须存在「有效个人报名」（individual 赛道、enrolled）。"""
    if not _has_active_enrollment_in_scope(
        db, competition_id, user_id, CompetitionEnrollmentScope.INDIVIDUAL
    ):
        raise HTTPException(
            status_code=403,
            detail="Not actively enrolled as individual participant in this competition",
        )


def _individual_sequence_no(db: Session, competition_id: int, enrollment: CompetitionEnrollment) -> int:
    """本竞赛个人赛道内序号（从 1 起，按报名时间、id 稳定排序）。"""
    n = (
        db.query(func.count(CompetitionEnrollment.id))
        .filter(
            CompetitionEnrollment.competition_id == competition_id,
            CompetitionEnrollment.enrollment_scope == CompetitionEnrollmentScope.INDIVIDUAL,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
            or_(
                CompetitionEnrollment.created_at < enrollment.created_at,
                and_(
                    CompetitionEnrollment.created_at == enrollment.created_at,
                    CompetitionEnrollment.id <= enrollment.id,
                ),
            ),
        )
        .scalar()
    )
    return int(n or 0)


def _team_sequence_no(db: Session, competition_id: int, team: Team) -> int:
    """本竞赛组队赛道内队伍序号（从 1 起，按队伍创建时间、队伍 id 排序）。"""
    n = (
        db.query(func.count(Team.id))
        .filter(
            Team.competition_id == competition_id,
            Team.status == TeamStatus.ACTIVE,
            or_(
                Team.created_at < team.created_at,
                and_(Team.created_at == team.created_at, Team.id <= team.id),
            ),
        )
        .scalar()
    )
    return int(n or 0)


def _ensure_submission_access(
    db: Session, submission: Submission, identity: AltAuthUserRecord
) -> None:
    """
    作品访问边界：
    - super_admin：可访问竞赛内全部
    - 已核验且被指派的 expert：仅可访问其被指派队伍的作品
    - student：仅可访问本人个人作品，或其所在队伍的提交作品
    """
    r = _effective_alt_role(identity.role)
    cid = submission.competition_id

    if r == "super_admin":
        return

    if r == "expert":
        if not _expert_gate_ok(identity):
            raise HTTPException(status_code=403, detail="Access denied")
        if submission.team_id is None:
            raise HTTPException(status_code=403, detail="Access denied")
        if not _is_assigned_expert_for_team(db, cid, submission.team_id, identity.id):
            raise HTTPException(status_code=403, detail="Access denied")
        return

    if r != "student":
        raise HTTPException(status_code=403, detail="Access denied")

    if submission.team_id is None:
        if submission.student_id != identity.id:
            raise HTTPException(status_code=403, detail="Access denied")
        return

    member = db.query(TeamMember).filter(
        TeamMember.team_id == submission.team_id,
        TeamMember.user_id == identity.id,
    ).first()
    if not member:
        raise HTTPException(status_code=403, detail="Access denied")


def _display_user_name(u: Optional[AltAuthUserRecord], fallback_id: Optional[int] = None) -> str:
    if not u:
        return str(fallback_id) if fallback_id is not None else "-"
    if u.full_name and str(u.full_name).strip():
        return str(u.full_name).strip()
    if u.username and str(u.username).strip():
        return str(u.username).strip()
    return str(fallback_id) if fallback_id is not None else str(u.id)


def _is_expert_anonymized_viewer(
    db: Session, competition_id: int, identity: AltAuthUserRecord
) -> bool:
    """专家阅卷视图：隐藏队员姓名/学校/指导老师等，保留真实队名。超管不受限。"""
    if _effective_alt_role(identity.role) == "super_admin":
        return False
    return _is_competition_assigned_expert(db, competition_id, identity)


def _anonymize_team_member_response(m: TeamMemberWithUserResponse) -> TeamMemberWithUserResponse:
    return TeamMemberWithUserResponse(
        id=m.id,
        team_id=m.team_id,
        user_id=m.user_id,
        username="",
        full_name=None,
        is_captain=m.is_captain,
        joined_at=m.joined_at,
    )


def _build_team_member_user_responses(
    members: list[TeamMember],
    users_by_id: dict[int, AltAuthUserRecord],
    *,
    anonymize: bool = False,
) -> list[TeamMemberWithUserResponse]:
    out: list[TeamMemberWithUserResponse] = []
    for m in sorted(members, key=lambda x: (not x.is_captain, x.joined_at or utc_now(), x.id)):
        u = users_by_id.get(m.user_id)
        item = TeamMemberWithUserResponse(
                id=m.id,
                team_id=m.team_id,
                user_id=m.user_id,
            username=(u.username if u else "") if not anonymize else "",
            full_name=(u.full_name if u else None) if not anonymize else None,
                is_captain=m.is_captain,
                joined_at=m.joined_at,
            )
        out.append(item)
    return out


def _team_detail_response(
    adb: Session, team: Team, *, anonymize: bool = False
) -> TeamDetailResponse:
    member_uids = {m.user_id for m in team.members}
    advisor_ids = set()
    if team.created_by_advisor_id is not None:
        advisor_ids.add(team.created_by_advisor_id)
    if getattr(team, "second_advisor_id", None) is not None:
        advisor_ids.add(int(team.second_advisor_id))
    users_by_id = _alt_users_by_id(adb, member_uids | advisor_ids | {team.captain_id})
    members_out = _build_team_member_user_responses(
        team.members, users_by_id, anonymize=anonymize
    )
    raw_div = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).strip().lower()
    try:
        division = CompetitionDivision(raw_div)
    except ValueError:
        division = CompetitionDivision.DEFAULT
    raw_track = getattr(team, "work_track", None)
    work_track = None
    if raw_track is not None and str(raw_track).strip():
        try:
            work_track = CompetitionWorkTrack(str(raw_track).strip().lower())
        except ValueError:
            work_track = None
    if anonymize:
        return TeamDetailResponse(
            id=team.id,
            competition_id=team.competition_id,
            name=team.name,
            captain_id=team.captain_id,
            created_by_advisor_id=None,
            advisor_name=None,
            second_advisor_id=None,
            second_advisor_name=None,
            division=division,
            work_track=work_track,
            status=TeamStatus(team.status),
            review_feedback=getattr(team, "review_feedback", None),
            created_at=team.created_at,
            members=members_out,
        )
    display_name = _team_advisor_display_name(team, users_by_id)
    advisor_name = display_name if display_name else getattr(team, "advisor_name", None)
    second_name = _team_second_advisor_display_name(team, users_by_id)
    return TeamDetailResponse(
        id=team.id,
        competition_id=team.competition_id,
        name=team.name,
        captain_id=team.captain_id,
        created_by_advisor_id=team.created_by_advisor_id,
        advisor_name=advisor_name,
        second_advisor_id=getattr(team, "second_advisor_id", None),
        second_advisor_name=second_name if second_name else getattr(team, "second_advisor_name", None),
        division=division,
        work_track=work_track,
        status=TeamStatus(team.status),
        review_feedback=getattr(team, "review_feedback", None),
        created_at=team.created_at,
        members=members_out,
    )


def _team_detail_responses(
    adb: Session, teams: list[Team], *, anonymize: bool = False
) -> list[TeamDetailResponse]:
    return [_team_detail_response(adb, team, anonymize=anonymize) for team in teams]


def _competition_stage_label(competition: Competition) -> str:
    stage = str(getattr(competition, "stage", None) or "single").lower()
    if stage == CompetitionStage.PRELIMINARY:
        return "初赛"
    if stage == CompetitionStage.FINAL:
        return "决赛"
    return "单阶段"


def _division_label_cn(raw: Optional[str]) -> str:
    v = str(raw or "").strip().lower()
    if v == "undergraduate":
        return "本科"
    if v == "vocational":
        return "高职"
    return v or "-"


def _work_track_label_cn(raw: Optional[str]) -> str:
    v = str(raw or "").strip().lower()
    return {
        "works": "作品",
        "software": "软件",
        "hardware": "硬件",
    }.get(v, v or "-")


def _work_track_section_label(raw: Optional[str]) -> str:
    """作品提交分区标题 / 导出压缩包名：作品赛道、软件赛道、硬件赛道。"""
    v = str(raw or "").strip().lower()
    return {
        "works": "作品赛道",
        "software": "软件赛道",
        "hardware": "硬件赛道",
    }.get(v, v or "赛道")


def _content_disposition_attachment(filename: str) -> str:
    """ASCII fallback + RFC5987 UTF-8 filename*，便于中文压缩包名下载。"""
    from urllib.parse import quote

    safe = _safe_export_filename(filename, fallback="export.zip")
    ascii_name = re.sub(r"[^\x20-\x7e]", "_", safe) or "export.zip"
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(safe)}"


def _division_track_project_label(division: Optional[str], work_track: Optional[str]) -> str:
    """组别+赛道，如「本科软件组」。"""
    div = _division_label_cn(division)
    track = _work_track_label_cn(work_track)
    parts = []
    if div and div != "-":
        parts.append(div)
    if track and track != "-":
        parts.append(track)
    if not parts:
        return "-"
    return f"{''.join(parts)}组"


def _question_score_headers_for_track(competition: Competition, work_track: str) -> List[str]:
    """队员之后的分题列表头：取自发布试卷时的分题配置名称。"""
    from app.competition_exam_config import get_competition_question_config

    cfg = get_competition_question_config(competition, work_track)
    headers: List[str] = []
    for q in cfg.get("questions") or []:
        try:
            no = int(q.get("no"))
        except (TypeError, ValueError):
            continue
        if no < 1 or no > COMPETITION_QUESTION_COUNT:
            continue
        name = str(q.get("name") or f"第{no}题").strip() or f"第{no}题"
        headers.append(name)
    if not headers:
        headers = [f"第{i}题" for i in range(1, 6)]
    return headers


def _grade_score_cells(grade: Optional[CompetitionTeamQuestionGrade], question_count: int) -> list:
    """按配置题数输出分题分数单元格。"""
    n = max(1, min(COMPETITION_QUESTION_COUNT, int(question_count) or 5))
    attrs = ("score_q1", "score_q2", "score_q3", "score_q4", "score_q5")
    if grade is None:
        return [""] * n
    return [getattr(grade, attrs[i], None) for i in range(n)]


def _append_team_mapping_rows(
    ws,
    *,
    competition: Competition,
    teams: list[Team],
    users_by_id: dict[int, AltAuthUserRecord],
    grades_by_team: Optional[dict] = None,
    question_count: int = 5,
    include_scores: bool = True,
) -> None:
    """对照表：一行一支队伍；表头含分题列与总分，include_scores=false 时分数单元格留空。"""
    grades_by_team = grades_by_team or {}
    q_count = max(1, min(COMPETITION_QUESTION_COUNT, int(question_count) or 5))
    for team in teams:
        advisor = _team_advisor_display_name(team, users_by_id) or "-"
        school = (getattr(team, "school", None) or "").strip() or "-"
        team_name = (team.name or "").strip() or f"队伍{team.id}"
        name_and_advisor = team_name if advisor in ("", "-") else f"{team_name} / {advisor}"
        members = sorted(team.members, key=lambda x: (0 if x.is_captain else 1, x.id))
        member_labels = []
        for m in members:
            if m.is_captain:
                continue
            if team.captain_id is not None and int(m.user_id) == int(team.captain_id):
                continue
            u = users_by_id.get(m.user_id)
            member_name = _display_user_name(u, m.user_id)
            member_labels.append(member_name)
        members_cell = "、".join(member_labels) if member_labels else "-"
        row = [
            school,
            competition.name or "-",
            _division_track_project_label(
                getattr(team, "division", None),
                getattr(team, "work_track", None),
            ),
            team.id,
            name_and_advisor,
            members_cell,
        ]
        if include_scores:
            grade = grades_by_team.get(int(team.id))
            row.extend(_grade_score_cells(grade, q_count))
            row.append(grade.total_score if grade else "")
        else:
            row.extend([""] * q_count)
            row.append("")
        ws.append(row)


def _load_active_teams_with_members(db: Session, competition_id: int) -> list[Team]:
    return (
        db.query(Team)
        .options(joinedload(Team.members))
        .filter(Team.competition_id == competition_id, Team.status == TeamStatus.ACTIVE)
        .order_by(Team.created_at.asc(), Team.id.asc())
        .all()
    )


def _team_captain_label(team: Team, users_by_id: dict[int, AltAuthUserRecord]) -> Tuple[Optional[int], Optional[str]]:
    """返回 (captain_id, captain_display_name)。仅统计已正式入队成员。"""
    captain_id = _resolved_team_captain_id(team)
    if captain_id is None:
        return None, None
    u = users_by_id.get(captain_id)
    return captain_id, _display_user_name(u, captain_id)


def _team_members_label(
    team: Team,
    users_by_id: dict[int, AltAuthUserRecord],
    *,
    exclude_captain: bool = True,
) -> str:
    """队员姓名列表；默认不含队长。"""
    members = sorted(team.members or [], key=lambda x: (0 if x.is_captain else 1, x.id))
    labels = []
    for m in members:
        if exclude_captain and m.is_captain:
            continue
        if exclude_captain and team.captain_id is not None and int(m.user_id) == int(team.captain_id):
            continue
        u = users_by_id.get(m.user_id)
        name = _display_user_name(u, m.user_id)
        labels.append(name)
    return "、".join(labels) if labels else "-"


def _filter_teams_by_division_query(teams: list[Team], division: Optional[str]) -> list[Team]:
    raw = str(division or "").strip().lower()
    if raw in ("", "default", "all"):
        return teams
    if raw not in ("undergraduate", "vocational"):
        return teams
    return [t for t in teams if str(getattr(t, "division", None) or "").strip().lower() == raw]


def _build_competition_score_team_items(
    *,
    teams: list[Team],
    users_by_id: dict[int, AltAuthUserRecord],
    grades_by_team: dict[int, CompetitionTeamQuestionGrade],
) -> List[CompetitionScoreTeamItem]:
    items: List[CompetitionScoreTeamItem] = []
    for team in teams:
        grade = grades_by_team.get(int(team.id))
        advisor = _team_advisor_display_name(team, users_by_id) or None
        school = (getattr(team, "school", None) or "").strip() or None
        team_name = (team.name or "").strip() or f"队伍{team.id}"
        captain_id, captain_name = _team_captain_label(team, users_by_id)
        items.append(
            CompetitionScoreTeamItem(
                team_id=int(team.id),
                team_name=team_name,
                school=school,
                advisor_name=advisor,
                captain_name=captain_name,
                captain_id=captain_id,
                members=_team_members_label(team, users_by_id, exclude_captain=True),
                score_q1=float(grade.score_q1) if grade else None,
                score_q2=float(grade.score_q2) if grade else None,
                score_q3=float(grade.score_q3) if grade else None,
                score_q4=float(grade.score_q4) if grade else None,
                score_q5=float(grade.score_q5) if grade else None,
                total_score=float(grade.total_score) if grade else None,
                graded=grade is not None,
                feedback=grade.feedback if grade else None,
            )
        )
    return items


def _parse_score_division_param(raw: Optional[str]) -> CompetitionDivision:
    v = str(raw or "").strip().lower()
    if v == "undergraduate":
        return CompetitionDivision.UNDERGRADUATE
    if v == "vocational":
        return CompetitionDivision.VOCATIONAL
    return CompetitionDivision.DEFAULT


@router.get("/{competition_id}/teams/export")
async def export_team_roster_excel(
    competition_id: int,
    scope: str = Query(
        "current",
        description="current=仅本场；paired=仅关联场次；both=初赛+决赛对照表",
    ),
    include_scores: bool = Query(
        True,
        description="是否填充分题分数与总分；false 时仍保留题目名称与总分表头，分数单元格留空",
    ),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    管理员导出参赛对照表：按作品/软件/硬件赛道各生成一份 Excel，打成 zip。
    每份表格列：学校名称、竞赛名称、组别项目、队伍编码、队伍名称指导老师、队员、
    分题列（表头取自该赛道发布试卷时的分题配置名称）、总分。
    include_scores=false 时分题列与总分表头保留，单元格不填分数。
    scope=both 时导出初赛+决赛全部队伍。
    """
    from app.competition_exam_config import WORK_TRACKS

    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    scope_norm = (scope or "current").strip().lower()
    if scope_norm not in ("current", "paired", "both"):
        raise HTTPException(status_code=400, detail="scope must be current, paired, or both")

    comps: list[Competition] = []
    if scope_norm == "current":
        comps = [competition]
    elif scope_norm == "paired":
        paired_id = getattr(competition, "paired_competition_id", None)
        if not paired_id:
            raise HTTPException(status_code=400, detail="当前竞赛无关联初赛/决赛")
        paired = _get_competition(db, int(paired_id))
        comps = [paired]
    else:
        comps = [competition]
        paired_id = getattr(competition, "paired_competition_id", None)
        if paired_id:
            other = _get_competition(db, int(paired_id))
            if _is_final_stage(competition):
                comps = [other, competition]
            elif str(getattr(competition, "stage", "") or "").lower() == CompetitionStage.PRELIMINARY:
                comps = [competition, other]
            else:
                comps = [competition, other]

    # 预加载各场次队伍、用户、成绩
    teams_by_comp: dict[int, list[Team]] = {}
    users_by_comp: dict[int, dict[int, AltAuthUserRecord]] = {}
    grades_by_comp: dict[int, dict[int, CompetitionTeamQuestionGrade]] = {}
    for comp in comps:
        teams = _load_active_teams_with_members(db, comp.id)
        teams_by_comp[comp.id] = teams
        all_uids: set[int] = set()
        for t in teams:
            all_uids.add(t.captain_id)
            if t.created_by_advisor_id:
                all_uids.add(t.created_by_advisor_id)
            for m in t.members:
                all_uids.add(m.user_id)
        users_by_comp[comp.id] = _alt_users_by_id(adb, all_uids)
        if include_scores:
            grade_rows = (
                db.query(CompetitionTeamQuestionGrade)
                .filter(CompetitionTeamQuestionGrade.competition_id == comp.id)
                .all()
            )
            grades_by_comp[comp.id] = {int(g.team_id): g for g in grade_rows}

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for track in WORK_TRACKS:
            # 表头以当前请求竞赛的该赛道分题配置为准（发布试卷时填写的题名）
            q_headers = _question_score_headers_for_track(competition, track)
            q_count = len(q_headers)

            wb = Workbook()
            ws = wb.active
            ws.title = _work_track_section_label(track)[:31] or "对照表"
            headers = [
                "学校名称",
                "竞赛名称",
                "组别项目",
                "队伍编码",
                "队伍名称指导老师",
                "队员",
                *q_headers,
                "总分",
            ]
            ws.append(headers)

            for comp in comps:
                all_teams = teams_by_comp.get(comp.id) or []
                track_teams = [
                    t
                    for t in all_teams
                    if _peek_team_work_track(db, comp.id, t) == track
                ]
                if not track_teams:
                    continue
                _append_team_mapping_rows(
                    ws,
                    competition=comp,
                    teams=track_teams,
                    users_by_id=users_by_comp.get(comp.id) or {},
                    grades_by_team=grades_by_comp.get(comp.id) or {},
                    question_count=q_count,
                    include_scores=include_scores,
                )

            xlsx_buf = BytesIO()
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            # 包内文件名：赛道英文 key + 中文名，便于 Windows 正确识别
            arcname = f"{track}_{_work_track_section_label(track)}.xlsx"
            info = zipfile.ZipInfo(filename=arcname)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.flag_bits |= 0x800  # UTF-8 文件名
            zf.writestr(info, xlsx_buf.getvalue())

    payload = zip_buf.getvalue()
    suffix = "roster" if include_scores else "participants"
    zip_name = f"competition_{competition_id}_{suffix}_{scope_norm}.zip"
    return Response(
        content=payload,
        media_type="application/zip",
        headers={
            "Content-Disposition": _content_disposition_attachment(zip_name),
            "Content-Type": "application/zip",
        },
    )


@router.post("/", response_model=CompetitionResponse, status_code=status.HTTP_201_CREATED)
async def create_competition(
    request: Request,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    创建竞赛。
    - **application/json**：与原先一致，请求体为 CompetitionCreate（无文件上传）。
    - **multipart/form-data**：文本字段与 JSON 相同（`name`、`description`…），可选文件字段
      **`qr_code_image`** 上传二维码、**`logo_image`** 上传竞赛 Logo（均为 png/jpg/gif/webp，最大 5MiB）。
    - **stage_mode=prelim_final**：一次创建初赛+决赛两场；`name` 为系列名，自动生成「{name}-初赛」「{name}-决赛」；
      `start_at`/`end_at` 为初赛时间，`final_start_at`/`final_end_at` 为决赛时间。返回初赛对象（含 paired_competition_id）。
    """
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)

    ct = (request.headers.get("content-type") or "").lower()
    qr_upload = None
    qr_upload_undergraduate = None
    qr_upload_vocational = None
    logo_upload = None
    if "application/json" in ct:
        try:
            body = await request.json()
        except Exception as e:
            raise HTTPException(status_code=400, detail="Invalid JSON body") from e
        competition = CompetitionCreate.model_validate(body)
    elif "multipart/form-data" in ct:
        form = await request.form()
        competition = _competition_create_from_form(form)
        qr_upload = _pick_qr_upload(form, "qr_code_image")
        qr_upload_undergraduate = _pick_qr_upload(form, "qr_code_image_undergraduate")
        qr_upload_vocational = _pick_qr_upload(form, "qr_code_image_vocational")
        logo_upload = _pick_qr_upload(form, "logo_image") or _pick_qr_upload(form, "logo")
    else:
        raise HTTPException(
            status_code=415,
            detail="Content-Type must be application/json or multipart/form-data",
        )

    division_mode = getattr(competition.division_mode, "value", competition.division_mode)
    qr_layout = getattr(competition.qr_layout, "value", competition.qr_layout)
    stage_mode = getattr(competition.stage_mode, "value", competition.stage_mode) or "single"
    series_name = (competition.name or "").strip()
    if not series_name:
        raise HTTPException(status_code=400, detail="name is required")

    async def _apply_uploads(comp: Competition) -> None:
        if qr_upload is not None:
            comp.qr_code_path = await _save_qr_code_upload(qr_upload, comp.id)
        if qr_upload_undergraduate is not None:
            comp.qr_code_path_undergraduate = await _save_qr_code_upload(
                qr_upload_undergraduate, comp.id
            )
        if qr_upload_vocational is not None:
            comp.qr_code_path_vocational = await _save_qr_code_upload(
                qr_upload_vocational, comp.id
            )
        if logo_upload is not None:
            comp.logo_path = await _save_logo_upload(logo_upload, comp.id)

    if stage_mode == "prelim_final":
        prelim_id = allocate_eight_digit_id(db, Competition)
        final_id = allocate_eight_digit_id(db, Competition, used_extra=[prelim_id])
        series_id = prelim_id
        prelim = Competition(
            id=prelim_id,
            name=f"{series_name}-初赛",
            description=competition.description,
            rules_text=competition.rules_text,
            target_audience=competition.target_audience,
            contact_name=competition.contact_name,
            contact_phone=competition.contact_phone,
            location=competition.location,
            environment=competition.environment,
            status="draft",
            start_at=competition.start_at,
            end_at=competition.end_at,
            allow_individual=competition.allow_individual,
            allow_team=competition.allow_team,
            division_mode=division_mode or "single",
            qr_layout=qr_layout or "shared",
            series_id=series_id,
            stage=CompetitionStage.PRELIMINARY,
            paired_competition_id=final_id,
            qr_code_path=None,
            qr_code_path_undergraduate=None,
            qr_code_path_vocational=None,
            logo_path=None,
        )
        final = Competition(
            id=final_id,
            name=f"{series_name}-决赛",
            description=competition.description,
            rules_text=competition.rules_text,
            target_audience=competition.target_audience,
            contact_name=competition.contact_name,
            contact_phone=competition.contact_phone,
            location=competition.location,
            environment=competition.environment,
            status="draft",
            start_at=competition.final_start_at,
            end_at=competition.final_end_at,
            allow_individual=competition.allow_individual,
            allow_team=competition.allow_team,
            division_mode=division_mode or "single",
            qr_layout=qr_layout or "shared",
            series_id=series_id,
            stage=CompetitionStage.FINAL,
            paired_competition_id=prelim_id,
            qr_code_path=None,
            qr_code_path_undergraduate=None,
            qr_code_path_vocational=None,
            logo_path=None,
        )
        db.add(prelim)
        db.add(final)
        db.flush()
        await _apply_uploads(prelim)
        # 决赛共用同一套二维码/Logo 文件路径引用（同相对路径可读）
        final.qr_code_path = prelim.qr_code_path
        final.qr_code_path_undergraduate = prelim.qr_code_path_undergraduate
        final.qr_code_path_vocational = prelim.qr_code_path_vocational
        final.logo_path = prelim.logo_path
        db.commit()
        db.refresh(prelim)
        return prelim

    comp = Competition(
        id=allocate_eight_digit_id(db, Competition),
        name=series_name,
        description=competition.description,
        rules_text=competition.rules_text,
        target_audience=competition.target_audience,
        contact_name=competition.contact_name,
        contact_phone=competition.contact_phone,
        location=competition.location,
        environment=competition.environment,
        status="draft",
        start_at=competition.start_at,
        end_at=competition.end_at,
        allow_individual=competition.allow_individual,
        allow_team=competition.allow_team,
        division_mode=division_mode or "single",
        qr_layout=qr_layout or "shared",
        series_id=None,
        stage=CompetitionStage.SINGLE,
        paired_competition_id=None,
        qr_code_path=None,
        qr_code_path_undergraduate=None,
        qr_code_path_vocational=None,
        logo_path=None,
    )
    db.add(comp)
    db.flush()
    await _apply_uploads(comp)
    # 单阶段：series_id 取自身 id，便于后续扩展
    comp.series_id = comp.id
    db.commit()
    db.refresh(comp)
    return comp


@router.get("/", response_model=List[CompetitionResponse])
async def list_competitions(
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    q = db.query(Competition)
    if _effective_alt_role(identity.role) == "expert":
        assigned_ids = [
            int(row[0])
            for row in db.query(CompetitionExpertAssignment.competition_id)
            .filter(CompetitionExpertAssignment.expert_id == identity.id)
            .all()
        ]
        if not assigned_ids:
            return []
        q = q.filter(Competition.id.in_(assigned_ids))
    return q.order_by(Competition.created_at.desc()).all()


@router.get("/experts", response_model=CompetitionExpertsListResponse)
async def list_all_experts(
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    管理员获取全部第二套专家帐号（``role=expert``），含各专家已指派的竞赛 id 列表。
    指派 / 取消指派仍使用 ``POST|DELETE /{competition_id}/experts/{expert_user_id}``。
    """
    _require_super_admin_identity(identity)

    expert_rows = (
        adb.query(AltAuthUserRecord)
        .filter(AltAuthUserRecord.role == "expert")
        .order_by(AltAuthUserRecord.id.asc())
        .all()
    )
    expert_ids = [u.id for u in expert_rows]

    assignments_by_expert: dict[int, List[int]] = {eid: [] for eid in expert_ids}
    teams_by_expert: dict[int, List[CompetitionExpertAssignedTeam]] = {eid: [] for eid in expert_ids}
    if expert_ids:
        assign_rows = (
            db.query(
                CompetitionExpertAssignment.expert_id,
                CompetitionExpertAssignment.competition_id,
            )
            .filter(CompetitionExpertAssignment.expert_id.in_(expert_ids))
            .all()
        )
        for expert_id, comp_id in assign_rows:
            assignments_by_expert.setdefault(int(expert_id), []).append(int(comp_id))

        team_rows = (
            db.query(
                CompetitionExpertTeamAssignment.expert_id,
                CompetitionExpertTeamAssignment.competition_id,
                CompetitionExpertTeamAssignment.team_id,
                Team.name,
                Team.division,
                Team.work_track,
            )
            .outerjoin(Team, Team.id == CompetitionExpertTeamAssignment.team_id)
            .filter(CompetitionExpertTeamAssignment.expert_id.in_(expert_ids))
            .all()
        )
        for expert_id, comp_id, team_id, team_name, division, work_track in team_rows:
            teams_by_expert.setdefault(int(expert_id), []).append(
                CompetitionExpertAssignedTeam(
                    competition_id=int(comp_id),
                    team_id=int(team_id),
                    team_name=team_name,
                    division=str(division) if division is not None else None,
                    work_track=str(work_track) if work_track is not None else None,
                )
            )

    items: List[CompetitionExpertListItem] = []
    for u in expert_rows:
        cids = sorted(set(assignments_by_expert.get(u.id, [])))
        assigned_teams = sorted(
            teams_by_expert.get(u.id, []),
            key=lambda t: (t.competition_id, t.team_id),
        )
        items.append(
            CompetitionExpertListItem(
                expert_user_id=u.id,
                username=u.username or "",
                email=u.email,
                phone=getattr(u, "phone", None),
                full_name=u.full_name,
                school=u.school,
                expert_verified=bool(getattr(u, "expert_verified", False)),
                expert_review_feedback=getattr(u, "expert_review_feedback", None),
                is_active=bool(getattr(u, "is_active", True)),
                assigned_competition_ids=cids,
                assigned_teams=assigned_teams,
            )
        )

    return CompetitionExpertsListResponse(total=len(items), items=items)


@router.get("/{competition_id}", response_model=CompetitionResponse)
async def get_competition_detail(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: Optional[AltAuthUserRecord] = Depends(get_optional_alt_identity),
):
    """
    获取单条竞赛详情（§8.1.1）。

    - **已登录**：须具备 VIEW_COMPETITIONS，可查看含草稿在内的竞赛。
    - **未登录**：仅可查看 status 为 published / closed 的竞赛（分享链接场景，无需 Bearer）。
    """
    competition = _get_competition(db, competition_id)
    if identity is None:
        _ensure_competition_shareable(competition)
    else:
        require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    return competition


@router.get("/{competition_id}/qr-code")
async def get_competition_qr_code(
    competition_id: int,
    division: Optional[str] = Query(
        None,
        description="dual 且 qr_layout=separate 时必填：undergraduate | vocational",
    ),
    db: Session = Depends(get_db),
    identity: Optional[AltAuthUserRecord] = Depends(get_optional_alt_identity),
):
    """
    下载/查看竞赛二维码图片。

    - **已登录**：须具备 VIEW_COMPETITIONS。
    - **未登录**：仅已发布/已结束竞赛可读（分享链接，无需 Bearer）。
    """
    competition = _get_competition(db, competition_id)
    if identity is None:
        _ensure_competition_shareable(competition)
    else:
        require_permission(identity.role, Permission.VIEW_COMPETITIONS)

    if division is not None and division not in ("undergraduate", "vocational"):
        raise HTTPException(
            status_code=400,
            detail="division must be undergraduate or vocational",
        )

    stored = _resolve_competition_qr_storage_path(competition, division)
    if not stored:
        raise HTTPException(
            status_code=404,
            detail="No QR code for this competition (division may be required)",
        )
    fs_path = _resolve_qr_fs_path(stored)
    if not os.path.isfile(fs_path):
        raise HTTPException(status_code=404, detail="QR code file missing on server")
    mime, _ = mimetypes.guess_type(fs_path)
    # 禁止浏览器/代理长期缓存：同一 URL 换图后否则会一直显示旧二维码（磁盘缓存）
    try:
        mtime = int(os.path.getmtime(fs_path))
    except OSError:
        mtime = 0
    return FileResponse(
        path=fs_path,
        filename=os.path.basename(fs_path),
        media_type=mime or "application/octet-stream",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            "ETag": f'W/"qr-{competition_id}-{mtime}-{os.path.basename(fs_path)}"',
        },
    )


@router.get("/{competition_id}/logo")
async def get_competition_logo(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: Optional[AltAuthUserRecord] = Depends(get_optional_alt_identity),
):
    """
    下载/查看竞赛 Logo。

    - **已登录**：须具备 VIEW_COMPETITIONS。
    - **未登录**：仅已发布/已结束竞赛可读（分享链接，无需 Bearer）。
    """
    competition = _get_competition(db, competition_id)
    if identity is None:
        _ensure_competition_shareable(competition)
    else:
        require_permission(identity.role, Permission.VIEW_COMPETITIONS)

    stored = getattr(competition, "logo_path", None)
    if not stored or not str(stored).strip():
        raise HTTPException(status_code=404, detail="No logo for this competition")
    fs_path = _resolve_logo_fs_path(stored)
    if not os.path.isfile(fs_path):
        raise HTTPException(status_code=404, detail="Logo file missing on server")
    mime, _ = mimetypes.guess_type(fs_path)
    try:
        mtime = int(os.path.getmtime(fs_path))
    except OSError:
        mtime = 0
    return FileResponse(
        path=fs_path,
        filename=os.path.basename(fs_path),
        media_type=mime or "application/octet-stream",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            "ETag": f'W/"logo-{competition_id}-{mtime}-{os.path.basename(fs_path)}"',
        },
    )


@router.get("/enrollments/me", response_model=List[MyEnrollmentResponse])
async def my_enrollments(
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查看当前用户报名的所有竞赛（含竞赛详情）"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    enrollments = (
        db.query(CompetitionEnrollment)
        .filter(
            CompetitionEnrollment.student_id == identity.id,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
        )
        .order_by(CompetitionEnrollment.created_at.desc())
        .all()
    )
    results = []
    for e in enrollments:
        comp = db.query(Competition).filter(Competition.id == e.competition_id).first()
        data = MyEnrollmentResponse.model_validate(e)
        if comp:
            data.competition = CompetitionResponse.model_validate(comp)
        results.append(data)
    return results


@router.get("/{competition_id}/teams/my-rejected", response_model=List[MyRejectedTeamItem])
async def my_rejected_teams_in_competition(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """当前用户在本竞赛下被校审驳回的队伍（队长或队员），供学生端顶部一次性提示驳回原因。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    member_team_ids = [
        int(r[0])
        for r in db.query(TeamMember.team_id)
        .filter(TeamMember.user_id == identity.id)
        .all()
    ]
    q = db.query(Team).filter(
        Team.competition_id == competition_id,
        Team.status == TeamStatus.REJECTED,
    )
    if member_team_ids:
        q = q.filter(
            or_(
                Team.captain_id == identity.id,
                Team.id.in_(member_team_ids),
            )
        )
    else:
        q = q.filter(Team.captain_id == identity.id)
    teams = q.order_by(Team.reviewed_at.desc(), Team.id.desc()).all()
    out: List[MyRejectedTeamItem] = []
    for team in teams:
        out.append(
            MyRejectedTeamItem(
                id=int(team.id),
                competition_id=int(team.competition_id),
                name=getattr(team, "name", None),
                division=CompetitionDivision(getattr(team, "division", None) or CompetitionDivision.DEFAULT),
                work_track=(
                    CompetitionWorkTrack(team.work_track)
                    if getattr(team, "work_track", None)
                    else None
                ),
                status=TeamStatus(team.status),
                review_feedback=getattr(team, "review_feedback", None),
                reviewed_at=getattr(team, "reviewed_at", None),
            )
        )
    return out


@router.put("/{competition_id}/publish", response_model=CompetitionResponse)
async def publish_competition(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    competition.status = "published"
    db.commit()
    db.refresh(competition)
    return competition


@router.put("/{competition_id}/lock", response_model=CompetitionResponse)
async def lock_competition(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """停止报名：将竞赛标记为 closed（禁止新报名/新建队伍/加入队伍，不禁止提交与评分等）。"""
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    competition.status = "closed"
    db.commit()
    db.refresh(competition)
    return competition


@router.post("/{competition_id}/exam-papers", response_model=CompetitionExamPapers)
async def publish_competition_exam_paper(
    competition_id: int,
    division: Optional[str] = Form(None),
    work_track: str = Form(..., description="works / software / hardware"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """超级管理员为已发布竞赛按组别+赛道上传/覆盖试卷（上传即发布）。"""
    from app.competition_exam_config import set_exam_paper_track_file

    _require_super_admin_identity(identity)
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    _ensure_competition_published_for_papers(competition)
    div = _normalize_exam_paper_division(competition, division)
    track = _resolve_work_track(work_track)
    rel, original = await _save_exam_paper_upload(file, competition_id, div, track)
    old_path = set_exam_paper_track_file(competition, div, track, rel, original)
    # 不再同步写入旧「按组别」字段，避免本科作品试卷被其它赛道回退误读
    competition.updated_at = utc_now()
    db.commit()
    db.refresh(competition)
    if old_path:
        _delete_stored_exam_paper_file(old_path)
    return _build_exam_papers_meta(competition)


@router.get("/{competition_id}/exam-papers", response_model=CompetitionExamPapers)
async def get_competition_exam_papers(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查询竞赛各组别/赛道试卷是否已发布（不返回本地路径）。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    return _build_exam_papers_meta(competition)


@router.put(
    "/{competition_id}/submission-question-config",
    response_model=CompetitionSubmissionQuestionConfigByTrack,
)
async def put_submission_question_config(
    competition_id: int,
    body: CompetitionSubmissionQuestionConfigByTrack,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """超级管理员按软件/硬件赛道分别配置分题：题数、题名、每题与总分分数区间。"""
    from app.competition_exam_config import dumps_json, normalize_submission_question_config_by_track

    _require_super_admin_identity(identity)
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    _ensure_competition_published_for_papers(competition)
    normalized = normalize_submission_question_config_by_track(body.model_dump())
    competition.submission_question_config = dumps_json(normalized)
    competition.updated_at = utc_now()
    db.commit()
    db.refresh(competition)
    return CompetitionSubmissionQuestionConfigByTrack.model_validate(normalized)


@router.get(
    "/{competition_id}/submission-question-config",
    response_model=CompetitionSubmissionQuestionConfigByTrack,
)
async def get_submission_question_config(
    competition_id: int,
    work_track: Optional[str] = Query(
        None, description="可选：software / hardware，仅返回该赛道配置"
    ),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    from app.competition_exam_config import (
        get_competition_question_config,
        get_competition_question_config_by_track,
    )

    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    by_track = get_competition_question_config_by_track(competition)
    track = (work_track or "").strip().lower()
    if track in ("software", "hardware"):
        # 仍返回完整 by-track，便于前端统一；单赛道也可只取对应字段
        return CompetitionSubmissionQuestionConfigByTrack.model_validate(by_track)
    return CompetitionSubmissionQuestionConfigByTrack.model_validate(by_track)


@router.get("/{competition_id}/exam-papers/download")
async def download_competition_exam_paper(
    competition_id: int,
    division: Optional[str] = Query(None, description="undergraduate / vocational"),
    work_track: Optional[str] = Query(None, description="works / software / hardware"),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    下载已发布试卷。
    仅学生按有效报名的组别（本科/高职）与赛道下载；指导老师不可下载。
    """
    from app.competition_exam_config import get_exam_paper_track_file

    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    _ensure_competition_published_for_papers(competition)
    div = _normalize_exam_paper_division(competition, division)
    track = _resolve_identity_work_track_for_paper(db, competition, identity, div, work_track)
    if not _can_download_exam_paper(db, competition, identity, div, track):
        raise HTTPException(
            status_code=403,
            detail=(
                "无权下载试卷：仅学生在对应赛道有效报名且组队已校审通过后可下载；"
                "指导老师不可下载试卷。"
            ),
        )
    path, filename = get_exam_paper_track_file(competition, div, track)
    # 兼容旧「不分学历组别」default 槽位（同一赛道）
    if not path:
        path, filename = get_exam_paper_track_file(competition, "default", track)
    # 仅当未指定赛道时，才允许回退到旧「按组别」槽位；有赛道时禁止串用其它赛道文件
    if not path and not track:
        path, filename = _exam_paper_path_and_filename(competition, div)
    if not path and not track:
        path, filename = _exam_paper_path_and_filename(competition, "default")
    if not path:
        raise HTTPException(
            status_code=404,
            detail=f"Exam paper not published for division={div}, work_track={track}",
        )
    fs_path = _resolve_exam_paper_fs_path(path)
    if not os.path.isfile(fs_path):
        raise HTTPException(status_code=404, detail="Exam paper file missing")
    download_name = filename or os.path.basename(fs_path)
    media_type, _ = mimetypes.guess_type(download_name)
    return FileResponse(
        path=fs_path,
        filename=download_name,
        media_type=media_type or "application/octet-stream",
    )


@router.patch("/admin/alt-users/{target_user_id}", response_model=AltUserAdminUpdateResult)
async def competition_admin_patch_alt_user(
    target_user_id: int,
    body: AltUserAdminPatch,
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """管理员调整第二套帐号（含专家审核、角色指派）。"""
    _require_super_admin_identity(identity)
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == target_user_id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Target user not found")
    if body.role is not None:
        r = body.role.value if hasattr(body.role, "value") else str(body.role)
        r = r.strip()
        if r not in {"student", "advisor", "teacher", "expert", "super_admin", "school_admin"}:
            raise HTTPException(status_code=400, detail="Invalid role for alt user")
        row.role = r
    if body.expert_verified is not None:
        if body.expert_verified and _effective_alt_role(row.role) != "expert":
            raise HTTPException(status_code=400, detail="expert_verified only applies to expert role")
        row.expert_verified = bool(body.expert_verified)
        # 通过核验时确保账号可用，并清空未通过原因
        if row.expert_verified:
            row.is_active = True
            row.expert_review_feedback = None
    if body.expert_review_feedback is not None:
        fb = str(body.expert_review_feedback).strip()
        row.expert_review_feedback = fb[:2000] if fb else None
    if body.is_active is not None:
        row.is_active = bool(body.is_active)
        if not row.is_active and _effective_alt_role(row.role) == "expert":
            row.expert_verified = False
        if row.is_active and not bool(getattr(row, "expert_verified", False)):
            # 恢复待审：清空未通过原因
            if body.expert_review_feedback is None:
                row.expert_review_feedback = None
    if body.school_admin_verified is not None:
        if body.school_admin_verified and _effective_alt_role(row.role) != "school_admin":
            raise HTTPException(status_code=400, detail="school_admin_verified only applies to school_admin role")
        row.school_admin_verified = bool(body.school_admin_verified)
        if row.school_admin_verified:
            row.school_admin_application_status = SchoolAdminApplicationStatus.APPROVED.value
        elif _school_admin_application_status(row) == SchoolAdminApplicationStatus.APPROVED:
            row.school_admin_application_status = SchoolAdminApplicationStatus.REJECTED.value
    adb.commit()
    adb.refresh(row)
    return AltUserAdminUpdateResult(
        id=row.id,
        role=_effective_alt_role(row.role),
        expert_verified=bool(getattr(row, "expert_verified", False)),
        school_admin_verified=bool(getattr(row, "school_admin_verified", False)),
        is_active=bool(getattr(row, "is_active", True)),
    )


@router.get("/school-admin/application/me", response_model=SchoolAdminApplicationMeResponse)
async def get_school_admin_application_me(
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管理员查看本人申请状态（未通过审核前不可查看组队校审列表）。"""
    _require_school_admin_role(identity)
    app_status = _school_admin_application_status(identity)
    photo_url = (
        "/api/v1/competitions/school-admin/application/photo"
        if getattr(identity, "school_admin_photo_path", None)
        else None
    )
    return SchoolAdminApplicationMeResponse(
        user_id=identity.id,
        school=identity.school,
        full_name=identity.full_name,
        school_admin_verified=bool(getattr(identity, "school_admin_verified", False)),
        application_status=app_status,
        application_contact=getattr(identity, "school_admin_application_contact", None),
        application_remark=getattr(identity, "school_admin_application_remark", None),
        application_submitted_at=getattr(identity, "school_admin_application_submitted_at", None),
        review_feedback=getattr(identity, "school_admin_review_feedback", None),
        reviewed_at=getattr(identity, "school_admin_reviewed_at", None),
        photo_url=photo_url,
        can_review_teams=_school_admin_can_review_teams(identity),
    )


@router.post("/school-admin/application", response_model=SchoolAdminApplicationMeResponse)
async def submit_school_admin_application(
    photo: UploadFile = File(..., description="校管理员申请照片（必填）"),
    contact: Optional[str] = Form(None, description="联系方式"),
    remark: Optional[str] = Form(None, description="申请备注"),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管理员提交资料申请（须含照片）；待 super_admin 审核通过后方可组队校审。"""
    _require_school_admin_role(identity)
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == identity.id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="User not found")

    app_status = _school_admin_application_status(row)
    if app_status == SchoolAdminApplicationStatus.PENDING:
        raise HTTPException(status_code=400, detail="Application already pending review")
    if _school_admin_can_review_teams(row):
        raise HTTPException(status_code=400, detail="School admin already verified")

    old_photo = getattr(row, "school_admin_photo_path", None)
    rel_path = await _save_school_admin_photo(photo, row.id)
    if old_photo and old_photo != rel_path:
        _delete_school_admin_photo(old_photo)

    row.school_admin_photo_path = rel_path
    row.school_admin_application_contact = (contact or "").strip() or None
    row.school_admin_application_remark = (remark or "").strip() or None
    row.school_admin_application_status = SchoolAdminApplicationStatus.PENDING.value
    row.school_admin_application_submitted_at = utc_now()
    row.school_admin_verified = False
    row.school_admin_review_feedback = None
    row.school_admin_reviewed_at = None
    row.school_admin_reviewed_by_id = None
    adb.commit()
    adb.refresh(row)

    return SchoolAdminApplicationMeResponse(
        user_id=row.id,
        school=row.school,
        full_name=row.full_name,
        school_admin_verified=False,
        application_status=SchoolAdminApplicationStatus.PENDING,
        application_contact=row.school_admin_application_contact,
        application_remark=row.school_admin_application_remark,
        application_submitted_at=row.school_admin_application_submitted_at,
        review_feedback=None,
        reviewed_at=None,
        photo_url="/api/v1/competitions/school-admin/application/photo",
        can_review_teams=False,
    )


@router.get("/school-admin/application/photo")
async def download_school_admin_application_photo_self(
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管理员下载本人申请照片。"""
    _require_school_admin_role(identity)
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == identity.id).first()
    if row is None or not getattr(row, "school_admin_photo_path", None):
        raise HTTPException(status_code=404, detail="Photo not found")
    fs_path = _resolve_school_admin_photo_fs_path(row.school_admin_photo_path)
    ext = os.path.splitext(fs_path)[1].lower()
    media = mimetypes.guess_type(fs_path)[0] or "application/octet-stream"
    return FileResponse(path=fs_path, filename=f"school_admin_{row.id}{ext}", media_type=media)


@router.get("/admin/school-admin-applications", response_model=SchoolAdminApplicationListResponse)
async def list_school_admin_applications(
    status_filter: Optional[str] = Query(
        "all",
        alias="status",
        description="申请状态：pending / approved / rejected / not_submitted / all",
    ),
    keyword: Optional[str] = Query(
        None,
        description="按用户ID / 用户名 / 姓名 / 学校模糊搜索",
    ),
    school: Optional[str] = Query(
        None,
        description="兼容旧参数：等同于 keyword（按学校等字段搜索）",
    ),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """超级管理员查看校管理员资料申请列表。"""
    _require_super_admin_identity(identity)
    q = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.role == "school_admin")
    if status_filter and status_filter != "all":
        if status_filter == SchoolAdminApplicationStatus.NOT_SUBMITTED.value:
            q = q.filter(
                or_(
                    AltAuthUserRecord.school_admin_application_status.is_(None),
                    AltAuthUserRecord.school_admin_application_status == "",
                )
            )
        else:
            q = q.filter(AltAuthUserRecord.school_admin_application_status == status_filter)
    search_kw = (keyword or school or "").strip()
    rows = q.order_by(
        AltAuthUserRecord.school_admin_application_submitted_at.desc(),
        AltAuthUserRecord.id.desc(),
    ).all()
    if search_kw:
        kw_cf = search_kw.casefold()

        def _row_matches(row: AltAuthUserRecord) -> bool:
            fields = [
                str(row.id),
                (row.username or "").strip(),
                (row.full_name or "").strip(),
                _normalize_school_name(getattr(row, "school", None)),
            ]
            return any(kw_cf in f.casefold() for f in fields if f)

        rows = [r for r in rows if _row_matches(r)]

    items: List[SchoolAdminApplicationListItem] = []
    for row in rows:
        app_status = _school_admin_application_status(row)
        items.append(
            SchoolAdminApplicationListItem(
                user_id=row.id,
                username=row.username or "",
                email=row.email,
                full_name=row.full_name,
                school=row.school,
                application_status=app_status,
                application_contact=getattr(row, "school_admin_application_contact", None),
                application_remark=getattr(row, "school_admin_application_remark", None),
                application_submitted_at=getattr(row, "school_admin_application_submitted_at", None),
                school_admin_verified=bool(getattr(row, "school_admin_verified", False)),
                review_feedback=getattr(row, "school_admin_review_feedback", None),
                reviewed_at=getattr(row, "school_admin_reviewed_at", None),
                photo_url=(
                    f"/api/v1/competitions/admin/school-admin-applications/{row.id}/photo"
                    if getattr(row, "school_admin_photo_path", None)
                    else None
                ),
            )
        )
    return SchoolAdminApplicationListResponse(total=len(items), items=items)


@router.get("/admin/school-admin-applications/{user_id}/photo")
async def download_school_admin_application_photo_admin(
    user_id: int,
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """超级管理员查看校管申请照片。"""
    _require_super_admin_identity(identity)
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == user_id).first()
    if row is None or _effective_alt_role(row.role) != "school_admin":
        raise HTTPException(status_code=404, detail="School admin user not found")
    if not getattr(row, "school_admin_photo_path", None):
        raise HTTPException(status_code=404, detail="Photo not found")
    fs_path = _resolve_school_admin_photo_fs_path(row.school_admin_photo_path)
    ext = os.path.splitext(fs_path)[1].lower()
    media = mimetypes.guess_type(fs_path)[0] or "application/octet-stream"
    return FileResponse(path=fs_path, filename=f"school_admin_{row.id}{ext}", media_type=media)


@router.put(
    "/admin/school-admin-applications/{user_id}",
    response_model=SchoolAdminApplicationReviewResult,
)
async def review_school_admin_application(
    user_id: int,
    body: SchoolAdminApplicationReviewRequest,
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    超级管理员审核校管理员资料申请。
    - approve：待审核 → 已通过
    - reject：待审核 / 已通过 → 已驳回（撤销校审权限）
    - reset_pending：已通过 → 待审核（撤销校审权限，重新进入待审）
    """
    _require_super_admin_identity(identity)
    row = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == user_id).first()
    if row is None or _effective_alt_role(row.role) != "school_admin":
        raise HTTPException(status_code=404, detail="School admin user not found")

    current = _school_admin_application_status(row)
    action = body.action.value if hasattr(body.action, "value") else str(body.action)
    feedback = (body.feedback or "").strip() or None
    now = utc_now()

    if action == "approve":
        if current != SchoolAdminApplicationStatus.PENDING:
            raise HTTPException(status_code=400, detail="Application is not pending review")
        if not getattr(row, "school_admin_photo_path", None):
            raise HTTPException(status_code=400, detail="Application has no photo")
        row.school_admin_verified = True
        row.school_admin_application_status = SchoolAdminApplicationStatus.APPROVED.value
    elif action == "reject":
        if current not in (
            SchoolAdminApplicationStatus.PENDING,
            SchoolAdminApplicationStatus.APPROVED,
        ):
            raise HTTPException(
                status_code=400,
                detail="仅待审核或已通过的校管申请可驳回",
            )
        if current == SchoolAdminApplicationStatus.PENDING and not getattr(
            row, "school_admin_photo_path", None
        ):
            raise HTTPException(status_code=400, detail="Application has no photo")
        row.school_admin_verified = False
        row.school_admin_application_status = SchoolAdminApplicationStatus.REJECTED.value
    elif action == "reset_pending":
        if current != SchoolAdminApplicationStatus.APPROVED:
            raise HTTPException(
                status_code=400,
                detail="仅已通过的校管可改回待审核",
            )
        # 改回待审核不强制照片：历史账号可能无照片字段，仍应允许超管撤销已通过状态
        row.school_admin_verified = False
        row.school_admin_application_status = SchoolAdminApplicationStatus.PENDING.value
    else:
        raise HTTPException(
            status_code=400,
            detail="action must be approve, reject, or reset_pending",
        )

    row.school_admin_review_feedback = feedback
    row.school_admin_reviewed_at = now
    row.school_admin_reviewed_by_id = identity.id
    adb.commit()
    adb.refresh(row)

    return SchoolAdminApplicationReviewResult(
        user_id=row.id,
        school_admin_verified=bool(row.school_admin_verified),
        application_status=_school_admin_application_status(row),
        review_feedback=row.school_admin_review_feedback,
        reviewed_at=row.school_admin_reviewed_at,
    )


def _normalize_team_review_work_track_filter(raw: Optional[str]) -> Optional[str]:
    """校审列表赛道筛选：合法值返回 works/software/hardware；空/all 返回 None；非法抛 400。"""
    track = (raw or "").strip().lower()
    if not track or track in ("all", "*"):
        return None
    if track not in ("works", "software", "hardware"):
        raise HTTPException(
            status_code=400,
            detail="work_track must be works, software, hardware, or all",
        )
    return track


def _team_matches_review_keyword(
    team: Team,
    keyword: str,
    users_by_id: dict,
) -> bool:
    """学校 / 队名 / 第一·第二指导老师用户名 模糊匹配（不区分大小写）。"""
    kw = (keyword or "").strip().casefold()
    if not kw:
        return True
    school_name = _normalize_school_name(getattr(team, "school", None)).casefold()
    if kw in school_name:
        return True
    team_name = (getattr(team, "name", None) or "").strip().casefold()
    if team_name and kw in team_name:
        return True
    advisor_ids = []
    if getattr(team, "created_by_advisor_id", None):
        advisor_ids.append(int(team.created_by_advisor_id))
    if getattr(team, "second_advisor_id", None):
        advisor_ids.append(int(team.second_advisor_id))
    for aid in advisor_ids:
        u = users_by_id.get(aid) if users_by_id else None
        uname = ((u.username or "") if u else "").strip().casefold()
        if uname and kw in uname:
            return True
    return False


@router.get("/school-admin/teams", response_model=SchoolAdminTeamReviewListResponse)
async def list_school_admin_teams(
    status_filter: Optional[str] = Query(
        None,
        alias="status",
        description="队伍状态筛选：pending_school_review / active / rejected；不传或 all 表示全部",
    ),
    username: Optional[str] = Query(
        None,
        description="按队长/队员/指导老师用户名模糊搜索（兼容旧参数）",
    ),
    school: Optional[str] = Query(
        None,
        description="模糊搜索：学校名 / 队伍名 / 指导老师用户名",
    ),
    work_track: Optional[str] = Query(
        None,
        description="赛道筛选：works / software / hardware；不传或 all 表示全部",
    ),
    competition_id: Optional[int] = Query(None, description="按竞赛 id 筛选"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    校管理员查看本校组队人员审核列表。
    列表字段：竞赛名、开始/结束时间、学校、指导老师、队伍名、队长、队员、组别、赛道、状态。
    """
    _require_school_admin_identity(identity)
    admin_school = _normalize_school_name(identity.school)

    q = (
        db.query(Team)
        .join(Competition, Team.competition_id == Competition.id)
        .options(joinedload(Team.members))
        .filter(Team.school.isnot(None))
    )
    status_raw = (status_filter or "").strip()
    status_norm = status_raw.lower()
    if status_norm and status_norm not in ("all", "*"):
        allowed = {
            TeamStatus.PENDING_SCHOOL_REVIEW,
            TeamStatus.ACTIVE,
            TeamStatus.REJECTED,
        }
        if status_norm not in allowed:
            raise HTTPException(
                status_code=400,
                detail="status must be pending_school_review, active, rejected, or all",
            )
        q = q.filter(Team.status == status_norm)
    track_filter = _normalize_team_review_work_track_filter(work_track)
    if track_filter:
        q = q.filter(Team.work_track == track_filter)
    if competition_id is not None:
        q = q.filter(Team.competition_id == competition_id)

    teams = q.order_by(Team.created_at.desc(), Team.id.desc()).all()
    teams = [t for t in teams if _team_matches_school(t, admin_school)]

    all_uids: set[int] = set()
    for t in teams:
        all_uids.add(t.captain_id)
        if t.created_by_advisor_id:
            all_uids.add(t.created_by_advisor_id)
        if getattr(t, "second_advisor_id", None):
            all_uids.add(int(t.second_advisor_id))
        for m in t.members:
            all_uids.add(m.user_id)
    users_by_id = _alt_users_by_id(adb, all_uids)

    school_kw = (school or "").strip()
    if school_kw:
        teams = [t for t in teams if _team_matches_review_keyword(t, school_kw, users_by_id)]

    username_kw = (username or "").strip().lower()
    if username_kw:
        matched: List[Team] = []
        for t in teams:
            uids = {t.captain_id}
            if t.created_by_advisor_id:
                uids.add(t.created_by_advisor_id)
            if getattr(t, "second_advisor_id", None):
                uids.add(int(t.second_advisor_id))
            for m in t.members:
                uids.add(m.user_id)
            hit = False
            for uid in uids:
                u = users_by_id.get(uid)
                uname = ((u.username or "") if u else "").strip().lower()
                if uname and username_kw in uname:
                    hit = True
                    break
            if hit:
                matched.append(t)
        teams = matched

    submitted_ids = _team_ids_with_submitted_work(db, teams)
    items = [
        _build_school_admin_team_item(
            t,
            t.competition,
            users_by_id,
            work_submitted=(int(t.id) in submitted_ids),
        )
        for t in teams
    ]
    return SchoolAdminTeamReviewListResponse(total=len(items), items=items)


@router.put("/teams/{team_id}/school-review", response_model=TeamSchoolReviewResult)
async def school_review_team(
    team_id: int,
    body: TeamSchoolReviewRequest,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管（本校）或超管审核队伍：通过或驳回。超管可将已通过改回待校审或驳回。双方共享同一 Team.status。"""
    role = _effective_alt_role(identity.role)
    is_super = role == "super_admin"
    is_school = role == "school_admin"
    if not is_super and not is_school:
        raise HTTPException(status_code=403, detail="Only school_admin or super_admin can review teams")
    if is_school:
        _require_school_admin_identity(identity)
    else:
        _require_super_admin_identity(identity)

    team = (
        db.query(Team)
        .options(joinedload(Team.members))
        .filter(Team.id == team_id)
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    if is_school:
        admin_school = _normalize_school_name(identity.school)
        if not _team_matches_school(team, admin_school):
            raise HTTPException(status_code=403, detail="Team does not belong to your school")

    action = body.action.value if hasattr(body.action, "value") else str(body.action)
    feedback = (body.feedback or "").strip() or None
    now = utc_now()

    if action == "approve":
        if team.status != TeamStatus.PENDING_SCHOOL_REVIEW:
            raise HTTPException(status_code=400, detail="Team is not pending school review")
        _assert_team_ready_for_school_approve(team)
        team.status = TeamStatus.ACTIVE
        team.reviewed_by_id = identity.id
        team.reviewed_at = now
        team.review_feedback = feedback
    elif action == "reject":
        # 校管仅可驳回待审；超管还可将已通过改为驳回
        if team.status == TeamStatus.PENDING_SCHOOL_REVIEW:
            pass
        elif is_super and team.status == TeamStatus.ACTIVE:
            pass
        else:
            raise HTTPException(
                status_code=400,
                detail="Team cannot be rejected in current status"
                if is_super
                else "Team is not pending school review",
            )
        team.status = TeamStatus.REJECTED
        team.reviewed_by_id = identity.id
        team.reviewed_at = now
        team.review_feedback = feedback
        enrollments = (
            db.query(CompetitionEnrollment)
            .filter(
                CompetitionEnrollment.competition_id == team.competition_id,
                CompetitionEnrollment.team_id == team.id,
                CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
            )
            .all()
        )
        for enr in enrollments:
            enr.status = CompetitionEnrollmentStatus.WITHDRAWN
            # 释放 uq_competition_student_work_track，否则后续其他队伍无法改到同赛道
            enr.work_track = None
        _cancel_pending_team_side_effects(db, team, now=now)
    elif action == "reset_pending":
        if not is_super:
            raise HTTPException(status_code=403, detail="Only super_admin can reset team to pending")
        if team.status != TeamStatus.ACTIVE:
            raise HTTPException(status_code=400, detail="Only approved teams can be reset to pending")
        team.status = TeamStatus.PENDING_SCHOOL_REVIEW
        team.reviewed_by_id = None
        team.reviewed_at = None
        team.review_feedback = feedback
    else:
        raise HTTPException(
            status_code=400,
            detail="action must be approve, reject, or reset_pending",
        )

    db.commit()
    db.refresh(team)
    return TeamSchoolReviewResult(
        team_id=team.id,
        status=team.status,
        reviewed_at=team.reviewed_at,
        review_feedback=team.review_feedback,
    )


def _resolve_advisor_row_from_refs(
    adb: Session,
    *,
    advisor_username: Optional[str] = None,
    advisor_id: Optional[int] = None,
    advisor_name: Optional[str] = None,
    label: str = "指导老师",
) -> Optional[AltAuthUserRecord]:
    username = (advisor_username or "").strip()
    name = (advisor_name or "").strip()
    if username:
        adv_row = _resolve_alt_user_by_username(adb, username, label=label)
    elif advisor_id is not None:
        adv_row = _ensure_alt_principal_is_advisor(adb, int(advisor_id))
    elif name:
        adv_row = _resolve_advisor_ref(adb, name)
    else:
        return None
    if _effective_alt_role(adv_row.role) not in {"advisor", "teacher"}:
        raise HTTPException(status_code=400, detail=f"{label}须为指导老师账号")
    return adv_row


@router.put("/teams/{team_id}/advisor", response_model=SchoolAdminSetTeamAdvisorResult)
async def set_team_advisor(
    team_id: int,
    body: SchoolAdminSetTeamAdvisorRequest,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管（本校）或超管为队伍添加/更换第一、第二指导老师。"""
    role = _effective_alt_role(identity.role)
    is_super = role == "super_admin"
    is_school = role == "school_admin"
    if not is_super and not is_school:
        raise HTTPException(status_code=403, detail="Only school_admin or super_admin can set team advisor")
    if is_school:
        _require_school_admin_identity(identity)
    else:
        _require_super_admin_identity(identity)

    team = db.query(Team).filter(Team.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    if is_school:
        admin_school = _normalize_school_name(identity.school)
        if not _team_matches_school(team, admin_school):
            raise HTTPException(status_code=403, detail="Team does not belong to your school")
    if team.status == TeamStatus.REJECTED:
        raise HTTPException(status_code=400, detail="Cannot set advisor on a rejected team")

    _assert_team_school_meta_mutable(db, team.competition_id, team.id)

    primary_row = _resolve_advisor_row_from_refs(
        adb,
        advisor_username=body.advisor_username,
        advisor_id=body.advisor_id,
        advisor_name=body.advisor_name,
        label="第一指导老师",
    )
    secondary_row = _resolve_advisor_row_from_refs(
        adb,
        advisor_username=body.second_advisor_username,
        advisor_id=body.second_advisor_id,
        advisor_name=body.second_advisor_name,
        label="第二指导老师",
    )
    clear_first = bool(getattr(body, "clear_first_advisor", False))
    clear_second = bool(body.clear_second_advisor)
    if (
        primary_row is None
        and secondary_row is None
        and not clear_first
        and not clear_second
    ):
        raise HTTPException(status_code=400, detail="请填写或勾选清除第一/第二指导老师")

    # 第一栏固定为第一指导老师，第二栏固定为第二指导老师
    if clear_first:
        first_id, first_name = None, None
    elif primary_row is not None:
        first_id = int(primary_row.id)
        first_name = _display_user_name(primary_row, primary_row.id)
    else:
        first_id = team.created_by_advisor_id
        first_name = team.advisor_name

    if clear_second:
        second_id, second_name = None, None
    elif secondary_row is not None:
        second_id = int(secondary_row.id)
        second_name = _display_user_name(secondary_row, secondary_row.id)
    else:
        second_id = getattr(team, "second_advisor_id", None)
        second_name = getattr(team, "second_advisor_name", None)

    if first_id is not None and second_id is not None and int(first_id) == int(second_id):
        raise HTTPException(status_code=400, detail="第一与第二指导老师不能为同一人")

    division = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    work_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    if not work_track:
        raise HTTPException(status_code=400, detail="队伍未设置赛道，无法校验指导老师额度")
    _assert_team_advisors_quota(
        db,
        competition_id=team.competition_id,
        division=division,
        work_track=work_track,
        first_advisor_id=int(first_id) if first_id is not None else None,
        second_advisor_id=int(second_id) if second_id is not None else None,
        exclude_team_id=team.id,
    )

    if clear_first or primary_row is not None:
        team.created_by_advisor_id = int(first_id) if first_id is not None else None
        team.advisor_name = first_name
    if clear_second or secondary_row is not None:
        team.second_advisor_id = int(second_id) if second_id is not None else None
        team.second_advisor_name = second_name
    db.commit()
    db.refresh(team)
    return SchoolAdminSetTeamAdvisorResult(
        team_id=team.id,
        advisor_id=team.created_by_advisor_id,
        advisor_name=team.advisor_name,
        second_advisor_id=team.second_advisor_id,
        second_advisor_name=team.second_advisor_name,
    )


@router.put("/teams/{team_id}/second-advisor", response_model=TeamSetSecondAdvisorResult)
async def set_team_second_advisor(
    team_id: int,
    body: TeamSetSecondAdvisorRequest,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """本队指导老师可添加/更换另一职位的指导老师（第一或第二）。"""
    role = _effective_alt_role(identity.role)
    if role not in {"advisor", "teacher", "school_admin", "super_admin"}:
        raise HTTPException(status_code=403, detail="Only advisor/teacher or admin can set peer advisor")

    team = db.query(Team).filter(Team.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    if team.status == TeamStatus.REJECTED:
        raise HTTPException(status_code=400, detail="Cannot set advisor on a rejected team")

    if role in {"advisor", "teacher"}:
        if not _team_advisor_managed(team, identity.id):
            raise HTTPException(status_code=403, detail="Only this team's advisors may set peer advisor")
    elif role == "school_admin":
        _require_school_admin_identity(identity)
        admin_school = _normalize_school_name(identity.school)
        if not _team_matches_school(team, admin_school):
            raise HTTPException(status_code=403, detail="Team does not belong to your school")
    else:
        _require_super_admin_identity(identity)

    _assert_team_school_meta_mutable(db, team.competition_id, team.id)

    target = str(getattr(body, "target_role", None) or "second").strip().lower()
    if target not in ("first", "second"):
        raise HTTPException(status_code=400, detail="target_role must be first or second")
    target_label = "第一指导老师" if target == "first" else "第二指导老师"

    if body.clear:
        if target == "first":
            team.created_by_advisor_id = None
            team.advisor_name = None
        else:
            team.second_advisor_id = None
            team.second_advisor_name = None
        db.commit()
        db.refresh(team)
        return TeamSetSecondAdvisorResult(
            team_id=team.id,
            advisor_id=team.created_by_advisor_id,
            advisor_name=team.advisor_name,
            second_advisor_id=team.second_advisor_id,
            second_advisor_name=team.second_advisor_name,
        )

    adv_row = _resolve_advisor_row_from_refs(
        adb,
        advisor_username=getattr(body, "second_advisor_username", None),
        advisor_id=body.second_advisor_id,
        advisor_name=body.second_advisor_name,
        label=target_label,
    )
    if adv_row is None:
        raise HTTPException(status_code=400, detail=f"请填写{target_label}用户名")

    new_id = int(adv_row.id)
    new_name = _display_user_name(adv_row, adv_row.id)
    if target == "first":
        other_id = getattr(team, "second_advisor_id", None)
        if other_id is not None and int(other_id) == new_id:
            raise HTTPException(status_code=400, detail="第一与第二指导老师不能为同一人")
        first_id, second_id = new_id, (
            int(other_id) if other_id is not None else None
        )
    else:
        other_id = team.created_by_advisor_id
        if other_id is not None and int(other_id) == new_id:
            raise HTTPException(status_code=400, detail="第二指导老师不能与第一指导老师相同")
        first_id, second_id = (
            int(other_id) if other_id is not None else None
        ), new_id

    division = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    work_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    if not work_track:
        raise HTTPException(status_code=400, detail="队伍未设置赛道，无法校验指导老师额度")
    _assert_team_advisors_quota(
        db,
        competition_id=team.competition_id,
        division=division,
        work_track=work_track,
        first_advisor_id=first_id,
        second_advisor_id=second_id,
        exclude_team_id=team.id,
    )

    if target == "first":
        team.created_by_advisor_id = new_id
        team.advisor_name = new_name
    else:
        team.second_advisor_id = new_id
        team.second_advisor_name = new_name
    db.commit()
    db.refresh(team)
    return TeamSetSecondAdvisorResult(
        team_id=team.id,
        advisor_id=team.created_by_advisor_id,
        advisor_name=team.advisor_name,
        second_advisor_id=team.second_advisor_id,
        second_advisor_name=team.second_advisor_name,
    )


def _free_stale_work_track_enrollment(
    db: Session,
    *,
    competition_id: int,
    student_id: int,
    work_track: str,
    keep_enrollment_id: int,
) -> None:
    """
    释放同竞赛+学生+赛道上已失效的报名槽位（撤回/驳回队），
    避免 uq_competition_student_work_track 阻止改赛道。
    """
    track = _normalize_optional_work_track(work_track)
    if not track:
        return
    rows = (
        db.query(CompetitionEnrollment)
        .filter(
            CompetitionEnrollment.competition_id == competition_id,
            CompetitionEnrollment.student_id == student_id,
            CompetitionEnrollment.work_track == track,
            CompetitionEnrollment.id != keep_enrollment_id,
        )
        .all()
    )
    for other in rows:
        if other.status == CompetitionEnrollmentStatus.WITHDRAWN:
            db.delete(other)
            continue
        if other.status != CompetitionEnrollmentStatus.ENROLLED:
            db.delete(other)
            continue
        other_team = (
            db.query(Team).filter(Team.id == other.team_id).first()
            if other.team_id is not None
            else None
        )
        if other_team is not None and other_team.status in (
            TeamStatus.REJECTED,
            TeamStatus.DISBANDED,
        ):
            other.status = CompetitionEnrollmentStatus.WITHDRAWN
            other.work_track = None
            continue
        raise HTTPException(
            status_code=400,
            detail=(
                f"队员 {student_id} 已在赛道 {track} 有有效报名"
                + (f"（队伍 {other.team_id}）" if other.team_id else "")
                + "，无法将本队改到该赛道"
            ),
        )
    db.flush()


def _sync_team_division_and_work_track(
    db: Session, team: Team, division: str, work_track: str
) -> None:
    team.division = division
    team.work_track = work_track
    enrollments = (
        db.query(CompetitionEnrollment)
        .filter(CompetitionEnrollment.team_id == team.id)
        .all()
    )
    for row in enrollments:
        row.division = division
        old_track = _normalize_optional_work_track(row.work_track)
        if old_track != work_track:
            _free_stale_work_track_enrollment(
                db,
                competition_id=team.competition_id,
                student_id=row.student_id,
                work_track=work_track,
                keep_enrollment_id=row.id,
            )
        row.work_track = work_track
    submissions = db.query(Submission).filter(Submission.team_id == team.id).all()
    for row in submissions:
        row.division = division


@router.put("/teams/{team_id}/division-track", response_model=SchoolAdminSetTeamDivisionTrackResult)
async def set_team_division_track(
    team_id: int,
    body: SchoolAdminSetTeamDivisionTrackRequest,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管（本校）或超管修改队伍组别与赛道，并同步该队报名记录。"""
    role = _effective_alt_role(identity.role)
    is_super = role == "super_admin"
    is_school = role == "school_admin"
    if not is_super and not is_school:
        raise HTTPException(
            status_code=403,
            detail="Only school_admin or super_admin can set team division and work track",
        )
    if is_school:
        _require_school_admin_identity(identity)
    else:
        _require_super_admin_identity(identity)

    team = db.query(Team).filter(Team.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    if is_school:
        admin_school = _normalize_school_name(identity.school)
        if not _team_matches_school(team, admin_school):
            raise HTTPException(status_code=403, detail="Team does not belong to your school")
    if team.status == TeamStatus.REJECTED:
        raise HTTPException(status_code=400, detail="Cannot update division on a rejected team")

    _assert_team_school_meta_mutable(db, team.competition_id, team.id)

    competition = _get_competition(db, team.competition_id)
    division = _resolve_enrollment_division(competition, body.division)
    work_track = _resolve_work_track(body.work_track)
    _assert_team_advisors_quota(
        db,
        competition_id=team.competition_id,
        division=division,
        work_track=work_track,
        first_advisor_id=team.created_by_advisor_id,
        second_advisor_id=getattr(team, "second_advisor_id", None),
        exclude_team_id=team.id,
    )
    try:
        _sync_team_division_and_work_track(db, team, division, work_track)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="赛道修改冲突：队员在目标赛道仍有占用记录，请刷新后重试",
        ) from e
    db.refresh(team)
    return SchoolAdminSetTeamDivisionTrackResult(
        team_id=team.id,
        division=str(team.division or division),
        work_track=str(team.work_track or work_track),
    )


@router.get("/admin/team-reviews", response_model=SchoolAdminTeamReviewListResponse)
async def list_admin_team_reviews(
    status_filter: Optional[str] = Query(
        None,
        alias="status",
        description="队伍状态筛选：pending_school_review / active / rejected；不传或 all 表示全部",
    ),
    school: Optional[str] = Query(
        None,
        description="模糊搜索：学校名 / 队伍名 / 指导老师用户名",
    ),
    keyword: Optional[str] = Query(
        None,
        description="与 school 同义：学校名 / 队伍名 / 指导老师用户名",
    ),
    work_track: Optional[str] = Query(
        None,
        description="赛道筛选：works / software / hardware；不传或 all 表示全部",
    ),
    competition_id: Optional[int] = Query(None, description="按竞赛 id 筛选"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """超级管理员查看全部学校的组队校审列表（与校管共享 Team.status）。"""
    _require_super_admin_identity(identity)

    q = (
        db.query(Team)
        .join(Competition, Team.competition_id == Competition.id)
        .options(joinedload(Team.members))
        .filter(Team.school.isnot(None))
    )
    status_raw = (status_filter or "").strip()
    status_norm = status_raw.lower()
    if status_norm and status_norm not in ("all", "*"):
        allowed = {
            TeamStatus.PENDING_SCHOOL_REVIEW,
            TeamStatus.ACTIVE,
            TeamStatus.REJECTED,
        }
        if status_norm not in allowed:
            raise HTTPException(
                status_code=400,
                detail="status must be pending_school_review, active, rejected, or all",
            )
        q = q.filter(Team.status == status_norm)
    track_filter = _normalize_team_review_work_track_filter(work_track)
    if track_filter:
        q = q.filter(Team.work_track == track_filter)
    if competition_id is not None:
        q = q.filter(Team.competition_id == competition_id)

    teams = q.order_by(Team.created_at.desc(), Team.id.desc()).all()

    all_uids: set[int] = set()
    for t in teams:
        all_uids.add(t.captain_id)
        if t.created_by_advisor_id:
            all_uids.add(t.created_by_advisor_id)
        if getattr(t, "second_advisor_id", None):
            all_uids.add(int(t.second_advisor_id))
        for m in t.members:
            all_uids.add(m.user_id)
    users_by_id = _alt_users_by_id(adb, all_uids)

    school_kw = (school or keyword or "").strip()
    if school_kw:
        teams = [t for t in teams if _team_matches_review_keyword(t, school_kw, users_by_id)]

    submitted_ids = _team_ids_with_submitted_work(db, teams)
    items = [
        _build_school_admin_team_item(
            t,
            t.competition,
            users_by_id,
            work_submitted=(int(t.id) in submitted_ids),
        )
        for t in teams
    ]
    return SchoolAdminTeamReviewListResponse(total=len(items), items=items)


def _assert_student_belongs_to_school(
    student: AltAuthUserRecord, admin_school: str, *, label: str = "学生"
) -> None:
    stu_school = _normalize_school_name(getattr(student, "school", None))
    if not stu_school:
        raise HTTPException(status_code=400, detail=f"{label}未配置学校")
    if stu_school.casefold() != _normalize_school_name(admin_school).casefold():
        raise HTTPException(status_code=400, detail=f"{label}不属于本校")


def _resolve_alt_user_by_username(adb: Session, username: str, *, label: str = "用户") -> AltAuthUserRecord:
    key = (username or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail=f"{label}用户名不能为空")
    row = (
        adb.query(AltAuthUserRecord)
        .filter(func.lower(AltAuthUserRecord.username) == key.lower())
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail=f"{label}「{key}」不存在")
    return row


def _upsert_team_enrollment(
    db: Session,
    *,
    competition: Competition,
    student_id: int,
    team: Team,
    is_captain: bool,
    division: str,
    work_track: str,
) -> None:
    track = _normalize_optional_work_track(work_track) or _resolve_work_track(work_track)
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        student_id,
        track,
        division,
        team_id=team.id,
        allow_same_team=True,
    )
    row_any = _get_enrollment_by_work_track(db, competition.id, student_id, track)
    if row_any and row_any.status == CompetitionEnrollmentStatus.ENROLLED:
        if row_any.team_id != team.id:
            raise HTTPException(
                status_code=400,
                detail=f"学生 {student_id} 已在该竞赛赛道 {track} 报名其他队伍",
            )
        row_any.is_captain = is_captain
        row_any.division = division
        return
    if row_any and row_any.status == CompetitionEnrollmentStatus.WITHDRAWN:
        row_any.team_id = team.id
        row_any.enrollment_scope = CompetitionEnrollmentScope.TEAM
        row_any.is_captain = is_captain
        row_any.status = CompetitionEnrollmentStatus.ENROLLED
        row_any.division = division
        row_any.work_track = track
        return
    db.add(
        CompetitionEnrollment(
            competition_id=competition.id,
            student_id=student_id,
            team_id=team.id,
            enrollment_scope=CompetitionEnrollmentScope.TEAM,
            division=division,
            work_track=track,
            is_captain=is_captain,
            status=CompetitionEnrollmentStatus.ENROLLED,
        )
    )


def _upsert_individual_enrollment(
    db: Session,
    *,
    competition: Competition,
    student_id: int,
    division: str,
    work_track: str,
) -> None:
    track = _normalize_optional_work_track(work_track) or _resolve_work_track(work_track)
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        student_id,
        track,
        division,
        team_id=None,
        allow_same_team=False,
    )
    row_any = _get_enrollment_by_work_track(db, competition.id, student_id, track)
    if row_any and row_any.status == CompetitionEnrollmentStatus.ENROLLED:
        raise HTTPException(
            status_code=400,
            detail=f"学生 {student_id} 已在该竞赛报名赛道 {track}",
        )
    if row_any and row_any.status == CompetitionEnrollmentStatus.WITHDRAWN:
        row_any.team_id = None
        row_any.enrollment_scope = CompetitionEnrollmentScope.INDIVIDUAL
        row_any.is_captain = False
        row_any.status = CompetitionEnrollmentStatus.ENROLLED
        row_any.division = division
        row_any.work_track = track
        return
    db.add(
        CompetitionEnrollment(
            competition_id=competition.id,
            student_id=student_id,
            team_id=None,
            enrollment_scope=CompetitionEnrollmentScope.INDIVIDUAL,
            division=division,
            work_track=track,
            is_captain=False,
            status=CompetitionEnrollmentStatus.ENROLLED,
        )
    )


@router.post(
    "/school-admin/proxy-teams",
    response_model=TeamResponse,
    status_code=status.HTTP_201_CREATED,
)
async def school_admin_proxy_create_team(
    body: SchoolAdminProxyTeamCreate,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管代建队：队伍直接 active（已通过），并为成员写入组队报名。"""
    _require_school_admin_identity(identity)
    admin_school = _normalize_school_name(identity.school)

    competition = _get_competition(db, body.competition_id)
    _ensure_enrollment_open(competition)
    if competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")
    if not competition.allow_team:
        raise HTTPException(status_code=400, detail="Team enrollment not allowed")
    _assert_final_stage_open_create_blocked(competition)

    captain_username = (body.captain_username or "").strip()
    if not captain_username:
        raise HTTPException(status_code=400, detail="captain_username 不能为空")

    seen_names: set[str] = set()
    ordered_usernames: List[str] = []
    for raw in body.member_usernames or []:
        name = (raw or "").strip()
        if not name:
            continue
        key = name.casefold()
        if key not in seen_names:
            seen_names.add(key)
            ordered_usernames.append(name)
    if not ordered_usernames:
        raise HTTPException(status_code=400, detail="member_usernames 至少一名学生")
    if captain_username.casefold() not in seen_names:
        ordered_usernames.insert(0, captain_username)
        seen_names.add(captain_username.casefold())

    team_division = _resolve_enrollment_division(competition, body.division)
    team_work_track = _resolve_work_track(getattr(body, "work_track", None))
    tn = _strip_team_name(body.team_name)
    _assert_team_name_unique_in_competition(db, competition.id, tn)

    ordered_ids: List[int] = []
    username_by_id: dict[int, str] = {}
    for uname in ordered_usernames:
        stu = _resolve_alt_user_by_username(adb, uname, label="学生")
        if _effective_alt_role(stu.role) != "student":
            raise HTTPException(status_code=400, detail=f"「{uname}」须为学生账号")
        _assert_student_belongs_to_school(stu, admin_school, label=f"学生「{uname}」")
        sid = int(stu.id)
        if sid in username_by_id:
            continue
        username_by_id[sid] = uname
        ordered_ids.append(sid)
        if _has_active_enrollment_in_scope(
            db,
            competition.id,
            sid,
            CompetitionEnrollmentScope.TEAM,
            work_track=team_work_track,
        ):
            raise HTTPException(
                status_code=400,
                detail=f"学生「{uname}」已在该竞赛报名赛道 {team_work_track}",
            )
        try:
            _assert_student_can_enroll_work_track(
                db,
                competition.id,
                sid,
                team_work_track,
                team_division,
                team_id=None,
                allow_same_team=False,
            )
        except HTTPException:
            raise
        _assert_final_stage_participant(db, competition, sid, team_id=None)

    captain_row = _resolve_alt_user_by_username(adb, captain_username, label="队长")
    if _effective_alt_role(captain_row.role) != "student":
        raise HTTPException(status_code=400, detail=f"队长「{captain_username}」须为学生账号")
    captain_id = int(captain_row.id)
    if captain_id not in username_by_id:
        raise HTTPException(status_code=400, detail="队长须出现在队员列表中或单独指定")

    created_by_advisor_id = None
    advisor_display_name = None
    advisor_username = (body.advisor_username or "").strip()
    if advisor_username:
        adv_row = _resolve_alt_user_by_username(adb, advisor_username, label="指导老师")
        if _effective_alt_role(adv_row.role) not in {"advisor", "teacher"}:
            raise HTTPException(status_code=400, detail=f"「{advisor_username}」须为指导老师账号")
        created_by_advisor_id = int(adv_row.id)
        advisor_display_name = _display_user_name(adv_row, adv_row.id)

    _assert_team_advisors_quota(
        db,
        competition_id=competition.id,
        division=team_division,
        work_track=team_work_track,
        first_advisor_id=created_by_advisor_id,
        second_advisor_id=None,
    )

    now = utc_now()
    team = Team(
        id=allocate_eight_digit_id(db, Team),
        competition_id=competition.id,
        name=tn,
        captain_id=captain_id,
        created_by_advisor_id=created_by_advisor_id,
        advisor_name=advisor_display_name,
        school=admin_school,
        division=team_division,
        work_track=team_work_track,
        status=TeamStatus.ACTIVE,
        reviewed_by_id=identity.id,
        reviewed_at=now,
        review_feedback="校管代建队，自动通过",
    )
    db.add(team)
    db.flush()

    for sid in ordered_ids:
        ic = sid == captain_id
        db.add(TeamMember(team_id=team.id, user_id=sid, is_captain=ic))
        _upsert_team_enrollment(
            db,
            competition=competition,
            student_id=sid,
            team=team,
            is_captain=ic,
            division=team_division,
            work_track=team_work_track,
        )

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Create team failed: {str(e)}")

    db.refresh(team)
    return team


@router.post(
    "/school-admin/proxy-enroll",
    response_model=SchoolAdminProxyEnrollResult,
    status_code=status.HTTP_201_CREATED,
)
async def school_admin_proxy_enroll(
    body: SchoolAdminProxyEnrollRequest,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """校管代报名：为本校队伍全员补组队报名，或为本校学生补个人报名。"""
    _require_school_admin_identity(identity)
    admin_school = _normalize_school_name(identity.school)

    has_team = body.team_id is not None
    has_student = body.student_id is not None
    if has_team == has_student:
        raise HTTPException(
            status_code=400,
            detail="须且仅传其一：team_id（队伍赛道）或 student_id（个人赛道）",
        )

    competition = _get_competition(db, body.competition_id)
    _ensure_enrollment_open(competition)
    if competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")

    enrolled_ids: List[int] = []

    if has_team:
        if not competition.allow_team:
            raise HTTPException(status_code=400, detail="Team enrollment not allowed")
        team = (
            db.query(Team)
            .options(joinedload(Team.members))
            .filter(Team.id == body.team_id, Team.competition_id == competition.id)
            .first()
        )
        if not team:
            raise HTTPException(status_code=404, detail="Team not found in this competition")
        if not _team_matches_school(team, admin_school):
            raise HTTPException(status_code=403, detail="Team does not belong to your school")
        if team.status == TeamStatus.REJECTED:
            raise HTTPException(status_code=400, detail="已驳回队伍不可补报名")

        team_division = str(getattr(team, "division", None) or "").lower()
        if team_division not in ("undergraduate", "vocational"):
            team_division = _resolve_enrollment_division(competition, body.division)
        team_work_track = str(getattr(team, "work_track", None) or "").lower()
        if team_work_track not in ("works", "software", "hardware"):
            team_work_track = _resolve_work_track(getattr(body, "work_track", None))

        members = list(team.members or [])
        if not members:
            raise HTTPException(status_code=400, detail="队伍暂无队员")

        for m in members:
            sid = int(m.user_id)
            stu = _ensure_alt_principal_is_student(adb, sid)
            _assert_student_belongs_to_school(stu, admin_school, label=f"学生 {sid}")
            _assert_final_stage_participant(db, competition, sid, team_id=team.id)
            before = _get_enrollment_by_scope(
                db, competition.id, sid, CompetitionEnrollmentScope.TEAM
            )
            already = (
                before is not None
                and before.status == CompetitionEnrollmentStatus.ENROLLED
                and before.team_id == team.id
            )
            _upsert_team_enrollment(
                db,
                competition=competition,
                student_id=sid,
                team=team,
                is_captain=bool(m.is_captain),
                division=team_division,
                work_track=team_work_track,
            )
            if not already:
                enrolled_ids.append(sid)

        if not enrolled_ids:
            raise HTTPException(status_code=400, detail="队伍成员均已报名，无需补报名")

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Proxy enroll failed: {str(e)}")

        return SchoolAdminProxyEnrollResult(
            competition_id=competition.id,
            enrolled_count=len(enrolled_ids),
            team_id=team.id,
            student_ids=enrolled_ids,
        )

    # individual
    if not competition.allow_individual:
        raise HTTPException(status_code=400, detail="Individual enrollment not allowed")
    sid = int(body.student_id)
    stu = _ensure_alt_principal_is_student(adb, sid)
    _assert_student_belongs_to_school(stu, admin_school, label=f"学生 {sid}")
    _assert_final_stage_participant(db, competition, sid, team_id=None)
    enroll_division = _resolve_enrollment_division(competition, body.division)
    enroll_work_track = _resolve_work_track(getattr(body, "work_track", None))
    _upsert_individual_enrollment(
        db,
        competition=competition,
        student_id=sid,
        division=enroll_division,
        work_track=enroll_work_track,
    )
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Proxy enroll failed: {str(e)}")

    return SchoolAdminProxyEnrollResult(
        competition_id=competition.id,
        enrolled_count=1,
        team_id=None,
        student_ids=[sid],
    )


@router.post(
    "/{competition_id}/experts/{expert_user_id}",
    status_code=status.HTTP_201_CREATED,
)
async def assign_competition_expert(
    competition_id: int,
    expert_user_id: int,
    body: CompetitionExpertAssignRequest,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    _require_super_admin_identity(identity)
    _get_competition(db, competition_id)
    user = adb.query(AltAuthUserRecord).filter(AltAuthUserRecord.id == expert_user_id).first()
    if user is None:
        raise HTTPException(status_code=404, detail="Expert user not found")
    if _effective_alt_role(user.role) != "expert":
        raise HTTPException(status_code=400, detail="Target role must be expert")
    if not bool(getattr(user, "expert_verified", False)):
        raise HTTPException(status_code=400, detail="Expert must be verified before assignment")

    raw_ids = body.team_ids or []
    team_ids: List[int] = []
    seen: set[int] = set()
    for tid in raw_ids:
        n = int(tid)
        if n in seen:
            continue
        seen.add(n)
        team_ids.append(n)
    if not team_ids:
        raise HTTPException(status_code=400, detail="team_ids must contain at least one team")

    teams = (
        db.query(Team)
        .filter(Team.id.in_(team_ids), Team.competition_id == competition_id)
        .all()
    )
    found = {int(t.id) for t in teams}
    missing = [tid for tid in team_ids if tid not in found]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Teams not found in this competition: {missing}",
        )

    dup = (
        db.query(CompetitionExpertAssignment)
        .filter(
            CompetitionExpertAssignment.competition_id == competition_id,
            CompetitionExpertAssignment.expert_id == expert_user_id,
        )
        .first()
    )
    if not dup:
        db.add(CompetitionExpertAssignment(competition_id=competition_id, expert_id=expert_user_id))

    existing = _assigned_team_ids_for_expert(db, competition_id, expert_user_id)
    added = 0
    for tid in team_ids:
        if tid in existing:
            continue
        db.add(
            CompetitionExpertTeamAssignment(
                competition_id=competition_id,
                expert_id=expert_user_id,
                team_id=tid,
            )
        )
        added += 1

    if not dup and added == 0:
        # 理论上不会：team_ids 非空且都缺失才会，但已校验存在
        pass
    if dup and added == 0:
        raise HTTPException(status_code=400, detail="Expert already assigned to selected teams")

    db.commit()
    return {"ok": True, "added_team_count": added}


@router.delete("/{competition_id}/experts/{expert_user_id}", status_code=status.HTTP_200_OK)
async def unassign_competition_expert(
    competition_id: int,
    expert_user_id: int,
    team_id: Optional[int] = Query(
        None,
        description="若传入则仅取消该队伍指派；不传则撤销该竞赛下全部队伍指派",
    ),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """取消专家指派：可按单支队伍撤销，或撤销该竞赛下全部队伍。"""
    _require_super_admin_identity(identity)
    _get_competition(db, competition_id)

    team_q = db.query(CompetitionExpertTeamAssignment).filter(
        CompetitionExpertTeamAssignment.competition_id == competition_id,
        CompetitionExpertTeamAssignment.expert_id == expert_user_id,
    )
    if team_id is not None:
        team_q = team_q.filter(CompetitionExpertTeamAssignment.team_id == int(team_id))
    team_rows = team_q.all()

    row = (
        db.query(CompetitionExpertAssignment)
        .filter(
            CompetitionExpertAssignment.competition_id == competition_id,
            CompetitionExpertAssignment.expert_id == expert_user_id,
        )
        .first()
    )

    if team_id is not None:
        if not team_rows:
            raise HTTPException(status_code=404, detail="该队伍指派记录不存在")
        for tr in team_rows:
            db.delete(tr)
        remaining = (
            db.query(CompetitionExpertTeamAssignment)
            .filter(
                CompetitionExpertTeamAssignment.competition_id == competition_id,
                CompetitionExpertTeamAssignment.expert_id == expert_user_id,
            )
            .count()
        )
        # 该竞赛下已无队伍时，同步移除竞赛级指派
        if remaining == 0 and row:
            db.delete(row)
        db.commit()
        return {"ok": True, "removed_team_id": int(team_id), "competition_cleared": remaining == 0}

    if not row and not team_rows:
        raise HTTPException(status_code=404, detail="Assignment not found")
    # 未指定 team_id：兼容旧行为，撤销该竞赛下全部
    all_teams = (
        db.query(CompetitionExpertTeamAssignment)
        .filter(
            CompetitionExpertTeamAssignment.competition_id == competition_id,
            CompetitionExpertTeamAssignment.expert_id == expert_user_id,
        )
        .all()
    )
    for tr in all_teams:
        db.delete(tr)
    if row:
        db.delete(row)
    db.commit()
    return {"ok": True}


@router.put("/{competition_id}", response_model=CompetitionResponse)
async def update_competition(
    competition_id: int,
    request: Request,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    修改竞赛（字段可选，只传需要改的）。
    - **application/json**：与原先一致，请求体为 CompetitionUpdate（无文件上传）。
    - **multipart/form-data**：文本字段同名且仅提交需修改项；可选 **`qr_code_image`** 替换二维码、
      **`logo_image`** 替换 Logo（规则同创建）。
    """
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)

    ct = (request.headers.get("content-type") or "").lower()
    qr_upload = None
    qr_upload_undergraduate = None
    qr_upload_vocational = None
    logo_upload = None
    if "application/json" in ct:
        try:
            body = await request.json()
        except Exception as e:
            raise HTTPException(status_code=400, detail="Invalid JSON body") from e
        payload = CompetitionUpdate.model_validate(body)
        update_data = payload.model_dump(exclude_unset=True)
    elif "multipart/form-data" in ct:
        form = await request.form()
        payload = _competition_update_from_form(form)
        update_data = payload.model_dump(exclude_unset=True)
        qr_upload = _pick_qr_upload(form, "qr_code_image")
        qr_upload_undergraduate = _pick_qr_upload(form, "qr_code_image_undergraduate")
        qr_upload_vocational = _pick_qr_upload(form, "qr_code_image_vocational")
        logo_upload = _pick_qr_upload(form, "logo_image") or _pick_qr_upload(form, "logo")
    else:
        raise HTTPException(
            status_code=415,
            detail="Content-Type must be application/json or multipart/form-data",
        )

    stage_mode = update_data.pop("stage_mode", None)
    has_final_start = "final_start_at" in update_data
    has_final_end = "final_end_at" in update_data
    final_start_at = update_data.pop("final_start_at", None)
    final_end_at = update_data.pop("final_end_at", None)

    for field, value in update_data.items():
        if field in ("division_mode", "qr_layout") and value is not None:
            value = getattr(value, "value", value)
        setattr(competition, field, value)

    if stage_mode is not None:
        mode_val = getattr(stage_mode, "value", stage_mode) or "single"
        current_stage = _competition_stage(competition)
        if mode_val == "single":
            if current_stage != CompetitionStage.SINGLE:
                raise HTTPException(
                    status_code=400,
                    detail="已关联初赛/决赛的竞赛不可改回单阶段；如需取消请分别删除对应竞赛",
                )
        elif mode_val == "prelim_final":
            if current_stage == CompetitionStage.FINAL:
                raise HTTPException(
                    status_code=400,
                    detail="请在初赛竞赛上修改赛程类型，或直接编辑本场决赛时间",
                )
            if current_stage == CompetitionStage.SINGLE:
                # 单阶段升级为初赛+决赛：当前场改为初赛，并创建关联决赛
                final_id = allocate_eight_digit_id(db, Competition, used_extra=[competition.id])
                series_id = competition.series_id or competition.id
                base_name = (competition.name or "").strip() or f"竞赛{competition.id}"
                if base_name.endswith("-决赛"):
                    base_name = base_name[: -len("-决赛")]
                if not base_name.endswith("-初赛"):
                    competition.name = f"{base_name}-初赛"
                else:
                    base_name = base_name[: -len("-初赛")]
                final = Competition(
                    id=final_id,
                    name=f"{base_name}-决赛",
                    description=competition.description,
                    rules_text=competition.rules_text,
                    target_audience=getattr(competition, "target_audience", None),
                    contact_name=getattr(competition, "contact_name", None),
                    contact_phone=getattr(competition, "contact_phone", None),
                    location=getattr(competition, "location", None),
                    environment=getattr(competition, "environment", None),
                    status=competition.status or "draft",
                    start_at=final_start_at,
                    end_at=final_end_at,
                    allow_individual=competition.allow_individual,
                    allow_team=competition.allow_team,
                    division_mode=competition.division_mode or "single",
                    qr_layout=competition.qr_layout or "shared",
                    series_id=series_id,
                    stage=CompetitionStage.FINAL,
                    paired_competition_id=competition.id,
                    qr_code_path=competition.qr_code_path,
                    qr_code_path_undergraduate=competition.qr_code_path_undergraduate,
                    qr_code_path_vocational=competition.qr_code_path_vocational,
                    logo_path=getattr(competition, "logo_path", None),
                )
                competition.series_id = series_id
                competition.stage = CompetitionStage.PRELIMINARY
                competition.paired_competition_id = final_id
                db.add(final)
                has_final_start = False
                has_final_end = False

    # 初赛：同步关联决赛时间
    if _is_preliminary_stage(competition) and (has_final_start or has_final_end):
        paired_id = getattr(competition, "paired_competition_id", None)
        if paired_id is None:
            raise HTTPException(status_code=400, detail="该初赛未关联决赛，无法更新决赛时间")
        final = db.query(Competition).filter(Competition.id == int(paired_id)).first()
        if not final or not _is_final_stage(final):
            raise HTTPException(status_code=400, detail="关联决赛不存在或阶段无效")
        if has_final_start:
            final.start_at = final_start_at
        if has_final_end:
            final.end_at = final_end_at

    # 初赛/决赛创建时共用同一套二维码/Logo 路径；替换后须同步关联场次，再删旧文件，避免关联场仍指向已删除文件
    old_paths_to_delete = []
    old_logo_paths_to_delete = []
    qr_changed = False
    logo_changed = False
    if qr_upload is not None:
        old_path = competition.qr_code_path
        competition.qr_code_path = await _save_qr_code_upload(qr_upload, competition_id)
        if old_path and old_path != competition.qr_code_path:
            old_paths_to_delete.append(old_path)
        qr_changed = True
    if qr_upload_undergraduate is not None:
        old_path = competition.qr_code_path_undergraduate
        competition.qr_code_path_undergraduate = await _save_qr_code_upload(
            qr_upload_undergraduate, competition_id
        )
        if old_path and old_path != competition.qr_code_path_undergraduate:
            old_paths_to_delete.append(old_path)
        qr_changed = True
    if qr_upload_vocational is not None:
        old_path = competition.qr_code_path_vocational
        competition.qr_code_path_vocational = await _save_qr_code_upload(
            qr_upload_vocational, competition_id
        )
        if old_path and old_path != competition.qr_code_path_vocational:
            old_paths_to_delete.append(old_path)
        qr_changed = True
    if logo_upload is not None:
        old_logo = getattr(competition, "logo_path", None)
        competition.logo_path = await _save_logo_upload(logo_upload, competition_id)
        if old_logo and old_logo != competition.logo_path:
            old_logo_paths_to_delete.append(old_logo)
        logo_changed = True

    if qr_changed or logo_changed:
        paired_id = getattr(competition, "paired_competition_id", None)
        if paired_id is not None:
            paired = db.query(Competition).filter(Competition.id == int(paired_id)).first()
            if paired is not None:
                if qr_changed:
                    paired.qr_code_path = competition.qr_code_path
                    paired.qr_code_path_undergraduate = competition.qr_code_path_undergraduate
                    paired.qr_code_path_vocational = competition.qr_code_path_vocational
                if logo_changed:
                    paired.logo_path = competition.logo_path

    # 简介/规则/联系人等文本字段仅改本场，不再同步到关联初赛/决赛

    db.commit()
    for old_path in old_paths_to_delete:
        _delete_stored_qr_file(old_path)
    for old_logo in old_logo_paths_to_delete:
        _delete_stored_logo_file(old_logo)
    db.refresh(competition)
    return competition


@router.delete("/{competition_id}", status_code=status.HTTP_200_OK)
async def delete_competition(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    qr_storage = competition.qr_code_path
    logo_storage = getattr(competition, "logo_path", None)

    db.query(Review).filter(
        Review.submission_id.in_(
            db.query(Submission.id).filter(Submission.competition_id == competition_id)
        )
    ).delete(synchronize_session=False)
    db.query(Submission).filter(Submission.competition_id == competition_id).delete(synchronize_session=False)
    db.query(CompetitionEnrollment).filter(CompetitionEnrollment.competition_id == competition_id).delete(synchronize_session=False)
    db.query(CompetitionPromotion).filter(
        or_(
            CompetitionPromotion.from_competition_id == competition_id,
            CompetitionPromotion.to_competition_id == competition_id,
        )
    ).delete(synchronize_session=False)
    paired_id = getattr(competition, "paired_competition_id", None)
    if paired_id is not None:
        paired = db.query(Competition).filter(Competition.id == int(paired_id)).first()
        if paired is not None:
            paired.paired_competition_id = None
    db.query(TeamMember).filter(
        TeamMember.team_id.in_(
            db.query(Team.id).filter(Team.competition_id == competition_id)
        )
    ).delete(synchronize_session=False)
    db.query(Team).filter(Team.competition_id == competition_id).delete(synchronize_session=False)
    db.delete(competition)

    db.commit()
    _delete_stored_qr_file(qr_storage)
    _delete_stored_logo_file(logo_storage)

    return {"ok": True, "detail": f"Competition {competition_id} and all related data deleted"}


def _peek_promotion_work_track(db: Session, row: CompetitionPromotion) -> Optional[str]:
    """从初赛源队伍/个人报名或决赛队伍解析晋级记录赛道。"""
    if row.source_team_id:
        st = db.query(Team).filter(Team.id == row.source_team_id).first()
        track = _peek_team_work_track(db, row.from_competition_id, st)
        if track:
            return track
    if row.source_student_id:
        enroll = _get_enrollment_by_scope(
            db,
            row.from_competition_id,
            row.source_student_id,
            CompetitionEnrollmentScope.INDIVIDUAL,
        )
        track = _normalize_optional_work_track(
            getattr(enroll, "work_track", None) if enroll else None
        )
        if track:
            return track
    if row.final_team_id:
        ft = db.query(Team).filter(Team.id == row.final_team_id).first()
        track = _peek_team_work_track(db, row.to_competition_id, ft)
        if track:
            return track
    return None


def _promotion_to_schema(
    db: Session,
    row: CompetitionPromotion,
    *,
    users_by_id: Optional[dict[int, AltAuthUserRecord]] = None,
) -> CompetitionPromotionResponse:
    users_by_id = users_by_id or {}
    src_name = None
    final_name = None
    advisor_name = None
    captain_id = None
    captain_name = None
    members_label = None
    source_team = None
    if row.source_team_id:
        source_team = (
            db.query(Team)
            .options(joinedload(Team.members))
            .filter(Team.id == row.source_team_id)
            .first()
        )
        if source_team:
            src_name = source_team.name
            advisor_name = _team_advisor_display_name(source_team, users_by_id)
            captain_id, captain_name = _team_captain_label(source_team, users_by_id)
            members_label = _team_members_label(source_team, users_by_id, exclude_captain=True)
    if row.final_team_id:
        ft = db.query(Team).filter(Team.id == row.final_team_id).first()
        final_name = ft.name if ft else None
    if row.source_student_id and not source_team:
        captain_id = int(row.source_student_id)
        captain_name = _display_user_name(users_by_id.get(captain_id), captain_id)
        members_label = "-"
    return CompetitionPromotionResponse(
        id=row.id,
        from_competition_id=row.from_competition_id,
        to_competition_id=row.to_competition_id,
        source_team_id=row.source_team_id,
        source_student_id=row.source_student_id,
        final_team_id=row.final_team_id,
        source_team_name=src_name,
        final_team_name=final_name,
        advisor_name=advisor_name,
        captain_name=captain_name,
        captain_id=captain_id,
        members=members_label,
        work_track=_peek_promotion_work_track(db, row),
        promoted_by=row.promoted_by,
        created_at=row.created_at,
    )


def _collect_promotion_user_ids(db: Session, rows: List[CompetitionPromotion]) -> set[int]:
    uids: set[int] = set()
    for row in rows:
        if row.source_student_id is not None:
            uids.add(int(row.source_student_id))
        if row.promoted_by is not None:
            uids.add(int(row.promoted_by))
        team_ids = [tid for tid in (row.source_team_id, row.final_team_id) if tid is not None]
        if not team_ids:
            continue
        teams = (
            db.query(Team)
            .options(joinedload(Team.members))
            .filter(Team.id.in_(team_ids))
            .all()
        )
        for t in teams:
            if t.captain_id is not None:
                uids.add(int(t.captain_id))
            if t.created_by_advisor_id is not None:
                uids.add(int(t.created_by_advisor_id))
            for m in t.members or []:
                if m.user_id is not None:
                    uids.add(int(m.user_id))
    return uids


def _assert_team_matches_promotion_track(db: Session, prelim: Competition, team: Team, track: Optional[str]) -> None:
    if not track:
        return
    team_track = _peek_team_work_track(db, prelim.id, team)
    if not team_track:
        raise HTTPException(
            status_code=400,
            detail=f"队伍 {team.id} 未设置赛道，无法按{_work_track_section_label(track)}晋级",
        )
    if team_track != track:
        raise HTTPException(
            status_code=400,
            detail=(
                f"队伍 {team.id} 属于{_work_track_section_label(team_track)}，"
                f"不能在{_work_track_section_label(track)}晋级"
            ),
        )


@router.get(
    "/{competition_id}/promotions/candidates",
    response_model=CompetitionPromotionCandidatesResponse,
)
async def list_promotion_candidates(
    competition_id: int,
    work_track: Optional[str] = Query(None, description="works / software / hardware"),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """初赛：列出可晋级决赛的队伍（须校审通过）。可按 work_track 筛选。"""
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    prelim = _get_competition(db, competition_id)
    final = _paired_final_competition(db, prelim)
    track = _normalize_optional_work_track(work_track)
    promoted_source_ids = {
        int(r[0])
        for r in db.query(CompetitionPromotion.source_team_id)
        .filter(
            CompetitionPromotion.to_competition_id == final.id,
            CompetitionPromotion.source_team_id.isnot(None),
        )
        .all()
        if r[0] is not None
    }
    teams = (
        db.query(Team)
        .filter(Team.competition_id == prelim.id)
        .order_by(Team.created_at.asc())
        .all()
    )
    items: List[CompetitionPromotionCandidateTeam] = []
    for t in teams:
        t_track = _peek_team_work_track(db, prelim.id, t)
        if track and t_track != track:
            continue
        member_ids = [
            int(m.user_id)
            for m in db.query(TeamMember).filter(TeamMember.team_id == t.id).all()
            if m.user_id is not None
            and not bool(m.is_captain)
            and (t.captain_id is None or int(m.user_id) != int(t.captain_id))
        ]
        items.append(
            CompetitionPromotionCandidateTeam(
                team_id=t.id,
                name=t.name,
                division=getattr(t, "division", None),
                work_track=t_track,
                captain_id=t.captain_id,
                member_ids=member_ids,
                status=str(t.status),
                already_promoted=t.id in promoted_source_ids,
            )
        )
    return CompetitionPromotionCandidatesResponse(
        from_competition_id=prelim.id,
        to_competition_id=final.id,
        teams=items,
    )


@router.get(
    "/{competition_id}/promotions",
    response_model=List[CompetitionPromotionResponse],
)
async def list_promotions(
    competition_id: int,
    work_track: Optional[str] = Query(None, description="works / software / hardware"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查看晋级名单：可传初赛或决赛 id。可按 work_track 筛选。"""
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    comp = _get_competition(db, competition_id)
    if _is_preliminary_stage(comp):
        q = db.query(CompetitionPromotion).filter(
            CompetitionPromotion.from_competition_id == comp.id
        )
    elif _is_final_stage(comp):
        q = db.query(CompetitionPromotion).filter(
            CompetitionPromotion.to_competition_id == comp.id
        )
    else:
        raise HTTPException(status_code=400, detail="仅初赛/决赛竞赛有晋级名单")
    track = _normalize_optional_work_track(work_track)
    rows = q.order_by(CompetitionPromotion.created_at.desc()).all()
    users_by_id = _alt_users_by_id(adb, _collect_promotion_user_ids(db, rows))
    out: List[CompetitionPromotionResponse] = []
    for row in rows:
        item = _promotion_to_schema(db, row, users_by_id=users_by_id)
        if track and item.work_track != track:
            continue
        out.append(item)
    return out


@router.post(
    "/{competition_id}/promotions",
    response_model=List[CompetitionPromotionResponse],
    status_code=status.HTTP_201_CREATED,
)
async def create_promotions(
    competition_id: int,
    body: CompetitionPromotionCreate,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """从初赛手动晋级队伍到决赛：自动在决赛建队并报名。可指定 work_track 仅晋级该赛道。"""
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    prelim = _get_competition(db, competition_id)
    final = _paired_final_competition(db, prelim)
    track = _normalize_optional_work_track(getattr(body, "work_track", None))
    team_ids = [int(x) for x in (body.team_ids or []) if x is not None]
    if not team_ids and not (body.student_ids or []):
        raise HTTPException(status_code=400, detail="请选择要晋级的队伍")

    results: List[CompetitionPromotion] = []
    for tid in team_ids:
        team = db.query(Team).filter(Team.id == tid, Team.competition_id == prelim.id).first()
        if not team:
            raise HTTPException(status_code=404, detail=f"队伍 {tid} 不存在或不属于该初赛")
        _assert_team_matches_promotion_track(db, prelim, team, track)
        results.append(
            _promote_prelim_team_to_final(db, prelim, final, team, identity.id)
        )

    # 个人晋级：写入资格并创建个人决赛报名（若允许个人）
    for sid in body.student_ids or []:
        sid_i = int(sid)
        if track:
            stu_track = _normalize_optional_work_track(
                getattr(
                    _get_enrollment_by_scope(
                        db, prelim.id, sid_i, CompetitionEnrollmentScope.INDIVIDUAL
                    ),
                    "work_track",
                    None,
                )
            )
            if stu_track != track:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"学生 {sid_i} 不属于{_work_track_section_label(track)}，无法按该赛道晋级"
                    ),
                )
        existing = (
            db.query(CompetitionPromotion)
            .filter(
                CompetitionPromotion.to_competition_id == final.id,
                CompetitionPromotion.source_student_id == sid_i,
            )
            .first()
        )
        if existing:
            results.append(existing)
            continue
        prelim_enroll = _get_enrollment_by_scope(
            db, prelim.id, sid_i, CompetitionEnrollmentScope.INDIVIDUAL
        )
        if not prelim_enroll or prelim_enroll.status != CompetitionEnrollmentStatus.ENROLLED:
            raise HTTPException(status_code=400, detail=f"学生 {sid_i} 未在初赛个人赛道有效报名")
        if not final.allow_individual:
            raise HTTPException(status_code=400, detail="决赛未开放个人参赛")
        promo = CompetitionPromotion(
            id=allocate_eight_digit_id(db, CompetitionPromotion),
            from_competition_id=prelim.id,
            to_competition_id=final.id,
            source_team_id=None,
            source_student_id=sid_i,
            final_team_id=None,
            promoted_by=identity.id,
        )
        db.add(promo)
        db.flush()
        row = _get_enrollment_by_scope(
            db, final.id, sid_i, CompetitionEnrollmentScope.INDIVIDUAL
        )
        division = str(getattr(prelim_enroll, "division", None) or "default")
        work_track = getattr(prelim_enroll, "work_track", None)
        if row and row.status == CompetitionEnrollmentStatus.WITHDRAWN:
            row.status = CompetitionEnrollmentStatus.ENROLLED
            row.division = division
            row.work_track = work_track
            row.team_id = None
        elif not row:
            db.add(
                CompetitionEnrollment(
                    competition_id=final.id,
                    student_id=sid_i,
                    team_id=None,
                    enrollment_scope=CompetitionEnrollmentScope.INDIVIDUAL,
                    division=division,
                    work_track=work_track,
                    is_captain=False,
                    status=CompetitionEnrollmentStatus.ENROLLED,
                )
            )
        results.append(promo)

    db.commit()
    out: List[CompetitionPromotionResponse] = []
    users_by_id = _alt_users_by_id(adb, _collect_promotion_user_ids(db, results))
    for row in results:
        db.refresh(row)
        out.append(_promotion_to_schema(db, row, users_by_id=users_by_id))
    return out


@router.post(
    "/{competition_id}/promotions/import",
    response_model=CompetitionPromotionImportResult,
    status_code=status.HTTP_200_OK,
)
async def import_promotions_excel(
    competition_id: int,
    file: UploadFile = File(...),
    work_track: Optional[str] = Form(None, description="works / software / hardware"),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    管理员上传 Excel 导入决赛名单。
    表头须含「队伍ID」（或 team_id / 队伍id）；可选「队伍名」用于校验。
    传入 work_track 时仅允许该赛道队伍晋级。
    """
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    prelim = _get_competition(db, competition_id)
    final = _paired_final_competition(db, prelim)
    required_track = _normalize_optional_work_track(work_track)

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="上传文件为空")
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析 Excel：{e}") from e

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel 无内容")

    def _norm_header(v) -> str:
        return str(v or "").strip().lower().replace(" ", "").replace("_", "")

    headers = [_norm_header(c) for c in header_row]
    id_keys = {"队伍id", "teamid", "队伍编号", "id"}
    name_keys = {"队伍名", "队名", "teamname", "name", "队伍名称"}

    col_id = None
    col_name = None
    for i, h in enumerate(headers):
        if col_id is None and h in id_keys:
            col_id = i
        if col_name is None and h in name_keys:
            col_name = i
    if col_id is None:
        raise HTTPException(
            status_code=400,
            detail="Excel 须包含列「队伍ID」（或 team_id）",
        )

    result = CompetitionPromotionImportResult()
    seen_ids: set[int] = set()

    for row_no, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        raw_id = row[col_id] if col_id < len(row) else None
        raw_name = row[col_name] if (col_name is not None and col_name < len(row)) else None
        team_name_hint = str(raw_name).strip() if raw_name is not None else None
        try:
            tid = int(float(str(raw_id).strip()))
        except Exception:
            result.failed += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=None,
                    team_name=team_name_hint,
                    status="error",
                    detail="队伍ID无效",
                )
            )
            continue
        if tid in seen_ids:
            result.skipped += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team_name_hint,
                    status="skipped",
                    detail="文件中重复的队伍ID",
                )
            )
            continue
        seen_ids.add(tid)

        team = db.query(Team).filter(Team.id == tid, Team.competition_id == prelim.id).first()
        if not team:
            result.failed += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team_name_hint,
                    status="error",
                    detail="队伍不存在或不属于该初赛",
                )
            )
            continue
        if required_track:
            team_track = _peek_team_work_track(db, prelim.id, team)
            if not team_track:
                result.failed += 1
                result.items.append(
                    CompetitionPromotionImportItemResult(
                        row=row_no,
                        team_id=tid,
                        team_name=team.name,
                        status="error",
                        detail=f"队伍未设置赛道，无法导入到{_work_track_section_label(required_track)}",
                    )
                )
                continue
            if team_track != required_track:
                result.failed += 1
                result.items.append(
                    CompetitionPromotionImportItemResult(
                        row=row_no,
                        team_id=tid,
                        team_name=team.name,
                        status="error",
                        detail=(
                            f"队伍属于{_work_track_section_label(team_track)}，"
                            f"不能导入到{_work_track_section_label(required_track)}"
                        ),
                    )
                )
                continue
        if team.status != TeamStatus.ACTIVE:
            result.failed += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team.name,
                    status="error",
                    detail=f"队伍状态为 {team.status}，须为 active",
                )
            )
            continue
        if team_name_hint:
            actual = (team.name or "").strip()
            if actual and actual != team_name_hint and team_name_hint not in actual:
                # soft warn but still promote if ID matches
                pass

        existing = (
            db.query(CompetitionPromotion)
            .filter(
                CompetitionPromotion.from_competition_id == prelim.id,
                CompetitionPromotion.source_team_id == tid,
            )
            .first()
        )
        if existing:
            result.skipped += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team.name,
                    status="skipped",
                    detail="已晋级",
                )
            )
            continue

        try:
            _promote_prelim_team_to_final(db, prelim, final, team, identity.id)
            result.imported += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team.name,
                    status="promoted",
                    detail="晋级成功",
                )
            )
        except HTTPException as he:
            result.failed += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team.name,
                    status="error",
                    detail=str(he.detail),
                )
            )
        except Exception as e:
            result.failed += 1
            result.items.append(
                CompetitionPromotionImportItemResult(
                    row=row_no,
                    team_id=tid,
                    team_name=team.name,
                    status="error",
                    detail=str(e),
                )
            )

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"导入提交失败：{e}") from e

    return result


@router.delete("/{competition_id}/promotions/{promotion_id}", status_code=status.HTTP_200_OK)
async def revoke_promotion(
    competition_id: int,
    promotion_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """撤销晋级（决赛开始前）。可传初赛或决赛 id。"""
    require_permission(identity.role, Permission.MANAGE_COMPETITIONS)
    comp = _get_competition(db, competition_id)
    promo = db.query(CompetitionPromotion).filter(CompetitionPromotion.id == promotion_id).first()
    if not promo:
        raise HTTPException(status_code=404, detail="晋级记录不存在")
    if promo.from_competition_id != comp.id and promo.to_competition_id != comp.id:
        raise HTTPException(status_code=400, detail="晋级记录与竞赛不匹配")

    final = _get_competition(db, promo.to_competition_id)
    if final.start_at is not None and ensure_utc(final.start_at) <= utc_now():
        raise HTTPException(status_code=400, detail="决赛已开始，无法撤销晋级")

    final_team_id = promo.final_team_id
    if final_team_id is not None:
        db.query(CompetitionEnrollment).filter(
            CompetitionEnrollment.competition_id == final.id,
            CompetitionEnrollment.team_id == final_team_id,
        ).delete(synchronize_session=False)
        db.query(TeamMember).filter(TeamMember.team_id == final_team_id).delete(
            synchronize_session=False
        )
        db.query(CompetitionQuestionAnswer).filter(
            CompetitionQuestionAnswer.team_id == final_team_id
        ).delete(synchronize_session=False)
        db.query(Team).filter(Team.id == final_team_id).delete(synchronize_session=False)
    elif promo.source_student_id is not None:
        db.query(CompetitionEnrollment).filter(
            CompetitionEnrollment.competition_id == final.id,
            CompetitionEnrollment.student_id == promo.source_student_id,
            CompetitionEnrollment.enrollment_scope == CompetitionEnrollmentScope.INDIVIDUAL,
        ).delete(synchronize_session=False)

    db.delete(promo)
    db.commit()
    return {"ok": True, "detail": f"Promotion {promotion_id} revoked"}


@router.post("/enroll", response_model=CompetitionEnrollmentResponse, status_code=status.HTTP_201_CREATED)
async def enroll_competition(
    enroll: CompetitionEnrollmentCreate,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.ENROLL_COMPETITIONS)

    competition = _get_competition(db, enroll.competition_id)
    _ensure_enrollment_open(competition)
    if competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")

    if _effective_alt_role(identity.role) != "student":
        # 当前版本仅允许 student 报名
        raise HTTPException(status_code=403, detail="Only students can enroll")

    # 允许个人/队伍模式校验
    is_team = enroll.team_id is not None
    if is_team and not competition.allow_team:
        raise HTTPException(status_code=400, detail="Team enrollment not allowed")
    if (not is_team) and not competition.allow_individual:
        raise HTTPException(status_code=400, detail="Individual enrollment not allowed")

    # 决赛：仅晋级名单可报名
    _assert_final_stage_participant(db, competition, identity.id, team_id=enroll.team_id)

    enroll_division = _resolve_enrollment_division(competition, enroll.division)
    enroll_work_track = _resolve_work_track(getattr(enroll, "work_track", None))

    scope = _enrollment_scope_for_team(enroll.team_id)
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        identity.id,
        enroll_work_track,
        enroll_division,
        team_id=enroll.team_id,
        allow_same_team=True,
    )
    existing_row = _get_enrollment_by_work_track(
        db, competition.id, identity.id, enroll_work_track
    )

    team: Optional[Team] = None
    if is_team:
        team = db.query(Team).filter(Team.id == enroll.team_id, Team.competition_id == competition.id).first()
        if not team:
            raise HTTPException(status_code=404, detail="Team not found in this competition")

        team_div = str(getattr(team, "division", None) or "default").lower()
        if team_div != enroll_division:
            raise HTTPException(
                status_code=400,
                detail="Team division does not match enrollment division",
            )

        # 同一队伍：成员表必须存在
        member = db.query(TeamMember).filter(
            TeamMember.team_id == team.id, TeamMember.user_id == identity.id
        ).first()
        if not member:
            raise HTTPException(status_code=403, detail="User is not a team member")

        is_captain = bool(member.is_captain)
        if existing_row and existing_row.status == CompetitionEnrollmentStatus.WITHDRAWN:
            existing_row.team_id = team.id
            existing_row.enrollment_scope = CompetitionEnrollmentScope.TEAM
            existing_row.is_captain = is_captain
            existing_row.status = CompetitionEnrollmentStatus.ENROLLED
            existing_row.division = enroll_division
            existing_row.work_track = enroll_work_track
            existing_row.student_no = enroll.student_no
            existing_row.real_name = enroll.real_name
            existing_row.college = enroll.college
            existing_row.grade = enroll.grade
            existing_row.contact = enroll.contact
            enrollment = existing_row
        else:
            enrollment = CompetitionEnrollment(
                competition_id=competition.id,
                student_id=identity.id,
                team_id=team.id,
                enrollment_scope=CompetitionEnrollmentScope.TEAM,
                division=enroll_division,
                work_track=enroll_work_track,
                is_captain=is_captain,
                status=CompetitionEnrollmentStatus.ENROLLED,
                student_no=enroll.student_no,
                real_name=enroll.real_name,
                college=enroll.college,
                grade=enroll.grade,
                contact=enroll.contact,
            )
            db.add(enrollment)
    else:
        if existing_row and existing_row.status == CompetitionEnrollmentStatus.WITHDRAWN:
            existing_row.team_id = None
            existing_row.enrollment_scope = CompetitionEnrollmentScope.INDIVIDUAL
            existing_row.is_captain = False
            existing_row.status = CompetitionEnrollmentStatus.ENROLLED
            existing_row.division = enroll_division
            existing_row.work_track = enroll_work_track
            existing_row.student_no = enroll.student_no
            existing_row.real_name = enroll.real_name
            existing_row.college = enroll.college
            existing_row.grade = enroll.grade
            existing_row.contact = enroll.contact
            enrollment = existing_row
        else:
            enrollment = CompetitionEnrollment(
                competition_id=competition.id,
                student_id=identity.id,
                team_id=None,
                enrollment_scope=CompetitionEnrollmentScope.INDIVIDUAL,
                division=enroll_division,
                work_track=enroll_work_track,
                is_captain=False,
                status=CompetitionEnrollmentStatus.ENROLLED,
                student_no=enroll.student_no,
                real_name=enroll.real_name,
                college=enroll.college,
                grade=enroll.grade,
                contact=enroll.contact,
            )
            db.add(enrollment)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Enrollment failed: {str(e)}")

    db.refresh(enrollment)
    base = CompetitionEnrollmentResponse.model_validate(enrollment)
    if not is_team:
        seq = _individual_sequence_no(db, competition.id, enrollment)
        return base.model_copy(update={"sequence_no": seq})
    assert team is not None
    seq = _team_sequence_no(db, competition.id, team)
    return base.model_copy(update={"sequence_no": seq})


@router.post("/{competition_id}/withdraw", response_model=CompetitionEnrollmentResponse)
async def withdraw_from_competition(
    competition_id: int,
    track: Optional[str] = Query(
        None,
        description="兼容旧参数：individual（个人）或 team（组队）。多赛道组队报名时请改用 work_track。",
    ),
    work_track: Optional[str] = Query(
        None,
        description="作品赛道：works / software / hardware。按赛道退赛时优先使用。",
    ),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    参赛学生退赛（取消当前竞赛中某一赛道的有效报名）。
    - 指定 work_track：仅撤回该作品赛道报名（及其关联队伍成员身份）
    - 个人赛道（track=individual）：将个人报名置为 withdrawn
    - 组队赛道：从 team_members 移除；队长若队伍内仍有其他成员，须先转让队长；
      若队长为队内唯一成员，则退赛同时解散队伍（team 标记为 disbanded）
    """
    require_permission(identity.role, Permission.ENROLL_COMPETITIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can withdraw")

    competition = _get_competition(db, competition_id)

    active_rows = _list_enrollments_for_student(
        db, competition.id, identity.id, enrolled_only=True
    )
    if not active_rows:
        raise HTTPException(status_code=404, detail="No active enrollment in this competition")

    wt = _normalize_optional_work_track(work_track)
    resolved_scope = (track or "").strip().lower() or None
    if resolved_scope not in (
        None,
        CompetitionEnrollmentScope.INDIVIDUAL,
        CompetitionEnrollmentScope.TEAM,
    ):
        raise HTTPException(
            status_code=400,
            detail="Invalid track; use individual or team",
        )

    enrollment: Optional[CompetitionEnrollment] = None
    if wt:
        enrollment = next(
            (
                r
                for r in active_rows
                if _normalize_optional_work_track(getattr(r, "work_track", None)) == wt
            ),
            None,
        )
        if enrollment is None:
            raise HTTPException(
                status_code=404,
                detail=f"No active enrollment for work_track={wt}",
            )
    elif resolved_scope == CompetitionEnrollmentScope.INDIVIDUAL:
        enrollment = next(
            (
                r
                for r in active_rows
                if r.enrollment_scope == CompetitionEnrollmentScope.INDIVIDUAL
            ),
            None,
        )
        if enrollment is None:
            raise HTTPException(
                status_code=404,
                detail="No active individual enrollment in this competition",
            )
    elif resolved_scope == CompetitionEnrollmentScope.TEAM:
        team_rows = [
            r
            for r in active_rows
            if r.enrollment_scope == CompetitionEnrollmentScope.TEAM
        ]
        if not team_rows:
            raise HTTPException(
                status_code=404,
                detail="No active team enrollment in this competition",
            )
        if len(team_rows) > 1:
            raise HTTPException(
                status_code=400,
                detail="Specify work_track=works|software|hardware when enrolled in multiple work tracks",
            )
        enrollment = team_rows[0]
    else:
        if len(active_rows) > 1:
            raise HTTPException(
                status_code=400,
                detail="Specify work_track=works|software|hardware (or track=individual|team) when enrolled in multiple tracks",
            )
        enrollment = active_rows[0]

    assert enrollment is not None

    if enrollment.enrollment_scope == CompetitionEnrollmentScope.INDIVIDUAL or enrollment.team_id is None:
        enrollment.status = CompetitionEnrollmentStatus.WITHDRAWN
        enrollment.is_captain = False
        db.commit()
        db.refresh(enrollment)
        return enrollment

    team = (
        db.query(Team)
        .filter(Team.id == enrollment.team_id, Team.competition_id == competition.id)
        .first()
    )
    member = db.query(TeamMember).filter(
        TeamMember.team_id == enrollment.team_id,
        TeamMember.user_id == identity.id,
    ).first()

    if team is None:
        enrollment.status = CompetitionEnrollmentStatus.WITHDRAWN
        enrollment.is_captain = False
        if member:
            db.delete(member)
        db.commit()
        db.refresh(enrollment)
        return enrollment

    is_captain = team.captain_id == identity.id or (member is not None and member.is_captain)

    if is_captain and team is not None:
        other_members = (
            db.query(TeamMember)
            .filter(TeamMember.team_id == team.id, TeamMember.user_id != identity.id)
            .count()
        )
        if other_members > 0:
            raise HTTPException(
                status_code=400,
                detail="Captain must transfer captaincy before withdrawing from the competition",
            )
        if member:
            db.delete(member)
        team.status = TeamStatus.DISBANDED
        enrollment.status = CompetitionEnrollmentStatus.WITHDRAWN
        enrollment.is_captain = False
        db.commit()
        db.refresh(enrollment)
        return enrollment

    if member:
        db.delete(member)
    enrollment.status = CompetitionEnrollmentStatus.WITHDRAWN
    enrollment.is_captain = False
    db.commit()
    db.refresh(enrollment)
    return enrollment

@router.get("/{competition_id}/teams/lookup", response_model=TeamResponse)
async def lookup_team_by_name(
    competition_id: int,
    name: str = Query(..., min_length=1, description="队名（精确匹配，忽略大小写）"),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """按队名查找可加入的队伍（pending_school_review / active）。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    target = _strip_team_name(name)
    if not target:
        raise HTTPException(status_code=404, detail="Team not found")
    target_key = target.casefold()
    teams = (
        db.query(Team)
        .filter(
            Team.competition_id == competition_id,
            Team.status.in_(_team_composition_open_statuses()),
        )
        .all()
    )
    for team in teams:
        tn = _strip_team_name(team.name)
        if tn and tn.casefold() == target_key:
            return team
    raise HTTPException(status_code=404, detail="Team not found")


@router.get("/{competition_id}/teams", response_model=List[TeamDetailResponse])
async def list_teams(
    competition_id: int,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查看某竞赛下队伍（含成员列表）；指导老师/教师仅见本人创建的队伍（含待校审）。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)

    role_eff = _effective_alt_role(identity.role)
    q = db.query(Team).filter(Team.competition_id == competition_id)
    if role_eff in {"advisor", "teacher"}:
        teams = (
            q.options(joinedload(Team.members))
            .filter(
                or_(
                    Team.created_by_advisor_id == identity.id,
                    Team.second_advisor_id == identity.id,
                ),
                Team.status.in_(
                    [
                        TeamStatus.ACTIVE,
                        TeamStatus.PENDING_SCHOOL_REVIEW,
                        TeamStatus.REJECTED,
                    ]
                ),
            )
            .order_by(Team.created_at.desc())
            .all()
        )
    else:
        if role_eff == "super_admin":
            teams_q = q.options(joinedload(Team.members)).filter(
                Team.status.in_([TeamStatus.ACTIVE, TeamStatus.PENDING_SCHOOL_REVIEW])
            )
        elif role_eff == "expert" and _is_competition_assigned_expert(db, competition_id, identity):
            allowed = _assigned_team_ids_for_expert(db, competition_id, identity.id)
            if not allowed:
                return []
            teams_q = (
            q.options(joinedload(Team.members))
                .filter(Team.status == TeamStatus.ACTIVE, Team.id.in_(allowed))
            )
        else:
            teams_q = q.options(joinedload(Team.members)).filter(Team.status == TeamStatus.ACTIVE)
        teams = teams_q.order_by(Team.created_at.desc()).all()
    anonymize = _is_expert_anonymized_viewer(db, competition_id, identity)
    return _team_detail_responses(adb, teams, anonymize=anonymize)


@router.get("/{competition_id}/participants/individual", response_model=List[IndividualParticipantItem])
async def list_individual_participants(
    competition_id: int,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    查看某竞赛「个人赛道」全部有效报名者（`team_id` 为空、`enrolled`）。
    `sequence_no` 为本竞赛个人赛道内序号（从 1 起）；`enrollment_id` 为数据库主键。
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    if not _can_view_full_participant_rosters(db, competition_id, identity):
        raise HTTPException(status_code=403, detail="Only super_admin or assigned verified experts may view roster")

    # 专家按队伍指派：个人赛道花名册对其不开放
    if _is_competition_assigned_expert(db, competition_id, identity) and not _can_view_all_competition_submissions(
        db, competition_id, identity
    ):
        return []

    rows = (
        db.query(CompetitionEnrollment)
        .filter(
            CompetitionEnrollment.competition_id == competition_id,
            CompetitionEnrollment.enrollment_scope == CompetitionEnrollmentScope.INDIVIDUAL,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
        )
        .order_by(CompetitionEnrollment.created_at.asc(), CompetitionEnrollment.id.asc())
        .all()
    )
    alt_ids = {r.student_id for r in rows}
    alt_map = _alt_users_by_id(adb, alt_ids)
    out: List[IndividualParticipantItem] = []
    for seq, enr in enumerate(rows, start=1):
        au = alt_map.get(enr.student_id)
        out.append(
            IndividualParticipantItem(
                sequence_no=seq,
                enrollment_id=enr.id,
                student_id=enr.student_id,
                username=(au.username or "") if au else "",
                full_name=au.full_name if au else None,
                student_no=enr.student_no,
                real_name=enr.real_name,
                college=enr.college,
                grade=enr.grade,
                contact=enr.contact,
                status=enr.status,
                created_at=enr.created_at,
            )
        )
    return out


@router.get("/{competition_id}/participants/teams", response_model=List[TeamParticipantDetailResponse])
async def list_team_participants(
    competition_id: int,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    查看某竞赛「组队赛道」全部活跃队伍及成员（含账号名）。
    `sequence_no` 为本竞赛内队伍序号（从 1 起）；队伍的 `id` 仍为全局队伍主键。
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    if not _can_view_full_participant_rosters(db, competition_id, identity):
        raise HTTPException(status_code=403, detail="Only super_admin or assigned verified experts may view roster")

    teams_q = (
        db.query(Team)
        .options(joinedload(Team.members))
        .filter(Team.competition_id == competition_id, Team.status == TeamStatus.ACTIVE)
    )
    if _is_competition_assigned_expert(db, competition_id, identity) and not _can_view_all_competition_submissions(
        db, competition_id, identity
    ):
        allowed = _assigned_team_ids_for_expert(db, competition_id, identity.id)
        if not allowed:
            return []
        teams_q = teams_q.filter(Team.id.in_(allowed))
    teams = teams_q.order_by(Team.created_at.asc(), Team.id.asc()).all()
    all_uids = {m.user_id for t in teams for m in t.members} | {t.captain_id for t in teams}
    users_by_id = _alt_users_by_id(adb, all_uids)
    anonymize = _is_expert_anonymized_viewer(db, competition_id, identity)

    out: List[TeamParticipantDetailResponse] = []
    for seq, team in enumerate(teams, start=1):
        members_out: List[TeamMemberWithUserResponse] = []
        for m in sorted(team.members, key=lambda x: (x.joined_at or utc_now(), x.id)):
            u = users_by_id.get(m.user_id)
            members_out.append(
                TeamMemberWithUserResponse(
                    id=m.id,
                    team_id=m.team_id,
                    user_id=m.user_id,
                    username="" if anonymize else ((u.username if u else "") or ""),
                    full_name=None if anonymize else (u.full_name if u else None),
                    is_captain=m.is_captain,
                    joined_at=m.joined_at,
                )
            )
        out.append(
            TeamParticipantDetailResponse(
                sequence_no=seq,
                id=team.id,
                competition_id=team.competition_id,
                name=team.name,
                captain_id=team.captain_id,
                status=team.status,
                created_at=team.created_at,
                members=members_out,
            )
        )
    return out


@router.post("/teams", response_model=TeamResponse, status_code=status.HTTP_201_CREATED)
async def create_team(
    team_create: TeamCreate,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    competition = _get_competition(db, team_create.competition_id)
    _ensure_enrollment_open(competition)
    if competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")
    if not competition.allow_team:
        raise HTTPException(status_code=400, detail="Team enrollment not allowed")

    _assert_final_stage_open_create_blocked(competition)

    role_eff = _effective_alt_role(identity.role)
    tn = _strip_team_name(team_create.name)
    team_division = _resolve_enrollment_division(competition, team_create.division)
    team_work_track = _resolve_work_track(getattr(team_create, "work_track", None))
    _assert_team_name_unique_in_competition(db, competition.id, tn)

    if role_eff in {"advisor", "teacher"}:
        ordered_ids: List[int] = []
        seen: set = set()
        member_refs = [
            str(x).strip()
            for x in (team_create.initial_members or [])
            if x is not None and str(x).strip()
        ]
        if member_refs:
            for ref in member_refs:
                sid = int(_resolve_student_ref(adb, ref, label="队员").id)
                if sid not in seen:
                    seen.add(sid)
                    ordered_ids.append(sid)
        else:
            for x in team_create.initial_member_ids or []:
                if x not in seen:
                    seen.add(x)
                    ordered_ids.append(x)

        captain_ref = (team_create.captain_student or "").strip()
        if captain_ref:
            captain_id = int(_resolve_student_ref(adb, captain_ref, label="队长").id)
            if ordered_ids and captain_id not in ordered_ids:
                raise HTTPException(status_code=400, detail="队长必须出现在初始队员列表中")
            if captain_id not in ordered_ids:
                ordered_ids = [captain_id] + ordered_ids
        elif team_create.captain_student_id is not None:
            captain_id = team_create.captain_student_id
            if ordered_ids and captain_id not in ordered_ids:
                raise HTTPException(
                    status_code=400, detail="captain_student_id 必须出现在 initial_member_ids 中"
                )
            if captain_id not in ordered_ids:
                ordered_ids = [captain_id] + ordered_ids
        elif ordered_ids:
            captain_id = ordered_ids[0]
        else:
            raise HTTPException(
                status_code=400,
                detail="指导老师建队须指定队长（姓名或用户ID），或提供至少一名初始队员",
            )

        if identity.id in ordered_ids:
            raise HTTPException(status_code=400, detail="指导老师不能以队员身份写入队伍名单")

        for sid in ordered_ids:
            _ensure_alt_principal_is_student(adb, sid)
            _assert_student_can_enroll_work_track(
                db,
                competition.id,
                sid,
                team_work_track,
                team_division,
                team_id=None,
                allow_same_team=False,
            )

        team_school = _assert_student_ids_same_school(adb, ordered_ids)
        # 指导老师/教师代建队：须显式选择本人担任第一或第二指导老师
        my_role = str(getattr(team_create, "creator_advisor_role", None) or "").strip().lower()
        if my_role not in ("first", "second"):
            raise HTTPException(
                status_code=400,
                detail="请选择您担任第一指导老师还是第二指导老师（creator_advisor_role）",
            )
        self_id = int(identity.id)
        self_name = _display_user_name(identity, identity.id)

        other_row = None
        # 另一位老师：仅支持用户名（advisor_username / second_advisor_username）
        if my_role == "first":
            other_row = _resolve_advisor_row_from_refs(
                adb,
                advisor_username=getattr(team_create, "second_advisor_username", None),
                advisor_id=team_create.second_advisor_id,
                advisor_name=team_create.second_advisor_name,
                label="第二指导老师",
            )
            first_advisor_id = self_id
            first_advisor_name = self_name
            second_advisor_id = int(other_row.id) if other_row else None
            second_advisor_name = (
                _display_user_name(other_row, other_row.id) if other_row else None
            )
        else:
            # 担任第二：第一指导老师选填（用户名）
            other_row = _resolve_advisor_row_from_refs(
                adb,
                advisor_username=getattr(team_create, "advisor_username", None),
                advisor_id=team_create.advisor_id,
                advisor_name=team_create.advisor_name,
                label="第一指导老师",
            )
            first_advisor_id = int(other_row.id) if other_row else None
            first_advisor_name = (
                _display_user_name(other_row, other_row.id) if other_row else None
            )
            second_advisor_id = self_id
            second_advisor_name = self_name

        if first_advisor_id is not None and second_advisor_id is not None:
            if int(first_advisor_id) == int(second_advisor_id):
                raise HTTPException(status_code=400, detail="第一与第二指导老师不能为同一人")

        _assert_team_advisors_quota(
            db,
            competition_id=competition.id,
            division=team_division,
            work_track=team_work_track,
            first_advisor_id=first_advisor_id,
            second_advisor_id=second_advisor_id,
        )

        team = Team(
            id=allocate_eight_digit_id(db, Team),
            competition_id=competition.id,
            name=tn,
            captain_id=captain_id,
            created_by_advisor_id=first_advisor_id,
            advisor_name=first_advisor_name,
            second_advisor_id=second_advisor_id,
            second_advisor_name=second_advisor_name,
            school=team_school,
            division=team_division,
            work_track=team_work_track,
            status=TeamStatus.PENDING_SCHOOL_REVIEW,
        )
        db.add(team)
        db.flush()

        # 队长与队员均需同意邀请后才正式入队（不直接写 TeamMember）
        for sid in ordered_ids:
            _create_pending_team_invite(
                db,
                team=team,
                competition=competition,
                invitee_id=sid,
                inviter_id=identity.id,
                as_captain=(sid == captain_id),
            )

    elif role_eff == "student":
        _assert_student_can_enroll_work_track(
            db,
            competition.id,
            identity.id,
            team_work_track,
            team_division,
            team_id=None,
            allow_same_team=False,
        )

        team_school = _resolve_team_school(adb, identity.id)

        def _resolve_student_picked_advisor_by_username(
            *,
            advisor_username: Optional[str],
            label: str,
        ) -> Tuple[Optional[int], Optional[str]]:
            # 学生添加指导老师：仅支持用户名，不再接受姓名或 8 位用户 ID
            raw = (advisor_username or "").strip()
            if not raw:
                return None, None
            adv_row = _resolve_alt_user_by_username(adb, raw, label=label)
            if _effective_alt_role(adv_row.role) not in {"advisor", "teacher"}:
                raise HTTPException(status_code=400, detail=f"「{raw}」须为指导老师账号")
            if int(adv_row.id) == int(identity.id):
                raise HTTPException(status_code=400, detail=f"不能将自己指定为{label}")
            return int(adv_row.id), _display_user_name(adv_row, adv_row.id)

        legacy_advisor_input = (
            team_create.advisor_id is not None
            or bool((team_create.advisor_name or "").strip())
            or team_create.second_advisor_id is not None
            or bool((team_create.second_advisor_name or "").strip())
        )
        username_advisor_input = bool(
            (getattr(team_create, "advisor_username", None) or "").strip()
            or (getattr(team_create, "second_advisor_username", None) or "").strip()
        )
        if legacy_advisor_input and not username_advisor_input:
            raise HTTPException(
                status_code=400,
                detail="学生添加指导老师请填写用户名，不支持姓名或用户 ID",
            )

        primary_id, primary_name = _resolve_student_picked_advisor_by_username(
            advisor_username=getattr(team_create, "advisor_username", None),
            label="第一指导老师",
        )
        secondary_id, secondary_name = _resolve_student_picked_advisor_by_username(
            advisor_username=getattr(team_create, "second_advisor_username", None),
            label="第二指导老师",
        )
        # 学生建队：第一/第二均为选填，栏位固定，不做互换
        created_by_advisor_id, advisor_display_name = primary_id, primary_name
        second_advisor_id, second_advisor_name = secondary_id, secondary_name
        _assert_team_advisors_quota(
            db,
            competition_id=competition.id,
            division=team_division,
            work_track=team_work_track,
            first_advisor_id=created_by_advisor_id,
            second_advisor_id=second_advisor_id,
        )

        team = Team(
            id=allocate_eight_digit_id(db, Team),
            competition_id=competition.id,
            name=tn,
            captain_id=identity.id,
            created_by_advisor_id=created_by_advisor_id,
            advisor_name=advisor_display_name,
            second_advisor_id=second_advisor_id,
            second_advisor_name=second_advisor_name,
            school=team_school,
            division=team_division,
            work_track=team_work_track,
            status=TeamStatus.PENDING_SCHOOL_REVIEW,
        )
        db.add(team)
        db.flush()

        captain_member = TeamMember(team_id=team.id, user_id=identity.id, is_captain=True)
        db.add(captain_member)

        _upsert_team_enrollment(
            db,
            competition=competition,
            student_id=identity.id,
            team=team,
            is_captain=True,
            division=team_division,
            work_track=team_work_track,
        )

        extras = team_create.initial_member_ids or []
        extras = [x for x in extras if x != identity.id]
        if extras:
            _assert_student_ids_same_school(adb, [identity.id] + list(extras))
        for sid in extras:
            _ensure_alt_principal_is_student(adb, sid)
            _assert_student_can_enroll_work_track(
                db,
                competition.id,
                sid,
                team_work_track,
                team_division,
                team_id=None,
                allow_same_team=False,
            )
            _assert_student_same_school_as_team(adb, team, sid, label="学生")
            db.add(TeamMember(team_id=team.id, user_id=sid, is_captain=False))
            _upsert_team_enrollment(
                db,
                competition=competition,
                student_id=sid,
                team=team,
                is_captain=False,
                division=team_division,
                work_track=team_work_track,
            )

    else:
        raise HTTPException(status_code=403, detail="Only student, advisor, or teacher can create teams")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Create team failed: {str(e)}")

    db.refresh(team)
    return team


@router.get("/teams/{team_id}", response_model=TeamDetailResponse)
async def get_team(
    team_id: int,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """队员/队长/建队指导老师查看本队详情（含校审状态 pending_school_review / active / rejected）。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    team = (
        db.query(Team)
        .options(joinedload(Team.members))
        .filter(Team.id == team_id)
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    _get_competition(db, team.competition_id)

    role_eff = _effective_alt_role(identity.role)
    if role_eff == "super_admin":
        return _team_detail_response(adb, team)
    if team.captain_id == identity.id:
        return _team_detail_response(adb, team)
    if _team_advisor_managed(team, identity.id):
        return _team_detail_response(adb, team)
    member = (
        db.query(TeamMember)
        .filter(TeamMember.team_id == team_id, TeamMember.user_id == identity.id)
        .first()
    )
    if member:
        return _team_detail_response(adb, team)
    raise HTTPException(status_code=403, detail="Not a member of this team")


@router.patch("/teams/{team_id}", response_model=TeamResponse)
async def patch_team(
    team_id: int,
    body: TeamPatch,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    body = TeamPatch.model_validate(body)
    require_permission(identity.role, Permission.MANAGE_TEAMS)
    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.status.in_(_team_composition_open_statuses()))
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    if team.captain_id != identity.id and not (
        _effective_alt_role(identity.role) in {"advisor", "teacher"} and _team_advisor_managed(team, identity.id)
    ):
        raise HTTPException(status_code=403, detail="Only captain or advising teacher can rename team")

    if body.name is not None:
        new_name = _strip_team_name(body.name)
        _assert_team_name_unique_in_competition(
            db, team.competition_id, new_name, exclude_team_id=team.id
        )
        team.name = new_name
    db.commit()
    db.refresh(team)
    return team


@router.post("/teams/{team_id}/invite", response_model=TeamInviteResponse, status_code=status.HTTP_201_CREATED)
async def invite_team_member(
    team_id: int,
    body: TeamInviteMember,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """队长或建队指导老师邀请学生入队：仅发出邀请，须对方同意后才正式入队。"""
    body = TeamInviteMember.model_validate(body)
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.status.in_(_team_composition_open_statuses()))
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    competition = team.competition
    _ensure_enrollment_open(competition)
    if competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")
    _assert_final_stage_roster_frozen(competition)
    _assert_team_roster_mutable(db, competition.id, team.id)

    if not _can_manage_team_composition(team, identity):
        raise HTTPException(status_code=403, detail="Only captain or advising teacher may invite")

    student_ref = (body.student or "").strip() if getattr(body, "student", None) else ""
    if student_ref:
        sid = int(_resolve_student_ref(adb, student_ref, label="学生").id)
    elif body.student_id is not None:
        sid = body.student_id
    else:
        raise HTTPException(status_code=400, detail="请填写学生姓名或用户 ID")

    if sid == identity.id and _effective_alt_role(identity.role) == "student":
        raise HTTPException(status_code=400, detail="不能邀请自己")

    _ensure_alt_principal_is_student(adb, sid)

    _assert_student_same_school_as_team(adb, team, sid, label="学生")

    if db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == sid).first():
        raise HTTPException(status_code=400, detail="Already a team member")

    invite_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    invite_div = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    if not invite_track:
        raise HTTPException(status_code=400, detail="队伍未设置赛道，无法邀请")
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        sid,
        invite_track,
        invite_div,
        team_id=team.id,
        allow_same_team=True,
    )

    as_captain = int(sid) == int(team.captain_id) and not db.query(TeamMember).filter(
        TeamMember.team_id == team.id, TeamMember.is_captain.is_(True)
    ).first()

    inv = _create_pending_team_invite(
        db,
        team=team,
        competition=competition,
        invitee_id=sid,
        inviter_id=identity.id,
        as_captain=as_captain,
    )

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Invite failed: {str(e)}")
    db.refresh(inv)
    return _build_team_invite_response(
        inv,
        team=team,
        competition=competition,
        inviter=identity,
    )


@router.get("/team-invites/me", response_model=List[TeamInviteResponse])
async def list_my_team_invites(
    status_filter: Optional[str] = Query("pending", alias="status"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """当前学生查看发给自己的入队邀请。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can view team invites")

    st = (status_filter or "pending").strip().lower()
    if st not in {
        TeamInviteStatus.PENDING,
        TeamInviteStatus.ACCEPTED,
        TeamInviteStatus.REJECTED,
        TeamInviteStatus.CANCELLED,
        "all",
    }:
        raise HTTPException(status_code=400, detail="Invalid invite status")

    q = db.query(TeamInvite).filter(TeamInvite.invitee_id == identity.id)
    if st != "all":
        q = q.filter(TeamInvite.status == st)
    rows = q.order_by(TeamInvite.created_at.desc(), TeamInvite.id.desc()).all()

    # 自愈：队伍已驳回但邀请仍 pending 时自动取消，避免学生端一直卡着
    if st in {TeamInviteStatus.PENDING, "all"}:
        pending_rows = [r for r in rows if r.status == TeamInviteStatus.PENDING]
        if pending_rows:
            team_map = {
                t.id: t
                for t in db.query(Team)
                .filter(Team.id.in_({r.team_id for r in pending_rows}))
                .all()
            }
            now = utc_now()
            healed = False
            for r in pending_rows:
                team = team_map.get(r.team_id)
                if team is None or team.status == TeamStatus.REJECTED:
                    r.status = TeamInviteStatus.CANCELLED
                    r.responded_at = now
                    healed = True
            if healed:
                db.commit()
                rows = [r for r in rows if r.status == st] if st != "all" else rows

    team_ids = {r.team_id for r in rows}
    comp_ids = {r.competition_id for r in rows}
    inviter_ids = {r.inviter_id for r in rows}
    teams = {t.id: t for t in db.query(Team).filter(Team.id.in_(list(team_ids))).all()} if team_ids else {}
    comps = (
        {c.id: c for c in db.query(Competition).filter(Competition.id.in_(list(comp_ids))).all()}
        if comp_ids
        else {}
    )
    inviters = _alt_users_by_id(adb, inviter_ids) if inviter_ids else {}
    return [
        _build_team_invite_response(
            r,
            team=teams.get(r.team_id),
            competition=comps.get(r.competition_id),
            inviter=inviters.get(r.inviter_id),
        )
        for r in rows
    ]


@router.post("/team-invites/{invite_id}/respond", response_model=TeamInviteResponse)
async def respond_team_invite(
    invite_id: int,
    body: TeamInviteAction,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """被邀请学生同意或拒绝入队邀请。"""
    body = TeamInviteAction.model_validate(body)
    require_permission(identity.role, Permission.MANAGE_TEAMS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can respond to team invites")

    inv = db.query(TeamInvite).filter(TeamInvite.id == invite_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invite not found")
    if int(inv.invitee_id) != int(identity.id):
        raise HTTPException(status_code=403, detail="Not your invite")
    if inv.status != TeamInviteStatus.PENDING:
        raise HTTPException(status_code=400, detail="Invite already responded")

    team_any = db.query(Team).filter(Team.id == inv.team_id).first()
    competition = None
    if team_any is not None:
        competition = team_any.competition or _get_competition(db, inv.competition_id)
    elif inv.competition_id:
        competition = _get_competition(db, inv.competition_id)

    if body.action == "reject":
        # 队伍已驳回/解散时仍允许学生关闭邀请，避免列表卡住
        inv.status = TeamInviteStatus.REJECTED
        inv.responded_at = utc_now()
        _handle_team_invite_rejected(db, inv, team_any)
        db.commit()
        db.refresh(inv)
        return _build_team_invite_response(inv, team=team_any, competition=competition)

    team = team_any if team_any is not None and team_any.status in _team_composition_open_statuses() else None
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    competition = team.competition or competition or _get_competition(db, inv.competition_id)

    _ensure_enrollment_open(competition)
    if competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")
    _assert_final_stage_roster_frozen(competition)
    _assert_team_roster_mutable(db, competition.id, team.id)
    _assert_student_same_school_as_team(adb, team, identity.id, label="当前账号")

    if db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == identity.id).first():
        inv.status = TeamInviteStatus.ACCEPTED
        inv.responded_at = utc_now()
        db.commit()
        db.refresh(inv)
        return _build_team_invite_response(inv, team=team, competition=competition)

    if _has_active_enrollment_in_scope(
        db,
        competition.id,
        identity.id,
        CompetitionEnrollmentScope.TEAM,
        work_track=_normalize_optional_work_track(getattr(team, "work_track", None)),
    ):
        raise HTTPException(
            status_code=400,
            detail="Already enrolled in this work track for this competition",
        )

    want_captain = bool(inv.as_captain) or int(team.captain_id) == int(identity.id)
    if want_captain:
        existing_captain = (
            db.query(TeamMember)
            .filter(TeamMember.team_id == team.id, TeamMember.is_captain.is_(True))
            .first()
        )
        if existing_captain and int(existing_captain.user_id) != int(identity.id):
            want_captain = False

    _add_student_to_team(db, competition, team, identity.id, is_captain=want_captain)
    if want_captain:
        team.captain_id = identity.id

    inv.status = TeamInviteStatus.ACCEPTED
    inv.responded_at = utc_now()

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Accept invite failed: {str(e)}")
    db.refresh(inv)
    return _build_team_invite_response(inv, team=team, competition=competition)


@router.delete("/teams/{team_id}/members/{user_id}", status_code=status.HTTP_200_OK)
async def kick_team_member(
    team_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.status.in_(_team_composition_open_statuses()))
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    competition = team.competition
    _assert_team_roster_mutable(db, competition.id, team.id)

    if not _can_manage_team_composition(team, identity):
        raise HTTPException(status_code=403, detail="Only captain or advising teacher may remove members")

    if user_id == team.captain_id:
        raise HTTPException(status_code=400, detail="Cannot kick captain")

    tm = db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == user_id).first()
    if not tm:
        raise HTTPException(status_code=404, detail="Member not found")

    db.delete(tm)
    enrollment = db.query(CompetitionEnrollment).filter(
        CompetitionEnrollment.competition_id == competition.id,
        CompetitionEnrollment.team_id == team.id,
        CompetitionEnrollment.student_id == user_id,
    ).first()
    if enrollment:
        enrollment.status = CompetitionEnrollmentStatus.WITHDRAWN

    db.commit()
    return {"ok": True}


@router.post("/teams/{team_id}/members", response_model=TeamJoinRequestResponse, status_code=status.HTTP_201_CREATED)
async def request_join_team(
    team_id: int,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """学生申请加入队伍（须队长或建队指导老师审核通过后正式入队）。"""
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can request to join teams")

    team = db.query(Team).filter(Team.id == team_id).first()
    if not team or team.status not in _team_composition_open_statuses():
        raise HTTPException(status_code=404, detail="Team not found")

    competition = team.competition
    if not competition or competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")
    _ensure_enrollment_open(competition)
    _assert_final_stage_roster_frozen(competition)

    if db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == identity.id).first():
        raise HTTPException(status_code=400, detail="Already a team member")

    _assert_student_same_school_as_team(adb, team, identity.id, label="当前账号")

    join_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    join_div = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    if not join_track:
        raise HTTPException(status_code=400, detail="队伍未设置赛道，无法申请加入")
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        identity.id,
        join_track,
        join_div,
        team_id=team.id,
        allow_same_team=True,
    )

    pending = (
        db.query(TeamJoinRequest)
        .filter(
            TeamJoinRequest.team_id == team.id,
            TeamJoinRequest.user_id == identity.id,
            TeamJoinRequest.status == TeamJoinRequestStatus.PENDING,
        )
        .first()
    )
    if pending:
        raise HTTPException(status_code=400, detail="您已提交过入队申请，请等待队长审核")

    req = TeamJoinRequest(team_id=team.id, user_id=identity.id, status=TeamJoinRequestStatus.PENDING)
    db.add(req)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Join request failed: {str(e)}")

    db.refresh(req)
    users_by_id = _alt_users_by_id(adb, {identity.id})
    return _build_team_join_request_response(req, users_by_id)


@router.get("/teams/{team_id}/join-requests", response_model=List[TeamJoinRequestResponse])
async def list_team_join_requests(
    team_id: int,
    status_filter: str = Query("pending", alias="status", description="pending | approved | rejected"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """队长或建队指导老师查看入队申请列表。"""
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    team = db.query(Team).filter(Team.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    _get_competition(db, team.competition_id)

    if not _can_manage_team_composition(team, identity):
        raise HTTPException(status_code=403, detail="Only captain or advising teacher may view join requests")

    st = (status_filter or "pending").strip().lower()
    if st not in {
        TeamJoinRequestStatus.PENDING,
        TeamJoinRequestStatus.APPROVED,
        TeamJoinRequestStatus.REJECTED,
    }:
        raise HTTPException(status_code=400, detail="Invalid status filter")

    rows = (
        db.query(TeamJoinRequest)
        .filter(TeamJoinRequest.team_id == team_id, TeamJoinRequest.status == st)
        .order_by(TeamJoinRequest.created_at.asc(), TeamJoinRequest.id.asc())
        .all()
    )
    users_by_id = _alt_users_by_id(adb, {r.user_id for r in rows})
    return [_build_team_join_request_response(r, users_by_id) for r in rows]


@router.post(
    "/teams/{team_id}/join-requests/{request_id}/review",
    response_model=TeamJoinRequestResponse,
)
async def review_team_join_request(
    team_id: int,
    request_id: int,
    body: TeamJoinRequestReview,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """队长或建队指导老师同意/拒绝入队申请。"""
    body = TeamJoinRequestReview.model_validate(body)
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.status.in_(_team_composition_open_statuses()))
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    competition = team.competition
    if not competition or competition.status != "published":
        raise HTTPException(status_code=400, detail="Competition not published")
    _ensure_enrollment_open(competition)

    if not _can_manage_team_composition(team, identity):
        raise HTTPException(status_code=403, detail="Only captain or advising teacher may review join requests")

    req = (
        db.query(TeamJoinRequest)
        .filter(
            TeamJoinRequest.id == request_id,
            TeamJoinRequest.team_id == team_id,
            TeamJoinRequest.status == TeamJoinRequestStatus.PENDING,
        )
        .first()
    )
    if not req:
        raise HTTPException(status_code=404, detail="Join request not found or already reviewed")

    if body.action == "reject":
        req.status = TeamJoinRequestStatus.REJECTED
        req.reviewed_at = utc_now()
        req.reviewed_by_id = identity.id
        db.commit()
        db.refresh(req)
        users_by_id = _alt_users_by_id(adb, {req.user_id})
        return _build_team_join_request_response(req, users_by_id)

    _assert_team_roster_mutable(db, competition.id, team.id)
    _ensure_alt_principal_is_student(adb, req.user_id)
    _assert_student_same_school_as_team(adb, team, req.user_id, label="申请人")
    if db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == req.user_id).first():
        raise HTTPException(status_code=400, detail="Already a team member")
    approve_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    approve_div = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    if not approve_track:
        raise HTTPException(status_code=400, detail="队伍未设置赛道")
    _assert_student_can_enroll_work_track(
        db,
        competition.id,
        req.user_id,
        approve_track,
        approve_div,
        team_id=team.id,
        allow_same_team=True,
    )

    _add_student_to_team(db, competition, team, req.user_id, is_captain=False)
    req.status = TeamJoinRequestStatus.APPROVED
    req.reviewed_at = utc_now()
    req.reviewed_by_id = identity.id

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Approve join request failed: {str(e)}")

    db.refresh(req)
    users_by_id = _alt_users_by_id(adb, {req.user_id})
    return _build_team_join_request_response(req, users_by_id)


@router.post("/teams/{team_id}/transfer-captain", response_model=TeamResponse)
async def transfer_captain(
    team_id: int,
    payload: TeamTransferCaptain,
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can transfer captain")

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.status.in_(_team_composition_open_statuses()))
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    if team.captain_id != identity.id:
        raise HTTPException(status_code=403, detail="Only current captain can transfer")

    _assert_team_roster_mutable(db, team.competition_id, team.id)

    new_captain_ref = (payload.new_captain or "").strip() if getattr(payload, "new_captain", None) else ""
    if new_captain_ref:
        new_captain_id = int(_resolve_student_ref(adb, new_captain_ref, label="新队长").id)
    elif payload.new_captain_id is not None:
        new_captain_id = int(payload.new_captain_id)
    else:
        raise HTTPException(status_code=400, detail="请填写新队长姓名或用户 ID")

    new_captain_member = db.query(TeamMember).filter(
        TeamMember.team_id == team.id,
        TeamMember.user_id == new_captain_id,
    ).first()
    if not new_captain_member:
        raise HTTPException(status_code=404, detail="New captain must be a team member")

    # 一致性：captain_id + team_members.is_captain
    old_captain_id = team.captain_id
    team.captain_id = new_captain_id

    db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.is_captain == True).update(  # noqa: E712
        {"is_captain": False}
    )
    new_captain_member.is_captain = True

    # 同步 enrollment.is_captain（这也是“队长身份”在报名维度上的一致性来源）
    db.query(CompetitionEnrollment).filter(
        CompetitionEnrollment.competition_id == team.competition_id,
        CompetitionEnrollment.team_id == team.id,
        CompetitionEnrollment.student_id == old_captain_id,
    ).update({"is_captain": False})
    db.query(CompetitionEnrollment).filter(
        CompetitionEnrollment.competition_id == team.competition_id,
        CompetitionEnrollment.team_id == team.id,
        CompetitionEnrollment.student_id == new_captain_id,
    ).update({"is_captain": True})

    # 新队长报名记录可能缺 division/work_track（入队时旧逻辑未写入），从队伍同步
    team_division = str(getattr(team, "division", None) or CompetitionDivision.DEFAULT.value).lower()
    team_work_track = _normalize_optional_work_track(getattr(team, "work_track", None))
    new_captain_enroll = (
        db.query(CompetitionEnrollment)
        .filter(
            CompetitionEnrollment.competition_id == team.competition_id,
            CompetitionEnrollment.team_id == team.id,
            CompetitionEnrollment.student_id == new_captain_id,
            CompetitionEnrollment.status == CompetitionEnrollmentStatus.ENROLLED,
        )
        .first()
    )
    if new_captain_enroll:
        if not str(getattr(new_captain_enroll, "division", None) or "").strip():
            new_captain_enroll.division = team_division
        if not _normalize_optional_work_track(getattr(new_captain_enroll, "work_track", None)) and team_work_track:
            new_captain_enroll.work_track = team_work_track

    db.commit()
    db.refresh(team)
    return team


@router.post("/teams/{team_id}/leave", status_code=status.HTTP_200_OK)
async def leave_team(
    team_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    队长退队：强制先转让。
    - 如果当前登录者为队长，则要求其先通过 transfer 接受“等价的队长身份转移”
    - 允许直接退队后成员仍可继续提交（本接口只处理“退队/移除成员”）
    """
    require_permission(identity.role, Permission.MANAGE_TEAMS)

    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can leave team")

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.status.in_(_team_composition_open_statuses()))
        .first()
    )
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    _assert_team_roster_mutable(db, team.competition_id, team.id)

    member = db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == identity.id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    if member.is_captain or team.captain_id == identity.id:
        raise HTTPException(status_code=400, detail="Captain must transfer before leaving")

    db.delete(member)

    # 更新报名状态（同一竞赛同一学生/同一队伍一条报名）
    enrollment = db.query(CompetitionEnrollment).filter(
        CompetitionEnrollment.competition_id == team.competition_id,
        CompetitionEnrollment.team_id == team.id,
        CompetitionEnrollment.student_id == identity.id,
    ).first()
    if enrollment:
        enrollment.status = CompetitionEnrollmentStatus.WITHDRAWN

    db.commit()
    return {"ok": True}


@router.get(
    "/{competition_id}/question-answers",
    response_model=CompetitionQuestionAnswersBoard,
)
async def get_team_question_answers_board(
    competition_id: int,
    team_id: int = Query(..., description="队伍 ID"),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查看某队 5 道题上传槽位（队员本人或超管/指派专家）。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    if not _can_view_team_question_answers(db, competition_id, team_id, identity):
        raise HTTPException(status_code=403, detail="Not allowed to view this team's question answers")
    return _build_question_answers_board(db, competition_id, team_id)


@router.get(
    "/{competition_id}/question-answers/overview",
    response_model=CompetitionQuestionAnswersOverviewResponse,
)
async def list_question_answers_overview(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    管理员/指派专家：按队伍查看 5 题答案上传概览（作品列表展示用）。
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    _assert_competition_uses_question_answers(competition)
    from app.competition_exam_config import get_competition_question_count
    q_count = get_competition_question_count(competition)
    is_admin = _can_view_all_competition_submissions(db, competition_id, identity)
    is_expert = _is_competition_assigned_expert(db, competition_id, identity)
    if not is_admin and not is_expert:
        raise HTTPException(status_code=403, detail="Not allowed to view competition answer overview")

    expert_team_ids: Optional[set[int]] = None
    if is_expert and not is_admin:
        expert_team_ids = _assigned_team_ids_for_expert(db, competition_id, identity.id)
        if not expert_team_ids:
            return CompetitionQuestionAnswersOverviewResponse(
                competition_id=competition_id,
                question_count=q_count,
                items=[],
            )

    teams_q = db.query(Team).filter(
        Team.competition_id == competition_id, Team.status == TeamStatus.ACTIVE
    )
    if expert_team_ids is not None:
        teams_q = teams_q.filter(Team.id.in_(expert_team_ids))
    teams = teams_q.order_by(Team.id.asc()).all()
    answers = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.competition_id == competition_id,
            CompetitionQuestionAnswer.status == CompetitionQuestionAnswerStatus.SUBMITTED,
        )
        .all()
    )
    by_team: dict[int, dict[int, CompetitionQuestionAnswer]] = {}
    for row in answers:
        by_team.setdefault(row.team_id, {})[row.question_no] = row

    grade_rows = (
        db.query(CompetitionTeamQuestionGrade)
        .filter(CompetitionTeamQuestionGrade.competition_id == competition_id)
        .all()
    )
    grade_by_team = {int(g.team_id): g for g in grade_rows}

    items: List[CompetitionQuestionAnswersTeamOverview] = []
    anonymize = _is_expert_anonymized_viewer(db, competition_id, identity)
    for team in teams:
        qmap = by_team.get(team.id, {})
        if not qmap:
            continue
        try:
            team_track = _resolve_team_work_track(db, competition_id, team)
        except HTTPException:
            team_track = str(getattr(team, "work_track", None) or "").strip().lower() or "software"
        if team_track not in ("software", "hardware"):
            continue
        team_q_count = get_competition_question_count(competition, team_track)
        slots: List[CompetitionQuestionAnswerSlot] = []
        uploaded = 0
        for q in range(1, team_q_count + 1):
            row = qmap.get(q)
            slot = _slot_from_row(db, q, row)
            if slot.submitted:
                uploaded += 1
            slots.append(slot)
        grade = grade_by_team.get(int(team.id))
        items.append(
            CompetitionQuestionAnswersTeamOverview(
                team_id=team.id,
                team_name=team.name,
                captain_id=None if anonymize else team.captain_id,
                status=team.status,
                work_track=team_track,
                uploaded_count=uploaded,
                question_count=team_q_count,
                slots=slots,
                graded=grade is not None,
                score_q1=grade.score_q1 if grade else None,
                score_q2=grade.score_q2 if grade else None,
                score_q3=grade.score_q3 if grade else None,
                score_q4=grade.score_q4 if grade else None,
                score_q5=grade.score_q5 if grade else None,
                total_score=grade.total_score if grade else None,
                feedback=grade.feedback if grade else None,
                reviewed_at=grade.reviewed_at if grade else None,
            )
        )

    return CompetitionQuestionAnswersOverviewResponse(
        competition_id=competition_id,
        question_count=q_count,
        items=items,
    )


@router.post(
    "/{competition_id}/questions/{question_no}/answers/upload",
    response_model=CompetitionQuestionAnswerResponse,
    status_code=status.HTTP_201_CREATED,
)
async def upload_question_answer(
    competition_id: int,
    question_no: int,
    team_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    队员向指定题号（1~5）上传答案文件；同队同题再次上传将覆盖。
    """
    require_permission(identity.role, Permission.SUBMIT_SUBMISSIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can upload question answers")

    competition = _get_competition(db, competition_id)
    _assert_competition_uses_question_answers(competition)
    team = _require_active_team_member_for_answers(db, competition, team_id, identity)
    track = _resolve_team_work_track(db, competition.id, team)
    _assert_work_track_allows_question_answers(track)
    qno = _validate_question_no(question_no, competition, track)
    _ensure_competition_allows_submissions(competition)
    _assert_team_answers_not_locked(db, competition.id, team_id)

    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Answer file is required")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Answer file is empty")
    if len(content) > MAX_QUESTION_ANSWER_BYTES:
        raise HTTPException(status_code=400, detail="Answer file exceeds 100MiB limit")

    file_uuid = str(uuid.uuid4())
    ext = os.path.splitext(file.filename)[1]
    stored_name = f"comp_{competition_id}_team_{team_id}_q{qno}_{file_uuid}{ext}"
    file_path = os.path.join(QUESTION_ANSWER_UPLOAD_DIR, stored_name)
    with open(file_path, "wb") as f:
        f.write(content)

    file_record = FileModel(
        filename=file.filename,
        file_type="question_answer",
        file_path=file_path,
        file_size=len(content),
        mime_type=file.content_type,
        sender_id=None,
    )
    db.add(file_record)
    db.flush()

    existing = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.competition_id == competition.id,
            CompetitionQuestionAnswer.team_id == team_id,
            CompetitionQuestionAnswer.question_no == qno,
        )
        .first()
    )
    old_file_path = None
    if existing:
        old = db.query(FileModel).filter(FileModel.id == existing.file_id).first()
        if old and old.file_path:
            old_file_path = old.file_path
        existing.submitter_id = identity.id
        existing.file_id = file_record.id
        existing.uploaded_at = utc_now()
        existing.status = CompetitionQuestionAnswerStatus.DRAFT
        existing.submitted_at = None
        row = existing
    else:
        row = CompetitionQuestionAnswer(
            competition_id=competition.id,
            team_id=team_id,
            question_no=qno,
            submitter_id=identity.id,
            file_id=file_record.id,
            status=CompetitionQuestionAnswerStatus.DRAFT,
            submitted_at=None,
        )
        db.add(row)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except OSError:
            pass
        raise HTTPException(status_code=400, detail=f"Upload question answer failed: {str(e)}")

    if old_file_path and old_file_path != file_path:
        try:
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        except OSError:
            pass

    db.refresh(row)
    return _question_answer_response(db, row)



@router.post(
    "/{competition_id}/question-answers/submit",
    response_model=CompetitionQuestionAnswersSubmitResult,
)
async def submit_team_question_answers(
    competition_id: int,
    team_id: int = Form(...),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    正式上传作品：将本队已选文件的题目答案全部标记为 submitted，
    之后管理员/专家「题目答案列表」才可见。
    """
    require_permission(identity.role, Permission.SUBMIT_SUBMISSIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can submit question answers")

    competition = _get_competition(db, competition_id)
    _assert_competition_uses_question_answers(competition)
    _ensure_competition_allows_submissions(competition)
    team = _require_active_team_member_for_answers(db, competition, team_id, identity)
    _assert_work_track_allows_question_answers(
        _resolve_team_work_track(db, competition.id, team)
    )

    rows = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.competition_id == competition.id,
            CompetitionQuestionAnswer.team_id == team_id,
        )
        .all()
    )
    if not rows:
        raise HTTPException(status_code=400, detail="请先为至少一道题选择并保存答案文件")

    if any(
        (getattr(r, "status", None) or "") == CompetitionQuestionAnswerStatus.SUBMITTED
        for r in rows
    ):
        raise HTTPException(status_code=400, detail="作品已正式提交，无法再次提交")

    now = utc_now()
    submitted_count = 0
    for row in rows:
        row.status = CompetitionQuestionAnswerStatus.SUBMITTED
        row.submitted_at = now
        row.submitter_id = identity.id
        submitted_count += 1

    db.commit()
    board = _build_question_answers_board(db, competition.id, team_id)
    return CompetitionQuestionAnswersSubmitResult(
        competition_id=competition.id,
        team_id=team_id,
        submitted_count=submitted_count,
        slots=board.slots,
    )


@router.get("/{competition_id}/question-answers/export")
async def export_question_answers_zip(
    competition_id: int,
    mode: str = Query(..., description="by_team | by_question"),
    work_track: str = Query(
        ...,
        description="works | software | hardware；按赛道导出，压缩包以赛道名称命名",
    ),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    赛后一键导出答案压缩包（按赛道）：
    - works + by_team：作品赛道队伍压缩包作品（外层「队伍名.zip」）
    - software/hardware + by_team：外层含「队伍名.zip」
    - software/hardware + by_question：外层含「题目名.zip」，内层文件夹为队伍名
    总压缩包文件名：作品赛道.zip / 软件赛道.zip / 硬件赛道.zip
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    competition = _get_competition(db, competition_id)
    is_admin = _can_view_all_competition_submissions(db, competition_id, identity)
    is_expert = _is_competition_assigned_expert(db, competition_id, identity)
    if not is_admin and not is_expert:
        raise HTTPException(status_code=403, detail="Not allowed to export answers")
    _ensure_competition_ended_for_export(competition)

    track = _resolve_work_track(work_track)
    mode_norm = (mode or "").strip().lower()
    if mode_norm not in ("by_team", "by_question"):
        raise HTTPException(status_code=400, detail="mode must be by_team or by_question")

    allowed_team_ids = None
    if is_expert and not is_admin:
        allowed_team_ids = _assigned_team_ids_for_expert(db, competition_id, identity.id) or set()

    if track == "works":
        if mode_norm != "by_team":
            raise HTTPException(
                status_code=400,
                detail="作品赛道仅支持按队伍导出（mode=by_team）",
            )
        buffer = _build_works_submissions_export_zip(
            db, competition, allowed_team_ids=allowed_team_ids
        )
    else:
        _assert_competition_uses_question_answers(competition)
        buffer = _build_answers_export_zip(
            db, competition, mode_norm, work_track=track, allowed_team_ids=allowed_team_ids
        )

    filename = f"{_work_track_section_label(track)}.zip"
    headers = {"Content-Disposition": _content_disposition_attachment(filename)}
    return StreamingResponse(buffer, media_type="application/zip", headers=headers)


@router.get("/{competition_id}/question-answers/{answer_id}/download")
async def download_question_answer_file(
    competition_id: int,
    answer_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    row = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.id == answer_id,
            CompetitionQuestionAnswer.competition_id == competition_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Question answer not found")
    if not _can_view_team_question_answers(db, competition_id, row.team_id, identity):
        raise HTTPException(status_code=403, detail="Not allowed to download this answer")

    file_record = db.query(FileModel).filter(FileModel.id == row.file_id).first()
    if not file_record or not file_record.file_path or not os.path.exists(file_record.file_path):
        raise HTTPException(status_code=404, detail="File missing on server")

    competition = _get_competition(db, competition_id)
    team = (
        db.query(Team)
        .filter(Team.id == row.team_id, Team.competition_id == competition_id)
        .first()
    )
    try:
        work_track = _resolve_team_work_track(db, competition_id, team) if team else None
    except HTTPException:
        work_track = str(getattr(team, "work_track", None) or "").strip().lower() or None
    download_name = _question_answer_download_filename(
        team=team,
        competition=competition,
        question_no=int(row.question_no),
        work_track=work_track,
        original_filename=file_record.filename or os.path.basename(file_record.file_path),
    )

    return FileResponse(
        path=file_record.file_path,
        media_type=file_record.mime_type or "application/octet-stream",
        headers={"Content-Disposition": _content_disposition_attachment(download_name)},
    )


@router.delete("/{competition_id}/question-answers/{answer_id}", status_code=status.HTTP_200_OK)
async def delete_question_answer(
    competition_id: int,
    answer_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """队员删除本队某题答案（竞赛须仍允许提交）。"""
    require_permission(identity.role, Permission.SUBMIT_SUBMISSIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can delete question answers")

    competition = _get_competition(db, competition_id)
    _ensure_competition_allows_submissions(competition)

    row = (
        db.query(CompetitionQuestionAnswer)
        .filter(
            CompetitionQuestionAnswer.id == answer_id,
            CompetitionQuestionAnswer.competition_id == competition_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Question answer not found")

    _require_active_team_member_for_answers(db, competition, row.team_id, identity)
    _assert_team_answers_not_locked(db, competition.id, row.team_id)

    file_record = db.query(FileModel).filter(FileModel.id == row.file_id).first()
    old_path = file_record.file_path if file_record else None
    db.delete(row)
    if file_record:
        db.delete(file_record)
    db.commit()

    if old_path:
        try:
            if os.path.exists(old_path):
                os.remove(old_path)
        except OSError:
            pass

    return {"ok": True}


@router.post("/submissions", response_model=SubmissionResponse, status_code=status.HTTP_201_CREATED)
async def create_submission(
    body: Union[SubmissionCreate, SubmissionCreateWrapped],
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    仅接受 `application/json`。需要上传文件请使用 `POST /competitions/submissions/upload`。
    请求体可为 **扁平** `SubmissionCreate`，或 `{"payload": { ...同上字段... }}`（兼容易与 multipart 字段名混淆的前端）。
    若与本接口混用 `UploadFile`，FastAPI 会强制 multipart 并要求名为 `payload` 的字段，导致纯 JSON 返回 422。
    """
    payload = body.payload if isinstance(body, SubmissionCreateWrapped) else body

    require_permission(identity.role, Permission.SUBMIT_SUBMISSIONS)

    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can submit")

    competition = _get_competition(db, payload.competition_id)
    _assert_competition_uses_zip_submission(competition)
    _ensure_competition_allows_submissions(competition)

    # 校验提交目标：个人 or 队伍
    team_id = payload.team_id
    if team_id is None:
        # 个人提交：student_id 必须是当前用户，且个人报名仍有效
        student_id = identity.id
        _ensure_active_individual_enrollment(db, competition.id, identity.id)
        work_track = _resolve_individual_work_track(db, competition.id, identity.id)
    else:
        team = db.query(Team).filter(Team.id == team_id, Team.competition_id == competition.id).first()
        if not team:
            raise HTTPException(status_code=404, detail="Team not found")
        if team.status != TeamStatus.ACTIVE:
            raise HTTPException(status_code=400, detail="Team must be approved by school admin before submitting")

        # 队伍提交：须为队长提交
        member = db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == identity.id).first()
        if not member:
            raise HTTPException(status_code=403, detail="User is not a team member")
        if team.captain_id != identity.id:
            raise HTTPException(status_code=403, detail="Only team captain may submit for the team")
        student_id = identity.id
        work_track = _resolve_team_work_track(db, competition.id, team)
    _assert_work_track_allows_zip(work_track)

    if not payload.file_id and not payload.content_text:
        raise HTTPException(status_code=400, detail="Provide file_id or content_text")

    file_id = payload.file_id
    content_text = payload.content_text

    _purge_prior_zip_submissions(
        db,
        competition.id,
        team_id=team_id,
        student_id=student_id if team_id is None else None,
    )

    submission = Submission(
        competition_id=competition.id,
        team_id=payload.team_id,
        student_id=student_id,
        submitter_id=identity.id,
        title=payload.title,
        description=payload.description,
        file_id=file_id,
        content_text=content_text,
        status=SubmissionStatus.SUBMITTED,
    )
    db.add(submission)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Create submission failed: {str(e)}")

    db.refresh(submission)
    return submission


@router.post("/submissions/upload", response_model=SubmissionResponse, status_code=status.HTTP_201_CREATED)
async def create_submission_upload(
    competition_id: int = Form(...),
    team_id: Optional[int] = Form(None),
    title: str = Form(...),
    description: Optional[str] = Form(None),
    content_text: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    multipart/form-data 方式提交作品（支持文件上传）。
    """
    require_permission(identity.role, Permission.SUBMIT_SUBMISSIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can submit")

    competition = _get_competition(db, competition_id)
    _assert_competition_uses_zip_submission(competition)
    _ensure_competition_allows_submissions(competition)

    # 校验提交目标：个人 or 队伍
    if team_id is None:
        student_id = identity.id
        _ensure_active_individual_enrollment(db, competition.id, identity.id)
        work_track = _resolve_individual_work_track(db, competition.id, identity.id)
    else:
        team = db.query(Team).filter(Team.id == team_id, Team.competition_id == competition.id).first()
        if not team:
            raise HTTPException(status_code=404, detail="Team not found")
        if team.status != TeamStatus.ACTIVE:
            raise HTTPException(status_code=400, detail="Team must be approved by school admin before submitting")
        member = db.query(TeamMember).filter(TeamMember.team_id == team.id, TeamMember.user_id == identity.id).first()
        if not member:
            raise HTTPException(status_code=403, detail="User is not a team member")
        if team.captain_id != identity.id:
            raise HTTPException(status_code=403, detail="Only team captain may submit for the team")
        student_id = identity.id
        work_track = _resolve_team_work_track(db, competition.id, team)
    _assert_work_track_allows_zip(work_track)

    if not (file and file.filename):
        raise HTTPException(status_code=400, detail="请上传作品压缩包（.zip）")

    original_name = os.path.basename(str(file.filename).strip())
    ext = os.path.splitext(original_name)[1].lower()
    if ext != ".zip":
        raise HTTPException(status_code=400, detail="作品赛道仅支持 .zip 压缩包")

    file_id = None
    if file is not None and file.filename:
        file_uuid = str(uuid.uuid4())
        filename = f"{file_uuid}{ext}"
        file_path = os.path.join(SUBMISSION_UPLOAD_DIR, filename)
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
        file_record = FileModel(
            filename=original_name,
            file_type="submission",
            file_path=file_path,
            file_size=len(content),
            mime_type=file.content_type,
            sender_id=None,
        )
        db.add(file_record)
        db.flush()
        file_id = file_record.id

    _purge_prior_zip_submissions(
        db,
        competition.id,
        team_id=team_id,
        student_id=student_id if team_id is None else None,
    )

    submission = Submission(
        competition_id=competition.id,
        team_id=team_id,
        student_id=student_id,
        submitter_id=identity.id,
        title=title,
        description=description,
        file_id=file_id,
        content_text=content_text,
        status=SubmissionStatus.SUBMITTED,
    )
    db.add(submission)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Create submission failed: {str(e)}")

    db.refresh(submission)
    return submission


@router.get("/{competition_id}/submissions", response_model=List[SubmissionResponse])
async def list_submissions(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)

    if _can_view_all_competition_submissions(db, competition_id, identity):
        rows = (
            db.query(Submission)
            .filter(Submission.competition_id == competition_id)
            .order_by(Submission.submitted_at.desc())
            .all()
        )
        return _submissions_to_responses(db, _keep_latest_submissions_only(rows))

    if _is_competition_assigned_expert(db, competition_id, identity):
        team_ids = _assigned_team_ids_for_expert(db, competition_id, identity.id)
        if not team_ids:
            return []
        rows = (
            db.query(Submission)
            .filter(
                Submission.competition_id == competition_id,
                Submission.team_id.in_(team_ids),
            )
            .order_by(Submission.submitted_at.desc())
            .all()
        )
        return _submissions_to_responses(db, _keep_latest_submissions_only(rows))

    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(
            status_code=403,
            detail="Only super_admin, assigned experts, or enrolled students may list submissions here",
        )

    # student：仅看自己的个人提交 + 自己所在队伍的提交
    team_ids = [tm.team_id for tm in db.query(TeamMember).join(Team).filter(
        TeamMember.user_id == identity.id,
        Team.competition_id == competition_id,
    ).all()]

    q = db.query(Submission).filter(Submission.competition_id == competition_id).filter(
        (Submission.student_id == identity.id) | (Submission.team_id.in_(team_ids) if team_ids else False)  # noqa: E712
    )
    rows = q.order_by(Submission.submitted_at.desc()).all()
    return _submissions_to_responses(db, _keep_latest_submissions_only(rows))


@router.get("/submissions/{submission_id}", response_model=SubmissionResponse)
async def get_submission(
    submission_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _ensure_submission_access(db, submission, identity)
    return _submission_to_response(
        submission,
        _team_names_by_ids(db, {int(submission.team_id)} if submission.team_id is not None else set()),
    )


@router.get("/submissions/{submission_id}/download")
async def download_submission_file(
    submission_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _ensure_submission_access(db, submission, identity)

    if not submission.file_id:
        raise HTTPException(status_code=404, detail="No file attached")

    file_record = db.query(FileModel).filter(FileModel.id == submission.file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found")
    if not file_record.file_path or not os.path.exists(file_record.file_path):
        raise HTTPException(status_code=404, detail="File missing on server")

    download_name = _submission_download_filename(db, submission, file_record)
    return FileResponse(
        path=file_record.file_path,
        filename=download_name,
        media_type=file_record.mime_type or "application/octet-stream",
    )


def _require_expert_reviewer_for_submission(
    db: Session, submission: Submission, identity: AltAuthUserRecord
) -> None:
    if _effective_alt_role(identity.role) != "expert":
        raise HTTPException(status_code=403, detail="Only competition experts may grade submissions")
    require_permission(identity.role, Permission.REVIEW_SUBMISSIONS)
    if not _expert_gate_ok(identity):
        raise HTTPException(status_code=403, detail="Expert account must be verified by admin before grading")
    if submission.team_id is None:
        raise HTTPException(status_code=403, detail="Only team submissions can be graded by assigned experts")
    if not _is_assigned_expert_for_team(
        db, submission.competition_id, submission.team_id, identity.id
    ):
        raise HTTPException(status_code=403, detail="You are not assigned to this team")


def _require_expert_reviewer_for_team(
    db: Session, competition_id: int, team_id: int, identity: AltAuthUserRecord
) -> Team:
    team = db.query(Team).filter(Team.id == team_id, Team.competition_id == competition_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    role = _effective_alt_role(identity.role)
    if role == "super_admin":
        return team
    if role != "expert":
        raise HTTPException(status_code=403, detail="Only competition experts or super_admin may grade submissions")
    require_permission(identity.role, Permission.REVIEW_SUBMISSIONS)
    if not _expert_gate_ok(identity):
        raise HTTPException(status_code=403, detail="Expert account must be verified by admin before grading")
    if not _is_assigned_expert_for_team(db, competition_id, team_id, identity.id):
        raise HTTPException(status_code=403, detail="You are not assigned to this team")
    return team


def _team_question_grade_configured_count(
    competition: Optional[Competition] = None,
    work_track: Optional[str] = None,
) -> int:
    from app.competition_exam_config import get_competition_question_count

    if competition is None:
        return COMPETITION_QUESTION_COUNT
    n = get_competition_question_count(competition, work_track)
    return max(1, min(COMPETITION_QUESTION_COUNT, int(n)))


def _team_question_grade_response(row: CompetitionTeamQuestionGrade) -> TeamQuestionGradeResponse:
    return TeamQuestionGradeResponse.model_validate(row)


def _apply_team_question_grade_body(
    row: CompetitionTeamQuestionGrade,
    body: TeamQuestionGradeRequest,
    reviewer_id: int,
    competition: Optional[Competition] = None,
    work_track: Optional[str] = None,
) -> None:
    from app.competition_exam_config import validate_grade_against_config

    scores = {
        1: float(body.score_q1),
        2: float(body.score_q2),
        3: float(body.score_q3),
        4: float(body.score_q4),
        5: float(body.score_q5),
    }
    count = _team_question_grade_configured_count(competition, work_track)
    for i in range(count + 1, COMPETITION_QUESTION_COUNT + 1):
        scores[i] = 0.0
    total = float(sum(scores[i] for i in range(1, count + 1)))
    if competition is not None:
        validate_grade_against_config(competition, scores, total, work_track=work_track)
    row.score_q1 = scores[1]
    row.score_q2 = scores[2]
    row.score_q3 = scores[3]
    row.score_q4 = scores[4]
    row.score_q5 = scores[5]
    row.total_score = total
    row.feedback = body.feedback
    row.reviewer_id = reviewer_id
    row.reviewed_at = utc_now()


def _sync_works_submission_review_from_team_grade(
    db: Session,
    competition: Competition,
    team: Team,
    grade_row: CompetitionTeamQuestionGrade,
    reviewer_id: int,
) -> None:
    """作品赛道按题评分后，同步压缩包作品上的 Review。"""
    try:
        track = _resolve_team_work_track(db, competition.id, team)
    except HTTPException:
        track = str(getattr(team, "work_track", None) or "").strip().lower() or ""
    if track != "works":
        return
    submission = (
        db.query(Submission)
        .options(joinedload(Submission.review))
        .filter(
            Submission.competition_id == competition.id,
            Submission.team_id == team.id,
        )
        .order_by(Submission.submitted_at.desc(), Submission.id.desc())
        .first()
    )
    if not submission:
        return
    now = utc_now()
    review = submission.review
    if review is None:
        db.add(
            Review(
                submission_id=submission.id,
                reviewer_id=reviewer_id,
                status=SubmissionStatus.APPROVED,
                score=float(grade_row.total_score),
                feedback=grade_row.feedback,
                reviewed_at=now,
            )
        )
    else:
        review.score = float(grade_row.total_score)
        review.feedback = grade_row.feedback
        review.reviewed_at = now
        review.reviewer_id = reviewer_id
    submission.status = SubmissionStatus.APPROVED


@router.get(
    "/{competition_id}/teams/{team_id}/question-grades",
    response_model=TeamQuestionGradeResponse,
)
async def get_team_question_grade(
    competition_id: int,
    team_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查询某队五题评分；未评分返回 404。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    if not _can_view_competition_score_insights(db, competition_id, identity):
        if _effective_alt_role(identity.role) == "student":
            member = (
                db.query(TeamMember)
                .filter(TeamMember.team_id == team_id, TeamMember.user_id == identity.id)
                .first()
            )
            if not member:
                raise HTTPException(status_code=403, detail="Not allowed to view this team's grades")
        else:
            raise HTTPException(status_code=403, detail="Not allowed to view grades for this competition")
    row = (
        db.query(CompetitionTeamQuestionGrade)
        .filter(
            CompetitionTeamQuestionGrade.competition_id == competition_id,
            CompetitionTeamQuestionGrade.team_id == team_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Grade not found")
    return _team_question_grade_response(row)


@router.put(
    "/{competition_id}/teams/{team_id}/question-grades",
    response_model=TeamQuestionGradeResponse,
)
async def put_team_question_grade(
    competition_id: int,
    team_id: int,
    body: TeamQuestionGradeRequest,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """首次按题评分；已评分请使用 PATCH。总分由五题相加。"""
    competition = _get_competition(db, competition_id)
    team = _require_expert_reviewer_for_team(db, competition_id, team_id, identity)
    try:
        track = _resolve_team_work_track(db, competition_id, team)
    except HTTPException:
        track = str(getattr(team, "work_track", None) or "").strip().lower() or None
    existing = (
        db.query(CompetitionTeamQuestionGrade)
        .filter(
            CompetitionTeamQuestionGrade.competition_id == competition_id,
            CompetitionTeamQuestionGrade.team_id == team_id,
        )
        .first()
    )
    if existing is not None:
        raise HTTPException(status_code=400, detail="Team already graded")
    row = CompetitionTeamQuestionGrade(
        competition_id=competition_id,
        team_id=team_id,
        reviewer_id=identity.id,
        created_at=utc_now(),
    )
    _apply_team_question_grade_body(row, body, identity.id, competition, track)
    db.add(row)
    _sync_works_submission_review_from_team_grade(db, competition, team, row, identity.id)
    db.commit()
    db.refresh(row)
    return _team_question_grade_response(row)


@router.patch(
    "/{competition_id}/teams/{team_id}/question-grades",
    response_model=TeamQuestionGradeResponse,
)
async def patch_team_question_grade(
    competition_id: int,
    team_id: int,
    body: TeamQuestionGradeRequest,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """修改某队五题评分；未评分请使用 PUT。"""
    competition = _get_competition(db, competition_id)
    team = _require_expert_reviewer_for_team(db, competition_id, team_id, identity)
    try:
        track = _resolve_team_work_track(db, competition_id, team)
    except HTTPException:
        track = str(getattr(team, "work_track", None) or "").strip().lower() or None
    row = (
        db.query(CompetitionTeamQuestionGrade)
        .filter(
            CompetitionTeamQuestionGrade.competition_id == competition_id,
            CompetitionTeamQuestionGrade.team_id == team_id,
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=400, detail="Team not graded yet")
    _apply_team_question_grade_body(row, body, identity.id, competition, track)
    _sync_works_submission_review_from_team_grade(db, competition, team, row, identity.id)
    db.commit()
    db.refresh(row)
    return _team_question_grade_response(row)


@router.get("/submissions/{submission_id}/review-grade", response_model=ReviewResponse)
async def get_review_grade(
    submission_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """查询作品评分记录（与 PUT/PATCH 响应体一致）；未评分返回 404。"""
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    submission = (
        db.query(Submission)
        .options(joinedload(Submission.review))
        .filter(Submission.id == submission_id)
        .first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _ensure_submission_access(db, submission, identity)
    if submission.review is None:
        raise HTTPException(status_code=404, detail="Review not found")
    return submission.review


@router.put("/submissions/{submission_id}/review-grade", response_model=ReviewResponse)
async def review_submission(
    submission_id: int,
    grade: ReviewGrade,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """首次评分/审核；已评分请使用 PATCH 同路径修改。"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _require_expert_reviewer_for_submission(db, submission, identity)

    if submission.review is not None:
        raise HTTPException(status_code=400, detail="Submission already reviewed")

    review = Review(
        submission_id=submission.id,
        reviewer_id=identity.id,
        status=SubmissionStatus.APPROVED,
        score=grade.score,
        feedback=grade.feedback,
        reviewed_at=utc_now(),
    )
    db.add(review)
    submission.status = SubmissionStatus.APPROVED
    db.commit()
    db.refresh(review)
    return review


@router.patch("/submissions/{submission_id}/review-grade", response_model=ReviewResponse)
async def update_review_grade(
    submission_id: int,
    grade: ReviewGrade,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """修改已有评分/反馈；未评分作品请使用 PUT 首次评分。"""
    submission = (
        db.query(Submission)
        .options(joinedload(Submission.review))
        .filter(Submission.id == submission_id)
        .first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _require_expert_reviewer_for_submission(db, submission, identity)

    review = submission.review
    if review is None:
        raise HTTPException(status_code=400, detail="Submission not reviewed yet")

    review.score = grade.score
    review.feedback = grade.feedback
    review.reviewed_at = utc_now()
    review.reviewer_id = identity.id
    if submission.status != SubmissionStatus.APPROVED:
        submission.status = SubmissionStatus.APPROVED

    db.commit()
    db.refresh(review)
    return review


@router.get("/{competition_id}/scores/summary", response_model=CompetitionScoreSummaryResponse)
async def score_summary(
    competition_id: int,
    division: Optional[str] = Query(None, description="default / undergraduate / vocational"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    竞赛评分汇总（按队伍列出：队名、学校、指导老师、队员、五题分、总分）。
    - super_admin：可查看任意竞赛汇总，并可在前端改分
    - verified expert：仅可查看被指派的竞赛汇总（只读）
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    if not _can_view_competition_score_insights(db, competition_id, identity):
        raise HTTPException(status_code=403, detail="Not allowed to view scores for this competition")

    teams = _filter_teams_by_division_query(
        _load_active_teams_with_members(db, competition_id),
        division,
    )
    all_uids: set[int] = set()
    for t in teams:
        all_uids.add(t.captain_id)
        if t.created_by_advisor_id:
            all_uids.add(t.created_by_advisor_id)
        for m in t.members or []:
            all_uids.add(m.user_id)
    users_by_id = _alt_users_by_id(adb, all_uids)
    grade_rows = (
        db.query(CompetitionTeamQuestionGrade)
        .filter(CompetitionTeamQuestionGrade.competition_id == competition_id)
        .all()
    )
    grades_by_team = {int(g.team_id): g for g in grade_rows}
    items = _build_competition_score_team_items(
        teams=teams,
        users_by_id=users_by_id,
        grades_by_team=grades_by_team,
    )

    graded_scores = [float(it.total_score) for it in items if it.graded and it.total_score is not None]
    submissions_total = (
        db.query(func.count(func.distinct(CompetitionQuestionAnswer.team_id)))
        .filter(
            CompetitionQuestionAnswer.competition_id == competition_id,
            CompetitionQuestionAnswer.status == CompetitionQuestionAnswerStatus.SUBMITTED,
        )
        .scalar()
        or 0
    )
    reviewed_total = len(graded_scores)
    avg_score = (sum(graded_scores) / reviewed_total) if reviewed_total else None
    max_score = max(graded_scores) if graded_scores else None
    min_score = min(graded_scores) if graded_scores else None

    return CompetitionScoreSummaryResponse(
        competition_id=competition_id,
        division=_parse_score_division_param(division),
        items=items,
        submissions_total=int(submissions_total),
        reviewed_total=int(reviewed_total),
        avg_score=avg_score,
        max_score=max_score,
        min_score=min_score,
    )


@router.get("/{competition_id}/scores/rankings", response_model=CompetitionScoreRankingResponse)
async def score_rankings(
    competition_id: int,
    limit: int = 50,
    division: Optional[str] = Query(None, description="default / undergraduate / vocational"),
    db: Session = Depends(get_db),
    adb: Session = Depends(get_alt_auth_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    排行榜：按队伍总分排序，含队名、学校、指导老师、队员与五题分。
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    _get_competition(db, competition_id)
    if not _can_view_competition_score_insights(db, competition_id, identity):
        raise HTTPException(status_code=403, detail="Not allowed to view rankings for this competition")

    teams = _filter_teams_by_division_query(
        _load_active_teams_with_members(db, competition_id),
        division,
    )
    all_uids: set[int] = set()
    for t in teams:
        all_uids.add(t.captain_id)
        if t.created_by_advisor_id:
            all_uids.add(t.created_by_advisor_id)
        for m in t.members or []:
            all_uids.add(m.user_id)
    users_by_id = _alt_users_by_id(adb, all_uids)
    grade_rows = (
        db.query(CompetitionTeamQuestionGrade)
        .filter(CompetitionTeamQuestionGrade.competition_id == competition_id)
        .all()
    )
    grades_by_team = {int(g.team_id): g for g in grade_rows}
    board = [
        it
        for it in _build_competition_score_team_items(
            teams=teams,
            users_by_id=users_by_id,
            grades_by_team=grades_by_team,
        )
        if it.graded and it.total_score is not None
    ]
    board.sort(key=lambda x: (-float(x.total_score), int(x.team_id)))

    items: List[CompetitionScoreRankingItem] = []
    rank_val = 1
    for i, row in enumerate(board):
        total = float(row.total_score)
        if i > 0 and total < float(board[i - 1].total_score):
            rank_val = i + 1
        items.append(
            CompetitionScoreRankingItem(
                rank=rank_val,
                team_id=row.team_id,
                student_id=None,
                team_name=row.team_name,
                school=row.school,
                advisor_name=row.advisor_name,
                captain_name=row.captain_name,
                captain_id=row.captain_id,
                members=row.members,
                best_score=total,
                reviewed_submissions=1,
                score_q1=row.score_q1,
                score_q2=row.score_q2,
                score_q3=row.score_q3,
                score_q4=row.score_q4,
                score_q5=row.score_q5,
            )
        )

    return CompetitionScoreRankingResponse(
        competition_id=competition_id,
        division=_parse_score_division_param(division),
        items=items[: max(1, min(int(limit or 50), 500))],
    )


@router.get("/{competition_id}/scores/me", response_model=MyCompetitionScoresResponse)
async def my_scores(
    competition_id: int,
    db: Session = Depends(get_db),
    identity: AltAuthUserRecord = Depends(get_current_alt_identity),
):
    """
    学生查看自己在某竞赛的提交与成绩。
    """
    require_permission(identity.role, Permission.VIEW_COMPETITIONS)
    if _effective_alt_role(identity.role) != "student":
        raise HTTPException(status_code=403, detail="Only students can view my scores")

    _get_competition(db, competition_id)
    # 自己个人提交 + 自己所在队伍提交
    team_ids = [tm.team_id for tm in db.query(TeamMember).join(Team).filter(
        TeamMember.user_id == identity.id,
        Team.competition_id == competition_id,
    ).all()]

    q = (
        db.query(Submission)
        .options(joinedload(Submission.review))
        .filter(Submission.competition_id == competition_id)
        .filter(
            (Submission.student_id == identity.id)
            | (Submission.team_id.in_(team_ids) if team_ids else False)  # noqa: E712
        )
    )
    submissions = _keep_latest_submissions_only(q.order_by(Submission.submitted_at.desc()).all())
    team_names = _team_names_by_ids(
        db, {int(s.team_id) for s in submissions if s.team_id is not None}
    )
    items: List[SubmissionForStudentScoreResponse] = []
    for s in submissions:
        base = _submission_to_response(s, team_names)
        rev = s.review
        items.append(
            SubmissionForStudentScoreResponse(
                **base.model_dump(),
                score=rev.score if rev else None,
                feedback=rev.feedback if rev else None,
                reviewed_at=rev.reviewed_at if rev else None,
            )
        )
    team_grades: List[TeamQuestionGradeResponse] = []
    if team_ids:
        grade_rows = (
            db.query(CompetitionTeamQuestionGrade)
            .filter(
                CompetitionTeamQuestionGrade.competition_id == competition_id,
                CompetitionTeamQuestionGrade.team_id.in_(team_ids),
            )
            .all()
        )
        team_grades = [_team_question_grade_response(g) for g in grade_rows]
    return MyCompetitionScoresResponse(
        competition_id=competition_id,
        submissions=items,
        team_grades=team_grades,
    )

